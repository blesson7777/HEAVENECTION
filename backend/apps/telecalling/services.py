import calendar
import csv
import io
import logging
import re
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_FLOOR, ROUND_HALF_UP
from threading import Thread
from urllib.parse import urlencode

from django.conf import settings
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.paginator import EmptyPage, Paginator
from django.core.mail import EmailMultiAlternatives
from django.db import IntegrityError, close_old_connections, transaction
from django.db.models import Case, Count, F, IntegerField, Max, Q, Sum, Value, When
from django.template.loader import render_to_string
from django.db.models.functions import Coalesce, TruncDate
from django.urls import reverse
from django.utils import timezone

from backend.apps.telecalling.models import (
    AppNotification,
    AppRelease,
    Call,
    CompanyProfile,
    InterestedLeadDetail,
    Lead,
    ReferralSubmission,
    ReferralReward,
    Salary,
    SalaryPaymentTransaction,
    Session,
    Staff,
    StaffAction,
    TrainingCompletion,
    TrainingLesson,
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - dependency installed in runtime
    Workbook = None
    load_workbook = None


ONLINE_WINDOW_SECONDS = 90
SHORT_CALL_SECONDS = 5
BACKGROUND_TIMEOUT_SECONDS = 5 * 60
IDLE_WARNING_AFTER_SECONDS = 5 * 60
IDLE_WARNING_GRACE_SECONDS = 5 * 60
IDLE_OFFLINE_AFTER_SECONDS = IDLE_WARNING_AFTER_SECONDS + IDLE_WARNING_GRACE_SECONDS
LIVE_CALL_STALE_SECONDS = 3 * 60 * 60
VERIFIED_CALL_ACTIVITY_TIMEOUT_SECONDS = 90
VERIFIED_CALL_TIME_SKEW_SECONDS = 2 * 60
CALL_ACTIVITY_IDLE_BREAK_SECONDS = 60
CONNECTED_CALL_COOLDOWN_SECONDS = 90
CALL_ACTIVITY_DIAL_LOOKBACK_SECONDS = 60
MIN_REAL_CALLS_PER_ATTEMPT_BLOCK = 10
VERIFIED_CALL_SOURCES = {
    "call_log",
    "call_log_short_resolution",
    "call_log_short_recall",
}
QUALITY_SCORE_LOOKBACK_DAYS = 30
MISSED_CALLBACK_AFTER_HOURS = 24
PENDING_STATUS_BLOCK_HOURS = 12
TWOPLACES = Decimal("0.01")
DEFAULT_LEAD_QUEUE_LIMIT = 1
CALLBACK_NOON_HOURS = range(12, 16)
logger = logging.getLogger(__name__)
CALLBACK_EVENING_HOURS = range(16, 20)
CALLBACK_NIGHT_HOURS = range(20, 24)
FOLLOWUP_NO_RESPONSE_LIMIT = 3
FOLLOWUP_STALE_EXPIRY_DAYS = 14
FOLLOWUP_STAFF_WARNING_DAYS = 7
FOLLOWUP_UNCALLED_ALERT_HOURS = 24
FOLLOWUP_EXPIRED_SCORE_PENALTY_POINTS = 4
FOLLOWUP_EXPIRED_SCORE_PENALTY_CAP = 24
FOLLOW_UP_AUTO_EXPIRE_STATUSES = (
    Lead.Status.CALL_BACK,
)
ACTIVE_QUEUE_STATUSES = (
    Lead.Status.NEW,
    Lead.Status.INTERESTED,
)
STAFF_CALL_QUEUE_STATUSES = (
    Lead.Status.NEW,
)
FOLLOW_UP_STATUSES = (
    Lead.Status.INTERESTED,
    Lead.Status.CALL_BACK,
)
RECOVERY_LEAD_STATUSES = (
    Lead.Status.NOT_INTERESTED,
    Lead.Status.NO_ANSWER,
)
TERMINAL_QUEUE_STATUSES = (
    Lead.Status.NOT_INTERESTED,
    Lead.Status.NO_ANSWER,
    Lead.Status.CONVERTED,
)
NAME_COLUMN_ALIASES = {
    "name",
    "leadname",
    "lead name",
    "customername",
    "customer name",
    "customer",
    "fullname",
    "full name",
    "clientname",
    "client name",
    "client",
}
PHONE_COLUMN_ALIASES = {
    "phone",
    "phone no",
    "phone no.",
    "phonenumber",
    "phone number",
    "mobile",
    "mobile no",
    "mobile no.",
    "mobilenumber",
    "mobile number",
    "contact",
    "contact no",
    "contact no.",
    "contactnumber",
    "contact number",
    "number",
    "whatsapp",
    "whatsapp number",
}

APP_NOTIFICATION_SEVERITY_ORDER = {
    AppNotification.Severity.CRITICAL: 0,
    AppNotification.Severity.WARNING: 1,
    AppNotification.Severity.NORMAL: 2,
    AppNotification.Severity.GOOD: 3,
}
LEAD_ID_COLUMN_ALIASES = {
    "id",
    "lead id",
    "leadid",
}
STATUS_COLUMN_ALIASES = {
    "status",
    "lead status",
    "call status",
    "result",
    "followup status",
    "follow up status",
}
CALLBACK_WINDOW_COLUMN_ALIASES = {
    "callback window",
    "callback slot",
    "call back slot",
    "time slot",
    "schedule",
    "callback time",
    "followup slot",
    "follow up slot",
    "followup time",
    "follow up time",
}
NOTES_COLUMN_ALIASES = {
    "notes",
    "remarks",
    "comment",
    "comments",
}
HANDOVER_STATUS_COLUMN_ALIASES = {
    "handover",
    "handover status",
    "handoverstate",
    "handover status",
    "client status",
    "client handover",
    "handoverstate",
}
ASSIGNED_STAFF_PHONE_COLUMN_ALIASES = {
    "assigned to phone",
    "assigned staff phone",
    "staff phone",
    "assignee phone",
}
ASSIGNED_STAFF_NAME_COLUMN_ALIASES = {
    "assigned to",
    "assigned staff",
    "staff",
    "staff name",
    "assignee",
}


class TrainingRequiredError(Exception):
    def __init__(self, payload):
        super().__init__("Complete mandatory training before starting work.")
        self.payload = payload


def get_company_profile():
    profile, _ = CompanyProfile.objects.get_or_create(
        id=1,
        defaults={
            "company_name": "Heavenection",
            "country": "India",
        },
    )
    return profile


def get_lead_queue_limit():
    profile = get_company_profile()
    return max(1, int(profile.lead_queue_target_per_staff or DEFAULT_LEAD_QUEUE_LIMIT))


def _work_review_rules():
    profile = get_company_profile()
    attempt_threshold = max(
        1,
        int(getattr(profile, "work_review_zero_talk_attempt_threshold", MIN_REAL_CALLS_PER_ATTEMPT_BLOCK) or 1),
    )
    idle_gap_seconds = max(
        1,
        int(getattr(profile, "work_review_idle_gap_seconds", CALL_ACTIVITY_IDLE_BREAK_SECONDS) or 1),
    )
    connected_cooldown_seconds = max(
        0,
        int(getattr(profile, "work_review_connected_cooldown_seconds", CONNECTED_CALL_COOLDOWN_SECONDS) or 0),
    )
    followup_expired_penalty_points = max(
        0,
        int(
            getattr(
                profile,
                "work_review_followup_expired_penalty_points",
                FOLLOWUP_EXPIRED_SCORE_PENALTY_POINTS,
            )
            or 0
        ),
    )
    followup_expired_penalty_cap = max(
        0,
        int(
            getattr(
                profile,
                "work_review_followup_expired_penalty_cap",
                FOLLOWUP_EXPIRED_SCORE_PENALTY_CAP,
            )
            or 0
        ),
    )
    return {
        "attempt_threshold": attempt_threshold,
        "idle_gap_seconds": idle_gap_seconds,
        "connected_cooldown_seconds": connected_cooldown_seconds,
        "followup_expired_penalty_points": followup_expired_penalty_points,
        "followup_expired_penalty_cap": followup_expired_penalty_cap,
    }


def _today_range(now=None):
    if now is None:
        now = timezone.now()
    local_now = timezone.localtime(now)
    today = local_now.date()
    start = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
    end = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.max.time()))
    return today, start, end


def _format_currency(value):
    rounded = round(float(value or 0), 2)
    return f"Rs. {rounded:,.2f}"


def _format_hours(total_seconds):
    return f"{round((total_seconds or 0) / 3600, 1)}h"


def _format_duration(total_seconds):
    total_seconds = int(total_seconds or 0)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return f"{minutes:02d}:{seconds:02d}"


def _normalize_datetime_value(value):
    if not value or not isinstance(value, datetime):
        return value
    if timezone.is_naive(value):
        return timezone.make_aware(value, timezone.get_current_timezone())
    return value


def _call_activity_window(call):
    try:
        raw_start = _normalize_datetime_value(
            call.get("start_time") if isinstance(call, dict) else getattr(call, "start_time", None)
        )
        if not call or not raw_start:
            return None, None, 0

        raw_end = _normalize_datetime_value(
            (
                call.get("end_time")
                if isinstance(call, dict)
                else getattr(call, "end_time", None)
            )
            or raw_start
        )
        if raw_end < raw_start:
            raw_end = raw_start

        activity_start = raw_start
        created_at = _normalize_datetime_value(
            call.get("created_at") if isinstance(call, dict) else getattr(call, "created_at", None)
        )
        if created_at and created_at < activity_start:
            dial_lead_seconds = max(0, int((activity_start - created_at).total_seconds()))
            if 0 < dial_lead_seconds <= CALL_ACTIVITY_DIAL_LOOKBACK_SECONDS:
                activity_start = created_at

        duration_value = call.get("duration_seconds") if isinstance(call, dict) else getattr(call, "duration_seconds", 0)
        duration_seconds = max(0, int(duration_value or 0))
        return activity_start, raw_end, duration_seconds
    except Exception:
        logger.warning(
            "Skipping malformed call activity window",
            extra={"call_id": str(getattr(call, "id", ""))},
            exc_info=True,
        )
        return None, None, 0


def _call_activity_cooldown_seconds(block_end, *, range_end=None, rules=None):
    if not block_end:
        return 0

    active_rules = rules or _work_review_rules()
    cooldown_seconds = int(active_rules.get("connected_cooldown_seconds", CONNECTED_CALL_COOLDOWN_SECONDS) or 0)
    if cooldown_seconds <= 0:
        return 0

    cooldown_end = block_end + timedelta(seconds=cooldown_seconds)
    if range_end is not None:
        cooldown_end = min(cooldown_end, range_end)

    local_block_end = timezone.localtime(block_end)
    day_end = timezone.make_aware(
        timezone.datetime.combine(local_block_end.date(), timezone.datetime.max.time())
    )
    cooldown_end = min(cooldown_end, day_end)
    return max(0, int((cooldown_end - block_end).total_seconds()))


def _split_call_activity_blocks(calls, *, rules=None):
    blocks = []
    block_calls = []
    block_start = None
    block_end = None
    active_rules = rules or _work_review_rules()
    idle_gap_seconds = int(active_rules.get("idle_gap_seconds", CALL_ACTIVITY_IDLE_BREAK_SECONDS) or 1)

    for call in calls:
        start_time, end_time, _duration_seconds = _call_activity_window(call)
        if not start_time:
            continue

        if block_start is None:
            block_calls = [call]
            block_start = start_time
            block_end = end_time
            continue

        gap_seconds = max(0, int((start_time - block_end).total_seconds()))
        if gap_seconds > idle_gap_seconds:
            blocks.append((block_calls, block_start, block_end))
            block_calls = [call]
            block_start = start_time
            block_end = end_time
            continue

        block_calls.append(call)
        if end_time > block_end:
            block_end = end_time

    if block_calls:
        blocks.append((block_calls, block_start, block_end))
    return blocks


def _call_activity_blocks_with_stats(calls, *, rules=None):
    blocks = []
    active_rules = rules or _work_review_rules()
    for block_calls, block_start, block_end in _split_call_activity_blocks(calls, rules=active_rules):
        real_call_activity_seconds = 0
        real_calls_in_block = 0
        zero_seconds_in_block = 0
        real_call_segments = []
        first_real_call_start = None

        for call in block_calls:
            activity_start, activity_end, duration_seconds = _call_activity_window(call)
            if not activity_start:
                continue

            activity_span_seconds = max(0, int((activity_end - activity_start).total_seconds()))
            if duration_seconds > 0:
                real_calls_in_block += 1
                real_call_activity_seconds += max(activity_span_seconds, duration_seconds)
                real_call_segments.append((activity_start, activity_end))
                if not first_real_call_start or activity_start < first_real_call_start:
                    first_real_call_start = activity_start
            else:
                zero_seconds_in_block += 1

        block_seconds = max(0, int((block_end - block_start).total_seconds())) if block_start and block_end else 0
        blocks.append(
            {
                "calls": block_calls,
                "block_start": block_start,
                "block_end": block_end,
                "block_seconds": block_seconds,
                "real_call_activity_seconds": real_call_activity_seconds,
                "real_calls_in_block": real_calls_in_block,
                "zero_seconds_in_block": zero_seconds_in_block,
                "real_call_segments": real_call_segments,
                "first_real_call_start": first_real_call_start,
            }
        )
    return blocks


def _call_activity_block_summary(calls, *, range_end=None, rules=None):
    total_seconds = 0
    attempt_count = 0
    suspicious_block_count = 0
    zero_only_block_count = 0
    suspicious_attempt_count = 0
    zero_second_attempt_count = 0
    real_call_attempt_count = 0

    active_rules = rules or _work_review_rules()
    attempt_threshold = int(active_rules.get("attempt_threshold", MIN_REAL_CALLS_PER_ATTEMPT_BLOCK) or 1)
    blocks = _call_activity_blocks_with_stats(calls, rules=active_rules)
    for block in blocks:
        attempt_count += len(block["calls"])
        zero_second_attempt_count += block["zero_seconds_in_block"]
        real_call_attempt_count += block["real_calls_in_block"]

    zero_streak = []
    zero_streak_attempts = 0

    def flush_zero_streak(*, force_zero_only=False):
        nonlocal total_seconds, suspicious_block_count, suspicious_attempt_count, zero_only_block_count
        nonlocal zero_streak, zero_streak_attempts
        if not zero_streak:
            return
        if force_zero_only or zero_streak_attempts >= attempt_threshold:
            zero_only_block_count += len(zero_streak)
        else:
            for block in zero_streak:
                block_seconds = block["block_seconds"]
                total_seconds += max(block_seconds, block["real_call_activity_seconds"])
                total_seconds += _call_activity_cooldown_seconds(block["block_end"], range_end=range_end, rules=active_rules)
        zero_streak = []
        zero_streak_attempts = 0

    for block in blocks:
        if block["real_calls_in_block"] <= 0:
            zero_streak.append(block)
            zero_streak_attempts += len(block["calls"])
            continue

        streak_reached = zero_streak_attempts >= attempt_threshold
        flush_zero_streak(force_zero_only=streak_reached)

        block_seconds = block["block_seconds"]
        if (
            len(block["calls"]) >= attempt_threshold
            and (block["real_calls_in_block"] * attempt_threshold) < len(block["calls"])
        ):
            suspicious_block_count += 1
            suspicious_attempt_count += len(block["calls"])
            total_seconds += block["real_call_activity_seconds"]
            continue

        if streak_reached and block["first_real_call_start"]:
            cooldown_seconds = _call_activity_cooldown_seconds(block["block_end"], range_end=range_end, rules=active_rules)
            adjusted_start = block["first_real_call_start"]
            adjusted_segment = (
                adjusted_start,
                block["block_end"] + timedelta(seconds=cooldown_seconds),
            )
            total_seconds += _segment_seconds(adjusted_segment)
        else:
            total_seconds += max(block_seconds, block["real_call_activity_seconds"])
            total_seconds += _call_activity_cooldown_seconds(block["block_end"], range_end=range_end, rules=active_rules)

    flush_zero_streak()

    return {
        "active_seconds": total_seconds,
        "attempt_count": attempt_count,
        "real_call_count": real_call_attempt_count,
        "zero_second_attempt_count": zero_second_attempt_count,
        "suspicious_block_count": suspicious_block_count,
        "zero_only_block_count": zero_only_block_count,
        "suspicious_attempt_count": suspicious_attempt_count,
    }


def _effective_active_seconds_map(*, start_at=None, end_at=None, staff_ids=None, rules=None):
    call_queryset = _payable_work_hour_call_queryset()
    active_rules = rules or _work_review_rules()

    if start_at is not None and end_at is not None:
        call_queryset = call_queryset.filter(start_time__range=(start_at, end_at))

    if staff_ids is not None:
        staff_ids = list(staff_ids)
        call_queryset = call_queryset.filter(staff_id__in=staff_ids)

    calls_by_staff_day = defaultdict(list)
    for call in call_queryset.annotate(activity_day=TruncDate("start_time")).only(
        "staff_id",
        "start_time",
        "end_time",
        "duration_seconds",
        "is_verified",
        "created_at",
    ).order_by("staff_id", "start_time", "end_time", "id"):
        calls_by_staff_day[(call.staff_id, getattr(call, "activity_day", None))].append(call)

    totals_by_staff = defaultdict(int)
    for day_key, calls in calls_by_staff_day.items():
        totals_by_staff[day_key[0]] += _call_activity_block_summary(
            calls,
            range_end=end_at,
            rules=active_rules,
        )["active_seconds"]

    if staff_ids is not None:
        for staff_id in staff_ids:
            totals_by_staff.setdefault(staff_id, 0)

    return dict(totals_by_staff)


def _effective_active_insights_map(*, start_at=None, end_at=None, staff_ids=None, rules=None):
    call_queryset = Call.objects.filter(end_time__isnull=False, is_verified=True)
    active_rules = rules or _work_review_rules()

    if start_at is not None and end_at is not None:
        call_queryset = call_queryset.filter(start_time__range=(start_at, end_at))

    if staff_ids is not None:
        staff_ids = list(staff_ids)
        call_queryset = call_queryset.filter(staff_id__in=staff_ids)

    calls_by_staff_day = defaultdict(list)
    for call in (
        call_queryset.annotate(activity_day=TruncDate("start_time"))
        .only(
            "staff_id",
            "start_time",
            "end_time",
            "duration_seconds",
            "is_verified",
            "created_at",
        )
        .order_by("staff_id", "start_time", "end_time", "id")
    ):
        calls_by_staff_day[(call.staff_id, getattr(call, "activity_day", None))].append(call)

    insights_by_staff = defaultdict(
        lambda: {
            "active_seconds": 0,
            "attempt_count": 0,
            "real_call_count": 0,
            "zero_second_attempt_count": 0,
            "suspicious_block_count": 0,
            "zero_only_block_count": 0,
            "suspicious_attempt_count": 0,
        }
    )
    for day_key, calls in calls_by_staff_day.items():
        summary = _call_activity_block_summary(calls, rules=active_rules)
        staff_summary = insights_by_staff[day_key[0]]
        for field_name, value in summary.items():
            staff_summary[field_name] += value

    if staff_ids is not None:
        for staff_id in staff_ids:
            insights_by_staff.setdefault(
                staff_id,
                {
                    "active_seconds": 0,
                    "attempt_count": 0,
                    "real_call_count": 0,
                    "zero_second_attempt_count": 0,
                    "suspicious_block_count": 0,
                    "zero_only_block_count": 0,
                    "suspicious_attempt_count": 0,
                },
            )

    return dict(insights_by_staff)


def _effective_active_seconds_for_staff(*, staff, start_at=None, end_at=None):
    return _effective_active_seconds_map(
        start_at=start_at,
        end_at=end_at,
        staff_ids=[staff.id],
    ).get(staff.id, 0)


def _format_datetime(value, fallback="--"):
    if not value:
        return fallback
    try:
        if isinstance(value, date) and not isinstance(value, datetime):
            return value.strftime("%d %b %Y")
        value = _normalize_datetime_value(value)
        return timezone.localtime(value).strftime("%d %b %Y, %I:%M %p")
    except Exception:
        logger.warning("Unable to format datetime value safely", exc_info=True)
        try:
            return str(value)
        except Exception:
            return fallback


def _notification_payload_from_row(row):
    created_at = row.created_at
    return {
        "id": str(row.id),
        "title": row.title or "",
        "message": row.message,
        "severity": row.severity,
        "severity_label": row.get_severity_display(),
        "source": row.source,
        "source_label": row.get_source_display(),
        "audience": row.audience,
        "audience_label": row.get_audience_display(),
        "target_staff_id": str(row.target_staff_id) if row.target_staff_id else "",
        "target_staff_name": row.target_staff.name if row.target_staff_id else "",
        "is_active": row.is_active,
        "allow_manual_close": row.allow_manual_close,
        "auto_dismiss_seconds": int(row.auto_dismiss_seconds or 0),
        "show_from": row.show_from.isoformat() if row.show_from else None,
        "show_from_label": _format_datetime(row.show_from),
        "show_until": row.show_until.isoformat() if row.show_until else None,
        "show_until_label": _format_datetime(row.show_until, fallback="Until closed")
        if row.show_until
        else "Until closed",
        "created_at": created_at.isoformat() if created_at else None,
        "created_at_label": _format_datetime(created_at),
        "created_by_name": row.created_by.name if row.created_by_id else "",
    }


def build_staff_active_notifications_payload(staff, *, now=None):
    now = now or timezone.now()
    queryset = (
        AppNotification.objects.select_related("target_staff", "created_by")
        .filter(is_active=True, show_from__lte=now)
        .filter(Q(show_until__isnull=True) | Q(show_until__gte=now))
        .filter(
            Q(audience=AppNotification.Audience.ALL_STAFF)
            | Q(audience=AppNotification.Audience.SINGLE_STAFF, target_staff=staff)
        )
    )
    rows = sorted(
        queryset,
        key=lambda row: (
            APP_NOTIFICATION_SEVERITY_ORDER.get(row.severity, 99),
            -(row.show_from.timestamp() if row.show_from else 0),
            -(row.created_at.timestamp() if row.created_at else 0),
        ),
    )
    return [_notification_payload_from_row(row) for row in rows]


def build_app_notification_management_payload():
    rows = [
        _notification_payload_from_row(row)
        for row in AppNotification.objects.select_related("target_staff", "created_by").all()[:50]
    ]
    active_count = sum(1 for row in rows if row["is_active"])
    critical_count = sum(
        1 for row in rows if row["is_active"] and row["severity"] == AppNotification.Severity.CRITICAL
    )
    return {
        "summary": {
            "total_notifications": len(rows),
            "active_notifications": active_count,
            "critical_notifications": critical_count,
        },
        "rows": rows,
    }


def build_admin_web_alert_payload(*, limit=10, monitoring_payload=None):
    monitoring_payload = monitoring_payload or build_live_monitoring_payload()
    company_profile = get_company_profile()
    now = timezone.now()
    today = timezone.localdate()
    _, today_start, today_end = _today_range()
    alerts = []
    staff_rows = monitoring_payload.get("staff_rows", [])

    def _minutes_from_duration_label(label):
        text = str(label or "").strip().lower()
        if not text:
            return 0
        total_minutes = Decimal("0")
        hour_match = re.search(r"(\d+(?:\.\d+)?)\s*h", text)
        minute_match = re.search(r"(\d+(?:\.\d+)?)\s*m", text)
        second_match = re.search(r"(\d+(?:\.\d+)?)\s*s", text)
        if hour_match:
            total_minutes += Decimal(hour_match.group(1)) * Decimal("60")
        if minute_match:
            total_minutes += Decimal(minute_match.group(1))
        if second_match:
            total_minutes += Decimal(second_match.group(1)) / Decimal("60")
        if total_minutes <= 0 and re.fullmatch(r"\d+(?:\.\d+)?", text):
            total_minutes = Decimal(text)
        return int(total_minutes)

    staff_ids = [uuid.UUID(str(row["id"])) for row in staff_rows if row.get("id")]
    pending_status_by_staff = {}
    repeated_sync_issue_counts = {}
    if staff_ids:
        pending_calls = (
            Call.objects.filter(
                staff_id__in=staff_ids,
                status=Call.Status.STARTED,
                end_time__isnull=False,
            )
            .select_related("lead", "staff")
            .order_by("-end_time", "-updated_at")
        )
        for call in pending_calls:
            pending_status_by_staff.setdefault(call.staff_id, call)

        repeated_sync_issue_counts = {
            row["staff"]: int(row["total"] or 0)
            for row in (
                Call.objects.filter(
                    staff_id__in=staff_ids,
                    auto_skipped_sync_issue=True,
                    start_time__gte=today_start,
                    start_time__lte=today_end,
                )
                .values("staff")
                .annotate(total=Count("id"))
            )
        }

    followup_due_now_count = Lead.objects.filter(
        status=Lead.Status.INTERESTED,
        callback_date=today,
    ).count()
    followup_overdue_count = Lead.objects.filter(
        status=Lead.Status.INTERESTED,
        callback_date__lt=today,
    ).count()
    unassigned_new_lead_count = Lead.objects.filter(
        status=Lead.Status.NEW,
        assigned_to__isnull=True,
    ).count()
    pending_salary_amount = Decimal("0.00")
    pending_salary_staff_count = 0
    for staff in _staff_queryset():
        (due_start, due_end), due_cycle = _due_period_for_staff(staff, today)
        due_snapshot = _salary_period_snapshot(
            staff,
            period_start=due_start,
            period_end=due_end,
            payout_cycle=due_cycle,
        )
        if due_snapshot["balance"] > Decimal("0.00"):
            pending_salary_amount += due_snapshot["balance"]
            pending_salary_staff_count += 1

    threshold_per_hour = int(company_profile.hourly_call_bonus_threshold or 0)
    bonus_enabled = bool(company_profile.hourly_call_bonus_enabled and threshold_per_hour > 0)

    if pending_salary_staff_count > 0:
        alerts.append(
            {
                "id": "salary-due-global",
                "severity": "warning",
                "severity_label": "Warning",
                "title": "Salary payment is due",
                "message": (
                    f"{pending_salary_staff_count} staff member(s) have due salary waiting. "
                    f"Total pending salary is {_format_currency(pending_salary_amount)}."
                ),
                "staff_name": "",
                "target_url": "/salary/",
                "target_label": "Open Salary Overview",
                "meta_label": "Payroll attention",
                "sort_score": 160 + min(pending_salary_staff_count, 25),
            }
        )

    if followup_overdue_count > 0:
        alerts.append(
            {
                "id": "followup-overdue-global",
                "severity": "critical",
                "severity_label": "Critical",
                "title": "Overdue follow-ups need attention",
                "message": f"{followup_overdue_count} follow-up lead(s) are now overdue and should be reviewed immediately.",
                "staff_name": "",
                "target_url": "/followups/",
                "target_label": "Open Follow-Ups",
                "meta_label": "Overdue queue pressure",
                "sort_score": 220 + (followup_overdue_count * 3),
            }
        )

    if followup_due_now_count > 0:
        alerts.append(
            {
                "id": "followup-due-global",
                "severity": "warning",
                "severity_label": "Warning",
                "title": "Follow-ups are due today",
                "message": f"{followup_due_now_count} follow-up lead(s) are scheduled for today and should stay visible to the team.",
                "staff_name": "",
                "target_url": "/followups/",
                "target_label": "Open Follow-Ups",
                "meta_label": "Due today",
                "sort_score": 145 + (followup_due_now_count * 2),
            }
        )

    if unassigned_new_lead_count > 0:
        alerts.append(
            {
                "id": "new-leads-unassigned-global",
                "severity": "warning",
                "severity_label": "Warning",
                "title": "Fresh leads are waiting for allocation",
                "message": f"{unassigned_new_lead_count} new lead(s) are still unassigned and can be allocated now.",
                "staff_name": "",
                "target_url": "/leads/",
                "target_label": "Open Lead Management",
                "meta_label": "Allocation pending",
                "sort_score": 120 + min(unassigned_new_lead_count, 25),
            }
        )

    for row in staff_rows:
        staff_id = str(row.get("id") or "").strip()
        if not staff_id:
            continue
        staff_uuid = uuid.UUID(staff_id)
        staff_name = row.get("name") or "Staff member"
        profile_url = f"/staff/{staff_id}/"
        work_review_url = "/work-review/"
        quality_label = row.get("quality_label") or ""
        quality_note = row.get("quality_note") or row.get("status_note") or ""
        suspicious_blocks = int(row.get("suspicious_block_count") or 0)
        zero_only_blocks = int(row.get("zero_only_block_count") or 0)
        missed_callbacks = int(row.get("missed_callbacks") or 0)
        invalid_short_count = int(row.get("invalid_short_count") or 0)
        zero_second_count = int(row.get("zero_second_attempt_count") or 0)
        calls_today = int(row.get("calls_today") or 0)
        assigned_leads = int(row.get("assigned_leads") or 0)
        online_label = row.get("online_label") or ""
        is_active_account = bool(row.get("is_active", True))
        work_minutes = _minutes_from_duration_label(row.get("active_hours_today"))
        pending_status_call = pending_status_by_staff.get(staff_uuid)
        sync_issue_count = int(repeated_sync_issue_counts.get(staff_uuid, 0) or 0)

        if sync_issue_count >= 2:
            alerts.append(
                {
                    "id": f"sync-issues-{staff_id}",
                    "severity": "critical",
                    "severity_label": "Critical",
                    "title": f"{staff_name} has repeated call sync issues",
                    "message": f"{sync_issue_count} call(s) were auto-skipped for sync issues today. This device or workflow needs review.",
                    "staff_name": staff_name,
                    "target_url": "/calls/",
                    "target_label": "Open Call Details",
                    "meta_label": "Repeated sync issue",
                    "sort_score": 185 + (sync_issue_count * 8),
                }
            )

        if pending_status_call:
            duration_label = _format_work_duration_label(pending_status_call.duration_seconds or 0)
            alerts.append(
                {
                    "id": f"pending-remark-{staff_id}",
                    "severity": "critical",
                    "severity_label": "Critical",
                    "title": f"{staff_name} has a call without final remark",
                    "message": f"The call with {pending_status_call.lead.name} ended at { _format_datetime(pending_status_call.end_time) } and still needs a final status update.",
                    "staff_name": staff_name,
                    "target_url": profile_url,
                    "target_label": "Open Staff Profile",
                    "meta_label": f"{duration_label} call pending",
                    "sort_score": 176 + int(pending_status_call.duration_seconds or 0),
                }
            )

        if is_active_account and online_label in {"Online", "Away", "Warning"} and assigned_leads <= 0 and not row.get("is_on_call"):
            alerts.append(
                {
                    "id": f"queue-empty-{staff_id}",
                    "severity": "warning",
                    "severity_label": "Warning",
                    "title": f"{staff_name} has no leads in queue",
                    "message": "This staff account is active in the app but does not currently have queue leads to work on.",
                    "staff_name": staff_name,
                    "target_url": "/leads/",
                    "target_label": "Open Lead Management",
                    "meta_label": online_label,
                    "sort_score": 92,
                }
            )

        if quality_label == "Review Needed" or suspicious_blocks > 0 or zero_only_blocks > 0:
            review_message = quality_note or "Call patterns need supervisor review."
            alerts.append(
                {
                    "id": f"review-{staff_id}",
                    "severity": "critical",
                    "severity_label": "Critical",
                    "title": f"{staff_name} needs review",
                    "message": review_message,
                    "staff_name": staff_name,
                    "target_url": work_review_url,
                    "target_label": "Open Work Review",
                    "meta_label": row.get("attempt_review_label") or row.get("online_label") or "Needs review",
                    "sort_score": 100 + suspicious_blocks * 10 + zero_only_blocks * 8 + missed_callbacks * 4,
                }
            )

        if missed_callbacks > 0:
            alerts.append(
                {
                    "id": f"callback-{staff_id}",
                    "severity": "warning",
                    "severity_label": "Warning",
                    "title": f"{staff_name} has missed follow-ups",
                    "message": f"{missed_callbacks} scheduled follow-up lead(s) now need attention.",
                    "staff_name": staff_name,
                    "target_url": "/followups/",
                    "target_label": "Open Follow-Ups",
                    "meta_label": row.get("online_label") or "Follow-up delay",
                    "sort_score": 70 + missed_callbacks * 6,
                }
            )

        if invalid_short_count >= 3:
            alerts.append(
                {
                    "id": f"invalid-short-{staff_id}",
                    "severity": "warning",
                    "severity_label": "Warning",
                    "title": f"{staff_name} has too many invalid short calls",
                    "message": f"{invalid_short_count} invalid short call(s) were recorded today. Review the calling pattern before it grows further.",
                    "staff_name": staff_name,
                    "target_url": "/calls/",
                    "target_label": "Open Call Details",
                    "meta_label": "Invalid short spike",
                    "sort_score": 88 + (invalid_short_count * 4),
                }
            )

        if quality_label == "Needs Attention" or invalid_short_count > 0 or zero_second_count > 0:
            alerts.append(
                {
                    "id": f"attention-{staff_id}",
                    "severity": "normal",
                    "severity_label": "Normal",
                    "title": f"{staff_name} shows lighter warning signals",
                    "message": quality_note or "Some empty or short call attempts should be reviewed.",
                    "staff_name": staff_name,
                    "target_url": profile_url,
                    "target_label": "Open Staff Profile",
                    "meta_label": (
                        f"{invalid_short_count} invalid short · {zero_second_count} zero-second"
                        if invalid_short_count or zero_second_count
                        else (row.get("online_label") or "Attention")
                    ),
                    "sort_score": 40 + invalid_short_count * 4 + zero_second_count,
                }
            )

        if bonus_enabled and work_minutes >= 60:
            completed_hours = work_minutes // 60
            threshold_calls = completed_hours * threshold_per_hour
            remaining_calls = threshold_calls - calls_today
            if remaining_calls > 0 and remaining_calls <= 5:
                alerts.append(
                    {
                        "id": f"bonus-near-{staff_id}",
                        "severity": "good",
                        "severity_label": "Good",
                        "title": f"{staff_name} is close to hourly bonus",
                        "message": f"Only {remaining_calls} more call(s) are needed to reach the current hourly bonus threshold.",
                        "staff_name": staff_name,
                        "target_url": profile_url,
                        "target_label": "Open Staff Profile",
                        "meta_label": f"{calls_today}/{threshold_calls} calls",
                        "sort_score": 18 + (5 - remaining_calls),
                    }
                )
            elif remaining_calls <= 0 and threshold_calls > 0:
                alerts.append(
                    {
                        "id": f"bonus-hit-{staff_id}",
                        "severity": "good",
                        "severity_label": "Good",
                        "title": f"{staff_name} reached hourly bonus pace",
                        "message": "This staff member has crossed the current hourly call threshold and is now earning extra call bonus for this completed hour block.",
                        "staff_name": staff_name,
                        "target_url": profile_url,
                        "target_label": "Open Staff Profile",
                        "meta_label": f"{calls_today} calls today",
                        "sort_score": 22,
                    }
                )

    alerts.sort(
        key=lambda item: (
            APP_NOTIFICATION_SEVERITY_ORDER.get(item["severity"], 99),
            -int(item.get("sort_score") or 0),
            item.get("title") or "",
        )
    )
    alerts = alerts[: max(1, int(limit or 10))]
    critical_count = sum(1 for item in alerts if item["severity"] == "critical")
    warning_count = sum(1 for item in alerts if item["severity"] == "warning")
    return {
        "summary": {
            "total_alerts": len(alerts),
            "critical_alerts": critical_count,
            "warning_alerts": warning_count,
            "generated_at_label": monitoring_payload.get("generated_at_label", _format_datetime(timezone.now())),
        },
        "alerts": alerts,
    }


def build_cached_admin_web_alert_payload(*, limit=10, cache_seconds=15, monitoring_payload=None):
    if monitoring_payload is not None:
        return build_admin_web_alert_payload(limit=limit, monitoring_payload=monitoring_payload)

    cache_key = f"telecalling:admin-web-alerts:v1:{int(limit or 10)}"
    cached_payload = cache.get(cache_key)
    if cached_payload is not None:
        return cached_payload

    payload = build_admin_web_alert_payload(limit=limit)
    cache.set(cache_key, payload, max(1, int(cache_seconds or 15)))
    return payload


def build_staff_document_url(staff, document_type, *, request=None, route_name="staff-document-page"):
    field_name = f"{document_type}_photo"
    file_field = getattr(staff, field_name, None)
    if not file_field:
        return ""
    if not file_field.name or not file_field.storage.exists(file_field.name):
        return ""

    if route_name == "api-staff-profile-document":
        url = reverse(route_name, args=[document_type])
    else:
        url = reverse(route_name, args=[staff.id, document_type])

    if request:
        url = request.build_absolute_uri(url)

    version = int((staff.updated_at or timezone.now()).timestamp())
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}v={version}"


def _week_range(now=None):
    now = now or timezone.now()
    local_now = timezone.localtime(now)
    start_date = local_now.date() - timedelta(days=local_now.weekday())
    start = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
    return start, now


def _month_range(now=None):
    now = now or timezone.now()
    local_now = timezone.localtime(now)
    start_date = local_now.date().replace(day=1)
    start = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
    return start, now


def _month_start(value):
    return value.replace(day=1)


def _month_end(value):
    return date(value.year, value.month, calendar.monthrange(value.year, value.month)[1])


def _month_range_for_reference(reference_date, *, end_at=None):
    start_date = _month_start(reference_date)
    end_date = _month_end(reference_date)
    start = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
    end = timezone.make_aware(timezone.datetime.combine(end_date, timezone.datetime.max.time()))
    if end_at:
        end = min(end, end_at)
    return start, end


def _shift_month(value, delta_months):
    month_index = (value.month - 1) + delta_months
    year = value.year + (month_index // 12)
    month = (month_index % 12) + 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def _parse_month_value(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        year_str, month_str = raw.split("-", 1)
        year = int(year_str)
        month = int(month_str)
        if month < 1 or month > 12:
            return None
        return date(year, month, 1)
    except (ValueError, AttributeError):
        return None


def _month_option_rows(*, reference_date, selected_value="", months_back=6):
    options = []
    for offset in range(months_back):
        month_date = _shift_month(reference_date.replace(day=1), -offset)
        value = month_date.strftime("%Y-%m")
        options.append(
            {
                "value": value,
                "label": month_date.strftime("%b %Y"),
                "is_selected": value == selected_value,
            }
        )
    return options


def _parse_date_value(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _day_range_for_date(target_date):
    start = timezone.make_aware(timezone.datetime.combine(target_date, timezone.datetime.min.time()))
    end = timezone.make_aware(timezone.datetime.combine(target_date, timezone.datetime.max.time()))
    return start, end


def _decimal_hours(total_seconds):
    return Decimal(str(total_seconds or 0)) / Decimal("3600")


def _decimal_minutes(total_seconds):
    return Decimal(str(total_seconds or 0)) / Decimal("60")


def _money(value):
    return Decimal(value or 0).quantize(TWOPLACES)


def _seconds_from_decimal_hours(total_hours):
    hours_value = Decimal(total_hours or 0)
    return max(
        0,
        int((hours_value * Decimal("3600")).to_integral_value(rounding=ROUND_HALF_UP)),
    )


def _format_work_duration_label(total_seconds):
    total_seconds = max(0, int(total_seconds or 0))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    parts = []
    if hours:
        parts.append(f"{hours}h")
    if minutes:
        parts.append(f"{minutes}m")
    if seconds or not parts:
        parts.append(f"{seconds}s")
    return " ".join(parts)


def _payable_work_hour_call_queryset():
    return Call.objects.filter(end_time__isnull=False, is_verified=True).exclude(
        status=Call.Status.INVALID_SHORT
    )


def _hourly_bonus_call_queryset(queryset=None):
    base_queryset = queryset if queryset is not None else Call.objects.all()
    return base_queryset.filter(is_verified=True).exclude(status=Call.Status.INVALID_SHORT)


def _empty_hourly_bonus_summary():
    return {
        "completed_hours": 0,
        "completed_bonus_calls": 0,
        "threshold_calls": 0,
        "extra_calls": 0,
        "bonus_amount": Decimal("0.00"),
    }


def _segment_seconds(segment):
    start_time, end_time = segment
    return max(0, int((end_time - start_time).total_seconds()))


def _truncate_segments_to_seconds(segments, total_seconds):
    remaining_seconds = max(0, int(total_seconds or 0))
    if remaining_seconds <= 0:
        return []

    truncated_segments = []
    for start_time, end_time in segments:
        if remaining_seconds <= 0:
            break
        segment_seconds = _segment_seconds((start_time, end_time))
        if segment_seconds <= 0:
            continue
        if segment_seconds <= remaining_seconds:
            truncated_segments.append((start_time, end_time))
            remaining_seconds -= segment_seconds
            continue

        truncated_segments.append((start_time, start_time + timedelta(seconds=remaining_seconds)))
        remaining_seconds = 0

    return truncated_segments


def _split_segments_into_buckets(segments, bucket_seconds):
    bucket_seconds = max(1, int(bucket_seconds or 0))
    buckets = []
    current_bucket = []
    remaining_bucket_seconds = bucket_seconds

    for start_time, end_time in segments:
        segment_start = start_time
        segment_end = end_time
        while segment_end > segment_start:
            slice_seconds = min(
                remaining_bucket_seconds,
                _segment_seconds((segment_start, segment_end)),
            )
            if slice_seconds <= 0:
                break

            slice_end = segment_start + timedelta(seconds=slice_seconds)
            current_bucket.append((segment_start, slice_end))
            remaining_bucket_seconds -= slice_seconds
            segment_start = slice_end

            if remaining_bucket_seconds <= 0:
                buckets.append(current_bucket)
                current_bucket = []
                remaining_bucket_seconds = bucket_seconds

    return buckets


def _call_bonus_timestamp(call):
    return _normalize_datetime_value((
        call.get("end_time")
        if isinstance(call, dict)
        else getattr(call, "end_time", None)
    ) or (
        call.get("start_time")
        if isinstance(call, dict)
        else getattr(call, "start_time", None)
    ))


def _count_calls_in_segments(calls, segments):
    if not calls or not segments:
        return 0

    ordered_segments = [
        (start_time, end_time)
        for start_time, end_time in sorted(segments, key=lambda value: value[0])
        if end_time >= start_time
    ]
    if not ordered_segments:
        return 0

    ordered_timestamps = sorted(
        timestamp
        for timestamp in (_call_bonus_timestamp(call) for call in calls)
        if timestamp is not None
    )
    if not ordered_timestamps:
        return 0

    count = 0
    segment_index = 0
    for timestamp in ordered_timestamps:
        while segment_index < len(ordered_segments) and timestamp > ordered_segments[segment_index][1]:
            segment_index += 1
        if segment_index >= len(ordered_segments):
            break
        if ordered_segments[segment_index][0] <= timestamp <= ordered_segments[segment_index][1]:
            count += 1
    return count


def _call_activity_block_analysis(calls, *, range_end=None, rules=None):
    analysis = {
        "active_seconds": 0,
        "attempt_count": 0,
        "real_call_count": 0,
        "zero_second_attempt_count": 0,
        "suspicious_block_count": 0,
        "zero_only_block_count": 0,
        "suspicious_attempt_count": 0,
        "segments": [],
    }

    active_rules = rules or _work_review_rules()
    attempt_threshold = int(active_rules.get("attempt_threshold", MIN_REAL_CALLS_PER_ATTEMPT_BLOCK) or 1)
    blocks = _call_activity_blocks_with_stats(calls, rules=active_rules)
    for block in blocks:
        analysis["attempt_count"] += len(block["calls"])
        analysis["zero_second_attempt_count"] += block["zero_seconds_in_block"]
        analysis["real_call_count"] += block["real_calls_in_block"]

    zero_streak = []
    zero_streak_attempts = 0

    def flush_zero_streak(*, force_zero_only=False):
        nonlocal zero_streak, zero_streak_attempts
        if not zero_streak:
            return
        if force_zero_only or zero_streak_attempts >= attempt_threshold:
            analysis["zero_only_block_count"] += len(zero_streak)
        else:
            for block in zero_streak:
                cooldown_seconds = _call_activity_cooldown_seconds(
                    block["block_end"],
                    range_end=range_end,
                    rules=active_rules,
                )
                block_segment = (
                    block["block_start"],
                    block["block_end"] + timedelta(seconds=cooldown_seconds),
                )
                analysis["active_seconds"] += _segment_seconds(block_segment)
                analysis["segments"].append(block_segment)
        zero_streak = []
        zero_streak_attempts = 0

    for block in blocks:
        if block["real_calls_in_block"] <= 0:
            zero_streak.append(block)
            zero_streak_attempts += len(block["calls"])
            continue

        streak_reached = zero_streak_attempts >= attempt_threshold
        flush_zero_streak(force_zero_only=streak_reached)

        if (
            len(block["calls"]) >= attempt_threshold
            and (block["real_calls_in_block"] * attempt_threshold) < len(block["calls"])
        ):
            analysis["suspicious_block_count"] += 1
            analysis["suspicious_attempt_count"] += len(block["calls"])
            analysis["active_seconds"] += block["real_call_activity_seconds"]
            analysis["segments"].extend(block["real_call_segments"])
            continue

        cooldown_seconds = _call_activity_cooldown_seconds(block["block_end"], range_end=range_end, rules=active_rules)
        block_start = block["block_start"]
        if streak_reached and block["first_real_call_start"]:
            block_start = block["first_real_call_start"]
        block_segment = (
            block_start,
            block["block_end"] + timedelta(seconds=cooldown_seconds),
        )
        analysis["active_seconds"] += _segment_seconds(block_segment)
        analysis["segments"].append(block_segment)

    flush_zero_streak()

    return analysis


def _hourly_bonus_summary_for_day_calls(calls, *, company_profile, range_end=None):
    if not company_profile.hourly_call_bonus_enabled:
        return _empty_hourly_bonus_summary()

    threshold_per_hour = int(company_profile.hourly_call_bonus_threshold or 0)
    bonus_rate = _money(company_profile.hourly_call_bonus_rate or 0)
    if threshold_per_hour <= 0 or bonus_rate <= Decimal("0.00"):
        return _empty_hourly_bonus_summary()

    analysis = _call_activity_block_analysis(calls, range_end=range_end, rules=_work_review_rules())
    completed_hours = analysis["active_seconds"] // 3600
    if completed_hours < 1:
        return {
            **_empty_hourly_bonus_summary(),
            "completed_hours": completed_hours,
        }

    completed_segments = _truncate_segments_to_seconds(
        analysis["segments"],
        completed_hours * 3600,
    )
    completed_hour_segments = _split_segments_into_buckets(completed_segments, 3600)
    completed_bonus_calls = 0
    extra_calls = 0
    for hour_segments in completed_hour_segments:
        hour_call_count = _count_calls_in_segments(calls, hour_segments)
        completed_bonus_calls += hour_call_count
        extra_calls += max(int(hour_call_count or 0) - threshold_per_hour, 0)

    threshold_calls = completed_hours * threshold_per_hour
    return {
        "completed_hours": int(completed_hours),
        "completed_bonus_calls": int(completed_bonus_calls or 0),
        "threshold_calls": int(threshold_calls),
        "extra_calls": int(extra_calls),
        "bonus_amount": _money(Decimal(extra_calls) * bonus_rate),
    }


def _hourly_bonus_summary_map(*, start_at=None, end_at=None, staff_ids=None, company_profile=None):
    company_profile = company_profile or get_company_profile()
    call_queryset = _hourly_bonus_call_queryset()

    if start_at is not None and end_at is not None:
        call_queryset = call_queryset.filter(start_time__range=(start_at, end_at))

    if staff_ids is not None:
        staff_ids = list(staff_ids)
        call_queryset = call_queryset.filter(staff_id__in=staff_ids)

    calls_by_staff_day = defaultdict(list)
    for call in (
        call_queryset.annotate(activity_day=TruncDate("start_time"))
        .only(
            "staff_id",
            "start_time",
            "end_time",
            "duration_seconds",
            "is_verified",
            "created_at",
        )
        .order_by("staff_id", "start_time", "end_time", "id")
    ):
        calls_by_staff_day[(call.staff_id, getattr(call, "activity_day", None))].append(call)

    summaries_by_staff = defaultdict(_empty_hourly_bonus_summary)
    for (_staff_id, _activity_day), calls in calls_by_staff_day.items():
        daily_summary = _hourly_bonus_summary_for_day_calls(
            calls,
            company_profile=company_profile,
            range_end=end_at,
        )
        staff_summary = summaries_by_staff[_staff_id]
        staff_summary["completed_hours"] += daily_summary["completed_hours"]
        staff_summary["completed_bonus_calls"] += daily_summary["completed_bonus_calls"]
        staff_summary["threshold_calls"] += daily_summary["threshold_calls"]
        staff_summary["extra_calls"] += daily_summary["extra_calls"]
        staff_summary["bonus_amount"] = _money(
            staff_summary["bonus_amount"] + daily_summary["bonus_amount"]
        )

    if staff_ids is not None:
        for staff_id in staff_ids:
            summaries_by_staff.setdefault(staff_id, _empty_hourly_bonus_summary())

    return dict(summaries_by_staff)


def _staff_period_totals(start, end):
    company_profile = get_company_profile()
    session_totals = _effective_active_seconds_map(start_at=start, end_at=end)
    call_totals = {
        row["staff_id"]: row["total"] or 0
        for row in Call.objects.filter(start_time__range=(start, end), is_qualifying=True)
        .values("staff_id")
        .annotate(total=Sum("duration_seconds"))
    }
    bonus_summaries = _hourly_bonus_summary_map(
        start_at=start,
        end_at=end,
        company_profile=company_profile,
    )
    converted_counts = Counter(
        Call.objects.filter(start_time__range=(start, end), status=Call.Status.CONVERTED, is_qualifying=True)
        .values_list("staff_id", flat=True)
    )
    return session_totals, call_totals, converted_counts, bonus_summaries


def _calculate_base_pay(staff, active_hours):
    active_hours = Decimal(active_hours or 0)
    return _money(active_hours * staff.hourly_rate)


def _calculate_hourly_call_bonus(*, company_profile, active_hours, bonus_calls):
    if not company_profile.hourly_call_bonus_enabled:
        return {
            "completed_hours": 0,
            "threshold_calls": 0,
            "extra_calls": 0,
            "bonus_amount": Decimal("0.00"),
        }

    threshold_per_hour = int(company_profile.hourly_call_bonus_threshold or 0)
    bonus_rate = _money(company_profile.hourly_call_bonus_rate or 0)
    completed_hours = int(Decimal(active_hours or 0).to_integral_value(rounding=ROUND_FLOOR))
    if (
        threshold_per_hour <= 0
        or bonus_rate <= Decimal("0.00")
        or active_hours <= Decimal("0.00")
        or completed_hours < 1
    ):
        return {
            "completed_hours": max(completed_hours, 0),
            "threshold_calls": 0,
            "extra_calls": 0,
            "bonus_amount": Decimal("0.00"),
        }

    threshold_calls = completed_hours * threshold_per_hour
    extra_calls = max(int(bonus_calls or 0) - threshold_calls, 0)
    bonus_amount = _money(Decimal(extra_calls) * bonus_rate)
    return {
        "completed_hours": completed_hours,
        "threshold_calls": threshold_calls,
        "extra_calls": extra_calls,
        "bonus_amount": bonus_amount,
    }


def calculate_staff_payout(
    staff,
    *,
    active_seconds=0,
    call_seconds=0,
    converted_leads=0,
    bonus_calls=0,
    company_profile=None,
    hourly_call_bonus_summary=None,
):
    company_profile = company_profile or get_company_profile()
    active_hours = _decimal_hours(active_seconds)
    call_minutes = _decimal_minutes(call_seconds)
    call_earnings = Decimal("0.00")
    conversion_bonus = _money(Decimal(str(converted_leads or 0)) * staff.bonus_per_conversion)
    hourly_call_bonus = hourly_call_bonus_summary or _calculate_hourly_call_bonus(
        company_profile=company_profile,
        active_hours=active_hours,
        bonus_calls=bonus_calls,
    )
    bonus_earnings = _money(conversion_bonus + hourly_call_bonus["bonus_amount"])
    base_pay = _calculate_base_pay(staff, active_hours)
    total_pay = _money(base_pay + bonus_earnings)
    return {
        "active_seconds": int(active_seconds or 0),
        "active_hours": active_hours,
        "call_seconds": int(call_seconds or 0),
        "call_minutes": call_minutes,
        "converted_leads": int(converted_leads or 0),
        "bonus_calls": int(hourly_call_bonus.get("completed_bonus_calls", bonus_calls or 0)),
        "base_pay": base_pay,
        "call_earnings": call_earnings,
        "conversion_bonus": conversion_bonus,
        "hourly_call_bonus": hourly_call_bonus["bonus_amount"],
        "hourly_call_bonus_completed_hours": hourly_call_bonus["completed_hours"],
        "hourly_call_bonus_threshold_calls": hourly_call_bonus["threshold_calls"],
        "hourly_call_bonus_extra_calls": hourly_call_bonus["extra_calls"],
        "bonus_earnings": bonus_earnings,
        "total_pay": total_pay,
    }


def _current_cycle_payout(staff, weekly_breakdown, monthly_breakdown):
    if staff.compensation_type == Staff.CompensationType.WEEKLY:
        return weekly_breakdown["total_pay"], "Weekly Balance"
    if staff.compensation_type == Staff.CompensationType.MONTHLY:
        return monthly_breakdown["total_pay"], "Monthly Balance"
    return monthly_breakdown["total_pay"], "Hourly Balance"


def _quantized_decimal(value):
    return Decimal(value or 0).quantize(TWOPLACES)


WEEKLY_PAYOUT_DAY_INDEX = {
    Staff.WeeklyPayoutDay.MONDAY: 0,
    Staff.WeeklyPayoutDay.TUESDAY: 1,
    Staff.WeeklyPayoutDay.WEDNESDAY: 2,
    Staff.WeeklyPayoutDay.THURSDAY: 3,
    Staff.WeeklyPayoutDay.FRIDAY: 4,
    Staff.WeeklyPayoutDay.SATURDAY: 5,
    Staff.WeeklyPayoutDay.SUNDAY: 6,
}


def _month_last_day(value):
    return date(value.year, value.month, calendar.monthrange(value.year, value.month)[1])


def _previous_month_range(value):
    previous_month_last_day = value.replace(day=1) - timedelta(days=1)
    return previous_month_last_day.replace(day=1), previous_month_last_day


def _weekly_due_period(value, payout_day):
    target_index = WEEKLY_PAYOUT_DAY_INDEX.get(
        payout_day or Staff.WeeklyPayoutDay.WEDNESDAY,
        WEEKLY_PAYOUT_DAY_INDEX[Staff.WeeklyPayoutDay.WEDNESDAY],
    )
    end = value - timedelta(days=(value.weekday() - target_index) % 7)
    start = end - timedelta(days=6)
    return start, end


def _weekly_running_period(value, payout_day):
    due_start, due_end = _weekly_due_period(value, payout_day)
    if value.weekday() == WEEKLY_PAYOUT_DAY_INDEX.get(
        payout_day or Staff.WeeklyPayoutDay.WEDNESDAY,
        WEEKLY_PAYOUT_DAY_INDEX[Staff.WeeklyPayoutDay.WEDNESDAY],
    ):
        return due_start, value
    return due_end + timedelta(days=1), value


def _monthly_due_period(value):
    current_month_end = _month_last_day(value)
    if value == current_month_end:
        return value.replace(day=1), value
    return _previous_month_range(value)


def _monthly_running_period(value):
    return value.replace(day=1), value


def _running_payout_cycle_for_staff(staff):
    if staff.compensation_type == Staff.CompensationType.WEEKLY:
        return Salary.PayoutCycle.WEEKLY
    if staff.compensation_type == Staff.CompensationType.MONTHLY:
        return Salary.PayoutCycle.MONTHLY
    return Salary.PayoutCycle.CUSTOM


def _due_period_for_staff(staff, today):
    if staff.compensation_type == Staff.CompensationType.WEEKLY:
        return _weekly_due_period(today, staff.weekly_payout_day), Salary.PayoutCycle.WEEKLY
    if staff.compensation_type == Staff.CompensationType.MONTHLY:
        return _monthly_due_period(today), Salary.PayoutCycle.MONTHLY
    return _monthly_running_period(today), Salary.PayoutCycle.CUSTOM


def _running_period_for_staff(staff, today):
    if staff.compensation_type == Staff.CompensationType.WEEKLY:
        return _weekly_running_period(today, staff.weekly_payout_day)
    return _monthly_running_period(today)


def _salary_period_snapshot(staff, *, period_start, period_end, payout_cycle):
    breakdown = calculate_staff_payout_for_dates(staff, period_start, period_end)
    record = Salary.objects.filter(
        staff=staff,
        period_start=period_start,
        period_end=period_end,
    ).first()
    paid_total = _salary_record_paid_total(record) if record else Decimal("0.00")
    earned_total = _money(breakdown["total_pay"])
    balance = max(earned_total - paid_total, Decimal("0.00"))
    conversion_reward_rows = _conversion_reward_rows_for_period(
        staff,
        period_start=period_start,
        period_end=period_end,
    )
    return {
        "period_start": period_start,
        "period_end": period_end,
        "period_label": f"{period_start.strftime('%d %b %Y')} to {period_end.strftime('%d %b %Y')}",
        "payout_cycle": payout_cycle,
        "hours_seconds": breakdown["active_seconds"],
        "hours": breakdown["active_hours"],
        "hours_label": _format_work_duration_label(breakdown["active_seconds"]),
        "earned_total": earned_total,
        "earned_total_label": _format_currency(earned_total),
        "paid_total": paid_total,
        "paid_total_label": _format_currency(paid_total),
        "balance": _money(balance),
        "balance_label": _format_currency(balance),
        "base_pay_label": _format_currency(breakdown["base_pay"]),
        "call_earnings_label": _format_currency(breakdown["call_earnings"]),
        "conversion_reward_label": _format_currency(breakdown["conversion_bonus"]),
        "hourly_call_bonus_label": _format_currency(breakdown["hourly_call_bonus"]),
        "bonus_earnings_label": _format_currency(breakdown["bonus_earnings"]),
        "converted_lead_count": int(breakdown["converted_leads"]),
        "conversion_reward_rows": conversion_reward_rows,
        "record": record,
    }


def _snapshot_with_paid_total(snapshot, paid_total):
    paid_total = _money(paid_total)
    earned_total = _money(snapshot.get("earned_total") or Decimal("0.00"))
    balance = max(earned_total - paid_total, Decimal("0.00"))
    updated_snapshot = dict(snapshot)
    updated_snapshot["paid_total"] = paid_total
    updated_snapshot["paid_total_label"] = _format_currency(paid_total)
    updated_snapshot["balance"] = _money(balance)
    updated_snapshot["balance_label"] = _format_currency(balance)
    return updated_snapshot


def _salary_schedule_label(staff):
    if staff.compensation_type == Staff.CompensationType.WEEKLY:
        return f"Every {staff.get_weekly_payout_day_display()}"
    if staff.compensation_type == Staff.CompensationType.MONTHLY:
        return "Last day of every month"
    return "Running earned amount"


def build_staff_current_salary_summary(staff, *, today=None):
    today = today or timezone.localdate()
    period_start, period_end = _running_period_for_staff(staff, today)
    breakdown = calculate_staff_payout_for_dates(staff, period_start, period_end)
    total_hours = breakdown["active_hours"]
    total_earned = breakdown["total_pay"]
    total_paid = (
        SalaryPaymentTransaction.objects.filter(salary_record__staff=staff)
        .aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))
        .get("total")
        or Decimal("0.00")
    )
    latest_transaction = (
        SalaryPaymentTransaction.objects.filter(salary_record__staff=staff)
        .order_by("-paid_at", "-created_at")
        .first()
    )
    return {
        "total_working_hours": float(total_hours),
        "total_working_hours_label": _format_work_duration_label(breakdown["active_seconds"]),
        "total_earned_amount": float(total_earned),
        "total_earned_amount_label": f"Rs. {float(total_earned):,.2f}",
        "total_paid_amount": float(total_paid),
        "total_paid_amount_label": _format_currency(total_paid),
        "latest_transaction_id": str(latest_transaction.id) if latest_transaction else "",
        "latest_paid_at_label": _format_datetime(latest_transaction.paid_at) if latest_transaction else "--",
    }


def _staff_total_active_hours(staff):
    total_seconds = _effective_active_seconds_for_staff(staff=staff)
    return _quantized_decimal(Decimal(total_seconds) / Decimal("3600"))


def _staff_active_hours_map(staff_ids):
    if not staff_ids:
        return {}
    return {
        staff_id: _quantized_decimal(Decimal(total_seconds or 0) / Decimal("3600"))
        for staff_id, total_seconds in _effective_active_seconds_map(staff_ids=staff_ids).items()
    }


def _build_referral_tracking_payload(staff, *, company_profile=None):
    company_profile = company_profile or get_company_profile()
    submissions = list(
        ReferralSubmission.objects.filter(referrer=staff)
        .select_related("joined_staff")
        .order_by("-created_at")
    )
    joined_staff_ids = [submission.joined_staff_id for submission in submissions if submission.joined_staff_id]
    active_hours_map = _staff_active_hours_map(joined_staff_ids)
    rewards = {
        reward.referred_staff_id: reward
        for reward in ReferralReward.objects.filter(
            referrer=staff,
            referred_staff_id__in=joined_staff_ids,
        ).select_related("referred_staff")
    }

    rows = []
    summary = {
        "enabled": company_profile.referral_program_enabled,
        "required_hours_label": f"{float(company_profile.referral_required_hours or 0):,.1f}h",
        "reward_amount_label": _format_currency(company_profile.referral_reward_amount or 0),
        "referred_by_name": staff.referred_by.name if staff.referred_by else "",
        "submitted_count": len(submissions),
        "not_joined_count": 0,
        "joined_count": 0,
        "started_working_count": 0,
        "completed_count": 0,
        "qualified_count": 0,
        "pending_count": 0,
        "paid_count": 0,
        "pending_total_label": _format_currency(0),
    }
    pending_total = Decimal("0.00")

    for submission in submissions:
        required_hours = _submission_required_hours(submission, company_profile=company_profile)
        reward_amount = _submission_reward_amount(submission, company_profile=company_profile)
        joined_staff = submission.joined_staff
        active_hours = active_hours_map.get(joined_staff.id, Decimal("0.00")) if joined_staff else Decimal("0.00")
        reward = rewards.get(joined_staff.id) if joined_staff else None
        is_completed = bool(reward) or (
            joined_staff is not None
            and required_hours <= Decimal("0.00")
        ) or (
            joined_staff is not None
            and required_hours > Decimal("0.00")
            and active_hours >= required_hours
        )

        if joined_staff is None:
            stage = "not_joined"
            stage_label = "Not Joined"
            progress_label = "Waiting to join the team."
            reward_status_label = "Reward unlocks after completion"
            reward_amount_label = "--"
            summary["not_joined_count"] += 1
        elif is_completed:
            stage = "completed"
            stage_label = "Completed"
            progress_label = (
                f"{float(active_hours):,.1f}h completed"
                if required_hours <= Decimal("0.00")
                else f"{float(active_hours):,.1f}h of {float(required_hours):,.1f}h completed"
            )
            reward_status_label = "Reward Paid" if reward and reward.is_paid else "Pending Reward"
            reward_amount_label = _format_currency(reward.reward_amount if reward else reward_amount)
            summary["completed_count"] += 1
            summary["qualified_count"] += 1
            if reward and reward.is_paid:
                summary["paid_count"] += 1
            else:
                summary["pending_count"] += 1
                pending_total += reward.reward_amount if reward else reward_amount
        elif active_hours > Decimal("0.00"):
            stage = "started_working"
            stage_label = "Started Working"
            progress_label = (
                f"{float(active_hours):,.1f}h started"
                if required_hours <= Decimal("0.00")
                else f"{float(active_hours):,.1f}h of {float(required_hours):,.1f}h completed"
            )
            reward_status_label = "Reward unlocks after completion"
            reward_amount_label = "--"
            summary["started_working_count"] += 1
        else:
            stage = "joined"
            stage_label = "Joined"
            progress_label = "Joined the team and waiting for work hours to begin."
            reward_status_label = "Reward unlocks after completion"
            reward_amount_label = "--"
            summary["joined_count"] += 1

        rows.append(
            {
                "id": str(submission.id),
                "referred_name": submission.referred_name,
                "referred_phone": submission.referred_phone,
                "status": submission.status,
                "status_label": submission.get_status_display(),
                "joined_staff_name": joined_staff.name if joined_staff else "",
                "created_at": _format_datetime(submission.created_at),
                "workflow_stage": stage,
                "workflow_stage_label": stage_label,
                "progress_label": progress_label,
                "active_hours_label": f"{float(active_hours):,.1f}h",
                "required_hours_label": f"{float(required_hours):,.1f}h",
                "reward_amount_label": reward_amount_label,
                "reward_status_label": reward_status_label,
            }
        )

    summary["pending_total_label"] = _format_currency(pending_total)
    return {"summary": summary, "rows": rows}


def _is_grandfathered_referral(staff):
    if not staff or not staff.referred_by_id:
        return False
    return ReferralSubmission.objects.filter(
        joined_staff=staff,
        referrer=staff.referred_by,
        program_enabled_at_submit=True,
    ).exists()


def _referral_submission_for_staff(staff):
    if not staff or not staff.referred_by_id:
        return None
    return (
        ReferralSubmission.objects.filter(
            joined_staff=staff,
            referrer=staff.referred_by,
        )
        .order_by("-created_at")
        .first()
    )


def _submission_required_hours(submission, *, company_profile=None):
    if submission is not None:
        return _quantized_decimal(submission.required_hours_at_submit or 0)
    company_profile = company_profile or get_company_profile()
    return _quantized_decimal(company_profile.referral_required_hours or 0)


def _submission_reward_amount(submission, *, company_profile=None):
    if submission is not None:
        return _money(submission.reward_amount_at_submit or 0)
    company_profile = company_profile or get_company_profile()
    return _money(company_profile.referral_reward_amount or 0)


def _sync_referral_reward_for_staff(staff, *, company_profile=None):
    company_profile = company_profile or get_company_profile()
    existing_reward = (
        ReferralReward.objects.select_related("referrer", "referred_staff")
        .filter(referred_staff=staff)
        .first()
    )
    if existing_reward:
        return existing_reward

    submission = _referral_submission_for_staff(staff)
    required_hours = _submission_required_hours(submission, company_profile=company_profile)
    reward_amount = _submission_reward_amount(submission, company_profile=company_profile)

    if (
        (not company_profile.referral_program_enabled and not _is_grandfathered_referral(staff))
        or not staff.referred_by_id
        or reward_amount <= Decimal("0")
    ):
        return None

    total_hours = _staff_total_active_hours(staff)
    if total_hours < required_hours:
        return None

    reward_defaults = {
        "referrer": staff.referred_by,
        "required_hours": required_hours,
        "reward_amount": reward_amount,
        "qualified_at": timezone.now(),
    }
    try:
        with transaction.atomic():
            reward, _ = ReferralReward.objects.get_or_create(
                referred_staff=staff,
                defaults=reward_defaults,
            )
    except IntegrityError:
        reward = ReferralReward.objects.select_related("referrer", "referred_staff").get(
            referred_staff=staff
        )
    return reward


def sync_referral_rewards(company_profile=None):
    company_profile = company_profile or get_company_profile()
    rewards = []
    for staff in Staff.objects.filter(role=Staff.Role.STAFF, referred_by__isnull=False).select_related("referred_by"):
        if not company_profile.referral_program_enabled and not _is_grandfathered_referral(staff):
            continue
        reward = _sync_referral_reward_for_staff(staff, company_profile=company_profile)
        if reward:
            rewards.append(reward)
    return rewards


def _date_range_bounds(period_start, period_end):
    start_at = timezone.make_aware(timezone.datetime.combine(period_start, timezone.datetime.min.time()))
    end_at = timezone.make_aware(timezone.datetime.combine(period_end, timezone.datetime.max.time()))
    return start_at, end_at


def calculate_staff_payout_for_dates(staff, period_start, period_end):
    start_at, end_at = _date_range_bounds(period_start, period_end)
    company_profile = get_company_profile()
    active_seconds = _effective_active_seconds_for_staff(
        staff=staff,
        start_at=start_at,
        end_at=end_at,
    )
    call_seconds = (
        Call.objects.filter(staff=staff, start_time__range=(start_at, end_at), is_qualifying=True)
        .aggregate(total=Sum("duration_seconds"))
        .get("total")
        or 0
    )
    converted_leads = Call.objects.filter(
        staff=staff,
        start_time__range=(start_at, end_at),
        status=Call.Status.CONVERTED,
        is_qualifying=True,
    ).count()
    hourly_call_bonus_summary = _hourly_bonus_summary_map(
        start_at=start_at,
        end_at=end_at,
        staff_ids=[staff.id],
        company_profile=company_profile,
    ).get(staff.id, _empty_hourly_bonus_summary())
    breakdown = calculate_staff_payout(
        staff,
        active_seconds=active_seconds,
        call_seconds=call_seconds,
        converted_leads=converted_leads,
        bonus_calls=hourly_call_bonus_summary["completed_bonus_calls"],
        company_profile=company_profile,
        hourly_call_bonus_summary=hourly_call_bonus_summary,
    )
    breakdown["period_start"] = period_start
    breakdown["period_end"] = period_end
    return breakdown


def _salary_history_row(record):
    return {
        "id": str(record.id),
        "period_start": record.period_start,
        "period_end": record.period_end,
        "period_label": f"{record.period_start.strftime('%d %b %Y')} to {record.period_end.strftime('%d %b %Y')}",
        "payout_cycle": record.payout_cycle,
        "payout_cycle_label": record.get_payout_cycle_display(),
        "total_hours": float(record.total_hours or 0),
        "total_hours_label": _format_work_duration_label(
            _seconds_from_decimal_hours(record.total_hours)
        ),
        "final_salary": _format_currency(record.final_salary),
        "paid_amount": _format_currency(record.paid_amount),
        "is_paid": record.is_paid,
        "paid_at": _format_datetime(record.paid_at),
        "payment_method": record.payment_method,
        "payment_method_label": record.get_payment_method_display() if record.payment_method else "Manual",
        "payment_reference": record.payment_reference or "--",
        "payment_note": record.payment_note or "--",
    }


def build_staff_salary_history_rows(staff, limit=20):
    transactions = (
        SalaryPaymentTransaction.objects.filter(salary_record__staff=staff)
        .select_related("salary_record")
        .order_by("-paid_at", "-created_at")[:limit]
    )
    rows = []
    for transaction in transactions:
        record = transaction.salary_record
        rows.append(
            {
                "id": str(transaction.id),
                "salary_record_id": str(record.id),
                "period_start": record.period_start,
                "period_end": record.period_end,
                "period_label": f"{record.period_start.strftime('%d %b %Y')} to {record.period_end.strftime('%d %b %Y')}",
                "payout_cycle": record.payout_cycle,
                "payout_cycle_label": record.get_payout_cycle_display(),
                "total_hours": float(record.total_hours or 0),
                "total_hours_label": _format_work_duration_label(
                    _seconds_from_decimal_hours(record.total_hours)
                ),
                "final_salary": _format_currency(record.final_salary),
                "final_salary_label": _format_currency(record.final_salary),
                "paid_amount": _format_currency(transaction.amount),
                "paid_amount_label": _format_currency(transaction.amount),
                "is_paid": record.is_paid,
                "paid_at_iso": transaction.paid_at.isoformat() if transaction.paid_at else None,
                "paid_at": _format_datetime(transaction.paid_at),
                "paid_at_label": _format_datetime(transaction.paid_at),
                "payment_kind": transaction.payment_kind,
                "payment_kind_label": transaction.get_payment_kind_display(),
                "payment_method": transaction.payment_method,
                "payment_method_label": transaction.get_payment_method_display()
                if transaction.payment_method
                else "Manual",
                "payment_reference": transaction.payment_reference or "--",
                "payment_note": transaction.payment_note or "--",
            }
        )
    return rows


def build_staff_monthly_salary_history_rows(staff, limit=12):
    records = list(
        Salary.objects.filter(staff=staff)
        .order_by("-period_end", "-period_start")
    )
    monthly_rows = []
    monthly_map = {}

    for record in records:
        month_key = record.period_end.strftime("%Y-%m")
        row = monthly_map.get(month_key)
        if row is None:
            row = {
                "month_key": month_key,
                "month_label": record.period_end.strftime("%b %Y"),
                "period_start": record.period_start,
                "period_end": record.period_end,
                "total_hours": Decimal("0.00"),
                "earned_total": Decimal("0.00"),
                "paid_total": Decimal("0.00"),
                "entry_count": 0,
                "paid_entry_count": 0,
                "last_paid_at": None,
            }
            monthly_map[month_key] = row
            monthly_rows.append(row)

        row["period_start"] = min(row["period_start"], record.period_start)
        row["period_end"] = max(row["period_end"], record.period_end)
        row["total_hours"] += _quantized_decimal(record.total_hours or 0)
        row["earned_total"] += _money(record.final_salary or 0)
        row["paid_total"] += _money(record.paid_amount or 0)
        row["entry_count"] += 1
        if record.is_paid:
            row["paid_entry_count"] += 1
        if record.paid_at and (row["last_paid_at"] is None or record.paid_at > row["last_paid_at"]):
            row["last_paid_at"] = record.paid_at

    formatted_rows = []
    for row in monthly_rows[: max(1, int(limit or 12))]:
        balance_total = max(row["earned_total"] - row["paid_total"], Decimal("0.00"))
        formatted_rows.append(
            {
                "month_key": row["month_key"],
                "month_label": row["month_label"],
                "period_label": (
                    f"{row['period_start'].strftime('%d %b %Y')} to {row['period_end'].strftime('%d %b %Y')}"
                ),
                "entry_count": row["entry_count"],
                "paid_entry_count": row["paid_entry_count"],
                "total_hours_label": _format_work_duration_label(
                    _seconds_from_decimal_hours(row["total_hours"])
                ),
                "earned_total_label": _format_currency(row["earned_total"]),
                "paid_total_label": _format_currency(row["paid_total"]),
                "balance_total_label": _format_currency(balance_total),
                "last_paid_at": _format_datetime(row["last_paid_at"]),
            }
        )
    return formatted_rows


def _conversion_reward_rows_for_period(staff, *, period_start, period_end):
    start_at, end_at = _date_range_bounds(period_start, period_end)
    reward_amount = _money(staff.bonus_per_conversion)
    converted_calls = (
        Call.objects.filter(
            staff=staff,
            start_time__range=(start_at, end_at),
            status=Call.Status.CONVERTED,
            is_qualifying=True,
        )
        .select_related("lead")
        .order_by("-end_time", "-start_time", "-id")
    )
    return [
        {
            "id": str(call.id),
            "lead_id": str(call.lead_id),
            "lead_name": call.lead.name or "Lead",
            "lead_phone": call.lead.phone or "--",
            "reward_amount_label": _format_currency(reward_amount),
            "converted_at_label": _format_datetime(call.end_time or call.start_time),
        }
        for call in converted_calls
    ]


def build_staff_referral_reward_rows(staff, limit=20):
    rewards = (
        ReferralReward.objects.filter(referrer=staff)
        .select_related("referred_staff")
        .order_by("-qualified_at", "-created_at")[:limit]
    )
    rows = []
    for reward in rewards:
        rows.append(
            {
                "id": str(reward.id),
                "referred_staff_name": reward.referred_staff.name,
                "referred_staff_phone": reward.referred_staff.phone,
                "required_hours_label": f"{float(reward.required_hours or 0):,.1f}h",
                "reward_amount_label": _format_currency(reward.reward_amount),
                "qualified_at_label": _format_datetime(reward.qualified_at),
                "is_paid": reward.is_paid,
                "paid_at_label": _format_datetime(reward.paid_at),
                "payment_method_label": reward.get_payment_method_display() if reward.payment_method else "Manual",
                "payment_reference": reward.payment_reference or "--",
                "payment_note": reward.payment_note or "--",
            }
        )
    return rows


def build_staff_referral_submission_rows(staff, limit=20):
    return _build_referral_tracking_payload(staff)["rows"][:limit]


def build_staff_referral_summary(staff, *, company_profile=None):
    company_profile = company_profile or get_company_profile()
    return _build_referral_tracking_payload(staff, company_profile=company_profile)["summary"]


def _staff_salary_snapshot_block(title, snapshot, *, subtitle=""):
    return {
        "title": title,
        "subtitle": subtitle,
        "period_start": snapshot["period_start"].isoformat(),
        "period_end": snapshot["period_end"].isoformat(),
        "period_label": snapshot["period_label"],
        "hours_label": snapshot["hours_label"],
        "earned_total_label": snapshot["earned_total_label"],
        "paid_total_label": snapshot["paid_total_label"],
        "balance_label": snapshot["balance_label"],
        "base_pay_label": snapshot["base_pay_label"],
        "call_earnings_label": snapshot["call_earnings_label"],
        "conversion_reward_label": snapshot["conversion_reward_label"],
        "hourly_call_bonus_label": snapshot["hourly_call_bonus_label"],
        "bonus_earnings_label": snapshot["bonus_earnings_label"],
        "converted_lead_count": snapshot["converted_lead_count"],
        "conversion_reward_rows": snapshot["conversion_reward_rows"],
    }


def _monthly_earnings_pattern_rows(staff, *, today):
    month_start = today.replace(day=1)
    month_end = today
    rows = []
    segment_start = month_start
    week_number = 1
    while segment_start <= month_end:
        segment_end = min(segment_start + timedelta(days=6), month_end)
        snapshot = _salary_period_snapshot(
            staff,
            period_start=segment_start,
            period_end=segment_end,
            payout_cycle=Salary.PayoutCycle.CUSTOM,
        )
        rows.append(
            {
                "title": f"Week {week_number}",
                "period_label": snapshot["period_label"],
                "hours_label": snapshot["hours_label"],
                "earned_total_label": snapshot["earned_total_label"],
                "paid_total_label": snapshot["paid_total_label"],
                "balance_label": snapshot["balance_label"],
            }
        )
        segment_start = segment_end + timedelta(days=1)
        week_number += 1
    return rows


def _weekly_earnings_pattern_rows(staff, *, today, count=4):
    latest_period_start, latest_period_end = _weekly_due_period(today, staff.weekly_payout_day)
    rows = []
    for index in range(count):
        period_end = latest_period_end - timedelta(days=7 * index)
        period_start = latest_period_start - timedelta(days=7 * index)
        snapshot = _salary_period_snapshot(
            staff,
            period_start=period_start,
            period_end=period_end,
            payout_cycle=Salary.PayoutCycle.WEEKLY,
        )
        rows.append(
            {
                "title": f"Week {index + 1}",
                "period_label": snapshot["period_label"],
                "hours_label": snapshot["hours_label"],
                "earned_total_label": snapshot["earned_total_label"],
                "paid_total_label": snapshot["paid_total_label"],
                "balance_label": snapshot["balance_label"],
            }
        )
    return rows


def build_staff_salary_details_payload(staff):
    today = timezone.localdate()
    company_profile = get_company_profile()
    sync_referral_rewards(company_profile)
    referral_tracking = _build_referral_tracking_payload(
        staff,
        company_profile=company_profile,
    )
    current_period_start, current_period_end = _running_period_for_staff(staff, today)
    current_snapshot = _salary_period_snapshot(
        staff,
        period_start=current_period_start,
        period_end=current_period_end,
        payout_cycle=_running_payout_cycle_for_staff(staff),
    )
    # Staff app must reflect credited advances in the running cycle
    # (admin salary views already apply this same adjustment).
    current_snapshot = _snapshot_with_paid_total(
        current_snapshot,
        _running_cycle_advance_paid_total(
            staff,
            period_start=current_period_start,
            period_end=current_period_end,
            payout_cycle=_running_payout_cycle_for_staff(staff),
        ),
    )
    previous_month_start, previous_month_end = _previous_month_range(today)
    previous_month_snapshot = _salary_period_snapshot(
        staff,
        period_start=previous_month_start,
        period_end=previous_month_end,
        payout_cycle=Salary.PayoutCycle.MONTHLY,
    )

    if staff.compensation_type == Staff.CompensationType.WEEKLY:
        pattern_title = "Recent weekly earnings"
        pattern_subtitle = (
            f"These payout blocks close every {staff.get_weekly_payout_day_display()}."
        )
        pattern_rows = _weekly_earnings_pattern_rows(staff, today=today)
    else:
        pattern_title = "Monthly earning pattern"
        pattern_subtitle = "See how the current month is building week by week."
        pattern_rows = _monthly_earnings_pattern_rows(staff, today=today)

    current_title = {
        Staff.CompensationType.MONTHLY: "Current month earnings",
        Staff.CompensationType.WEEKLY: "Current week earnings",
        Staff.CompensationType.HOURLY: "Current earned amount",
    }.get(staff.compensation_type, "Current earnings")

    target_hours_value = (
        staff.target_hours_per_month
        if staff.compensation_type == Staff.CompensationType.MONTHLY
        else staff.target_hours_per_week
    )
    target_hours_unit = "month" if staff.compensation_type == Staff.CompensationType.MONTHLY else "week"

    return {
        "summary": {
            "compensation_type": staff.compensation_type,
            "compensation_type_label": staff.get_compensation_type_display(),
            "payout_schedule_label": _salary_schedule_label(staff),
            "weekly_payout_day_label": staff.get_weekly_payout_day_display() if staff.weekly_payout_day else "",
            "hourly_rate_label": _format_currency(staff.hourly_rate),
            "call_rate_label": _format_currency(staff.call_rate),
            "bonus_per_conversion_label": _format_currency(staff.bonus_per_conversion),
            "target_hours_label": f"{float(target_hours_value or 0):,.1f}h per {target_hours_unit}",
        },
        "current_cycle": _staff_salary_snapshot_block(
            current_title,
            current_snapshot,
            subtitle="This is the earning progress for the running payout cycle.",
        ),
        "previous_month": _staff_salary_snapshot_block(
            "Previous month earnings",
            previous_month_snapshot,
            subtitle="Review the full earning and payment position from the last completed month.",
        ),
        "pattern": {
            "title": pattern_title,
            "subtitle": pattern_subtitle,
            "rows": pattern_rows,
        },
        "payment_history": build_staff_salary_history_rows(staff, limit=20),
        "referral": {
            **referral_tracking["summary"],
            "tracking_rows": referral_tracking["rows"][:20],
            "history": build_staff_referral_reward_rows(staff, limit=20),
        },
    }


def build_recent_salary_payment_rows(limit=40):
    transactions = (
        SalaryPaymentTransaction.objects.select_related("salary_record", "salary_record__staff")
        .order_by("-paid_at", "-created_at")[:limit]
    )
    rows = []
    for transaction in transactions:
        try:
            record = transaction.salary_record
            rows.append(
                {
                    "id": str(transaction.id),
                    "staff_id": str(record.staff_id),
                    "staff_name": record.staff.name,
                    "staff_phone": record.staff.phone,
                    "period_label": f"{record.period_start.strftime('%d %b %Y')} to {record.period_end.strftime('%d %b %Y')}",
                    "payout_cycle_label": record.get_payout_cycle_display(),
                    "paid_amount": _format_currency(transaction.amount),
                    "paid_at": _format_datetime(transaction.paid_at),
                    "payment_kind_label": transaction.get_payment_kind_display(),
                    "payment_method_label": transaction.get_payment_method_display()
                    if transaction.payment_method
                    else "Manual",
                    "payment_reference": transaction.payment_reference or "--",
                    "payment_note": transaction.payment_note or "--",
                    "final_salary": _format_currency(record.final_salary),
                    "total_hours_label": _format_work_duration_label(
                        _seconds_from_decimal_hours(record.total_hours)
                    ),
                }
            )
        except Exception:
            logger.warning(
                "Skipping malformed salary payment transaction row",
                extra={"transaction_id": str(getattr(transaction, "id", ""))},
                exc_info=True,
            )
    return rows


def _salary_record_paid_total(record):
    total = record.payment_transactions.aggregate(
        total=Coalesce(Sum("amount"), Decimal("0.00"))
    ).get("total")
    return _money(total)


def _running_cycle_advance_paid_total(staff, *, period_start, period_end, payout_cycle):
    total = (
        SalaryPaymentTransaction.objects.filter(
            salary_record__staff=staff,
            salary_record__payout_cycle=payout_cycle,
            salary_record__period_start=period_start,
            salary_record__period_end__gte=period_start,
            salary_record__period_end__lte=period_end,
            payment_kind=SalaryPaymentTransaction.PaymentKind.ADVANCE,
        ).aggregate(total=Coalesce(Sum("amount"), Decimal("0.00"))).get("total")
        or Decimal("0.00")
    )
    return _money(total)


def _salary_record_remaining_balance(record):
    remaining = _money(record.final_salary) - _salary_record_paid_total(record)
    if remaining < Decimal("0.00"):
        return Decimal("0.00")
    return _money(remaining)


def _refresh_salary_record_payment_state(record):
    totals = record.payment_transactions.aggregate(
        total=Coalesce(Sum("amount"), Decimal("0.00")),
        latest_paid_at=Max("paid_at"),
    )
    latest_transaction = record.payment_transactions.order_by("-paid_at", "-created_at").first()
    total_paid = _money(totals.get("total") or Decimal("0.00"))
    final_salary = _money(record.final_salary)

    record.paid_amount = total_paid
    record.is_paid = bool(final_salary > Decimal("0.00") and total_paid >= final_salary)
    record.paid_at = totals.get("latest_paid_at")
    if latest_transaction:
        record.payment_method = latest_transaction.payment_method
        record.payment_reference = latest_transaction.payment_reference.strip()
        record.payment_note = latest_transaction.payment_note.strip()
    else:
        record.payment_method = ""
        record.payment_reference = ""
        record.payment_note = ""
    record.save(
        update_fields=[
            "paid_amount",
            "is_paid",
            "paid_at",
            "payment_method",
            "payment_reference",
            "payment_note",
            "updated_at",
        ]
    )
    return record


def _payroll_email_is_ready():
    if not settings.PAYROLL_NOTIFY_EMAILS:
        return False, "Payroll email notifications are disabled."
    if not settings.EMAIL_HOST:
        return False, "SMTP is not configured yet."
    if settings.EMAIL_BACKEND == "django.core.mail.backends.console.EmailBackend":
        return False, "SMTP email backend is not configured yet."
    return True, ""


def send_salary_payment_acknowledgement(salary_record):
    staff = salary_record.staff
    if not staff.email:
        return {"sent": False, "message": f"{staff.name} does not have an email address saved."}

    email_ready, reason = _payroll_email_is_ready()
    if not email_ready:
        return {"sent": False, "message": reason}

    company_profile = get_company_profile()
    period_label = f"{salary_record.period_start.strftime('%d %b %Y')} to {salary_record.period_end.strftime('%d %b %Y')}"
    paid_at = timezone.localtime(salary_record.paid_at) if salary_record.paid_at else timezone.localtime(timezone.now())
    context = {
        "company_name": company_profile.company_name or "Heavenection",
        "company_support_email": company_profile.company_email or settings.DEFAULT_FROM_EMAIL,
        "company_phone": company_profile.company_phone or company_profile.support_phone or "",
        "staff_name": staff.name,
        "paid_amount": _format_currency(salary_record.paid_amount),
        "final_salary": _format_currency(salary_record.final_salary),
        "base_pay": _format_currency(salary_record.base_pay),
        "call_earnings": _format_currency(salary_record.call_earnings),
        "bonus_earnings": _format_currency(salary_record.bonus_earnings),
        "payment_method": salary_record.get_payment_method_display() if salary_record.payment_method else "Recorded Payment",
        "payment_reference": salary_record.payment_reference or "--",
        "payment_note": salary_record.payment_note or "",
        "payout_cycle": salary_record.get_payout_cycle_display(),
        "period_label": period_label,
        "paid_at": paid_at.strftime("%d %b %Y, %I:%M %p"),
        "total_hours": f"{round(float(salary_record.total_hours or 0), 2)}",
    }
    subject = f"{context['company_name']} Salary Credited - {context['period_label']}"
    from_email = settings.DEFAULT_FROM_EMAIL
    if company_profile.company_name and settings.DEFAULT_FROM_EMAIL:
        from_email = f"{company_profile.company_name} <{settings.DEFAULT_FROM_EMAIL}>"

    text_body = render_to_string("emails/salary_paid_notification.txt", context).strip()
    html_body = render_to_string("emails/salary_paid_notification.html", context)
    reply_to = [company_profile.company_email] if company_profile.company_email else None

    try:
        message = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=from_email,
            to=[staff.email],
            reply_to=reply_to,
        )
        message.attach_alternative(html_body, "text/html")
        message.send(fail_silently=False)
    except Exception as error:  # pragma: no cover - depends on SMTP runtime
        logger.exception("Salary payment acknowledgement email failed for staff %s", staff.id)
        return {
            "sent": False,
            "message": f"Salary was marked paid, but the email could not be sent: {error}",
        }

    return {
        "sent": True,
        "message": f"Salary acknowledgement email sent to {staff.email}.",
    }


def queue_salary_payment_acknowledgement(salary_record):
    staff = salary_record.staff
    if not staff.email:
        return {"queued": False, "message": f"{staff.name} does not have an email address saved."}

    email_ready, reason = _payroll_email_is_ready()
    if not email_ready:
        return {"queued": False, "message": reason}

    salary_record_id = salary_record.id
    staff_id = staff.id
    staff_email = staff.email

    def _send_in_background():
        close_old_connections()
        try:
            queued_record = Salary.objects.select_related("staff").get(id=salary_record_id)
            send_salary_payment_acknowledgement(queued_record)
        except Exception:  # pragma: no cover - depends on runtime thread scheduling
            logger.exception("Queued salary payment acknowledgement email failed for staff %s", staff_id)
        finally:
            close_old_connections()

    Thread(
        target=_send_in_background,
        name=f"salary-ack-{salary_record_id}",
        daemon=True,
    ).start()

    return {
        "queued": True,
        "message": f"Salary acknowledgement will be sent to {staff_email} shortly.",
    }


def record_referral_reward_payment(
    reward,
    *,
    payment_method="",
    payment_reference="",
    payment_note="",
):
    if reward.is_paid:
        raise ValidationError({"reward": ["This referral reward is already paid."]})

    reward.is_paid = True
    reward.paid_at = timezone.now()
    reward.payment_method = payment_method or ""
    reward.payment_reference = payment_reference or ""
    reward.payment_note = payment_note or ""
    reward.save(
        update_fields=[
            "is_paid",
            "paid_at",
            "payment_method",
            "payment_reference",
            "payment_note",
            "updated_at",
        ]
    )
    return reward


def record_staff_salary_payment(
    staff,
    *,
    payout_cycle,
    period_start,
    period_end,
    paid_amount,
    payment_kind=SalaryPaymentTransaction.PaymentKind.SALARY,
    payment_method="",
    payment_reference="",
    payment_note="",
):
    breakdown = calculate_staff_payout_for_dates(staff, period_start, period_end)
    recommended_amount = _money(breakdown["total_pay"])
    salary_record, created = Salary.objects.update_or_create(
        staff=staff,
        period_start=period_start,
        period_end=period_end,
        defaults={
            "payout_cycle": payout_cycle,
            "total_hours": _quantized_decimal(breakdown["active_hours"]),
            "total_call_minutes": _quantized_decimal(breakdown["call_minutes"]),
            "converted_leads": int(breakdown["converted_leads"]),
            "base_pay": _money(breakdown["base_pay"]),
            "call_earnings": _money(breakdown["call_earnings"]),
            "bonus_earnings": _money(breakdown["bonus_earnings"]),
            "incentives": _money(breakdown["bonus_earnings"]),
            "final_salary": recommended_amount,
            "paid_amount": Decimal("0.00"),
            "is_paid": False,
            "paid_at": None,
            "payment_method": "",
            "payment_reference": "",
            "payment_note": "",
        },
    )
    current_paid_total = _salary_record_paid_total(salary_record)
    remaining_balance = _salary_record_remaining_balance(salary_record)
    if recommended_amount <= Decimal("0.00"):
        raise ValidationError(
            {"paid_amount": "No earnings are available for this salary period yet."}
        )

    paid_amount_value = _money(
        paid_amount if paid_amount is not None else remaining_balance
    )
    if paid_amount_value <= Decimal("0.00"):
        raise ValidationError({"paid_amount": "Enter a credited amount greater than zero."})
    if remaining_balance <= Decimal("0.00"):
        raise ValidationError({"paid_amount": "This salary period is already fully paid."})
    if paid_amount_value > remaining_balance:
        raise ValidationError(
            {
                "paid_amount": (
                    f"Only the remaining balance of Rs. {float(remaining_balance):,.2f} "
                    "can be credited for this salary period."
                )
            }
        )

    transaction = SalaryPaymentTransaction.objects.create(
        salary_record=salary_record,
        amount=paid_amount_value,
        payment_kind=payment_kind,
        payment_method=payment_method,
        payment_reference=payment_reference.strip(),
        payment_note=payment_note.strip(),
        paid_at=timezone.now(),
    )
    salary_record = _refresh_salary_record_payment_state(salary_record)
    return salary_record, transaction, created


def delete_salary_payment_transaction(transaction):
    salary_record = transaction.salary_record
    transaction.delete()
    return _refresh_salary_record_payment_state(salary_record)


def _salary_setting_target_label(staff):
    hourly_label = f"Rs. {float(staff.hourly_rate):,.2f} / hour"
    if staff.compensation_type == Staff.CompensationType.WEEKLY:
        return f"{hourly_label} | Paid Weekly"
    if staff.compensation_type == Staff.CompensationType.MONTHLY:
        return f"{hourly_label} | Paid Monthly"
    return f"{hourly_label} | Running Hourly"


def _payout_cycle_label(staff):
    if staff.compensation_type == Staff.CompensationType.WEEKLY:
        return "Weekly"
    if staff.compensation_type == Staff.CompensationType.MONTHLY:
        return "Monthly"
    return "Hourly"


def _bounded_elapsed_seconds(previous, current, max_seconds=ONLINE_WINDOW_SECONDS):
    if not previous or not current:
        return 0
    return max(0, min(int((current - previous).total_seconds()), max_seconds))


def _dedupe_fields(field_names):
    return list(dict.fromkeys(field_names + ["updated_at"]))


def _log_staff_action(
    staff,
    action_type,
    *,
    session=None,
    call=None,
    lead=None,
    app_state=None,
    metadata=None,
):
    return StaffAction.objects.create(
        staff=staff,
        session=session,
        call=call,
        lead=lead,
        action_type=action_type,
        app_state=app_state,
        metadata=metadata or {},
    )


def _session_has_recent_verified_activity(session, now):
    if not session or not session.last_verified_call_at:
        return False
    return (now - session.last_verified_call_at).total_seconds() < VERIFIED_CALL_ACTIVITY_TIMEOUT_SECONDS


def _active_elapsed_until(session, now, *, has_live_customer_call=False):
    if session.last_known_state != Session.AppState.FOREGROUND:
        return 0

    if not _session_has_recent_verified_activity(session, now):
        return 0

    active_until = now
    if session.last_verified_call_at:
        verified_cutoff = session.last_verified_call_at + timedelta(
            seconds=VERIFIED_CALL_ACTIVITY_TIMEOUT_SECONDS
        )
        if verified_cutoff < active_until:
            active_until = verified_cutoff

    return _bounded_elapsed_seconds(session.last_heartbeat_at, active_until)


def _mark_session_foreground(session, now, *, metadata=None):
    previous_state = session.last_known_state
    session.last_interaction_at = now
    session.last_heartbeat_at = now

    update_fields = ["last_interaction_at", "last_heartbeat_at"]
    if previous_state != Session.AppState.FOREGROUND:
        session.last_known_state = Session.AppState.FOREGROUND
        session.state_changed_at = now
        session.warning_started_at = None
        update_fields.extend(["last_known_state", "state_changed_at", "warning_started_at"])

    session.save(update_fields=_dedupe_fields(update_fields))

    if previous_state == Session.AppState.WARNING:
        _log_staff_action(
            session.staff,
            StaffAction.ActionType.IDLE_WARNING_ACKNOWLEDGED,
            session=session,
            app_state=Session.AppState.FOREGROUND,
            metadata=metadata or {},
        )
    elif previous_state == Session.AppState.OFFLINE:
        _log_staff_action(
            session.staff,
            StaffAction.ActionType.RETURNED_ONLINE,
            session=session,
            app_state=Session.AppState.FOREGROUND,
            metadata=metadata or {},
        )
    elif previous_state == Session.AppState.BACKGROUND:
        _log_staff_action(
            session.staff,
            StaffAction.ActionType.APP_FOREGROUNDED,
            session=session,
            app_state=Session.AppState.FOREGROUND,
            metadata=metadata or {},
        )

    mark_staff_seen(session.staff, now)
    return session


def _credit_call_duration_to_session(
    session,
    call_end_time,
    call_duration_seconds,
    *,
    metadata=None,
    mark_verified=False,
):
    if not session or not session.is_open:
        return session

    session = _mark_session_foreground(session, call_end_time, metadata=metadata)
    qualifying_verified_call = mark_verified and int(call_duration_seconds or 0) >= SHORT_CALL_SECONDS
    update_fields = []
    if call_duration_seconds:
        session.active_seconds += max(0, int(call_duration_seconds))
        update_fields.append("active_seconds")
    if qualifying_verified_call and session.last_verified_call_at != call_end_time:
        session.last_verified_call_at = call_end_time
        update_fields.append("last_verified_call_at")
    if update_fields:
        session.save(update_fields=_dedupe_fields(update_fields))
    return session


def _resolve_requested_state(session, requested_state, now, interaction, *, has_live_customer_call=False):
    if requested_state == Session.AppState.BACKGROUND:
        return Session.AppState.BACKGROUND
    if requested_state == Session.AppState.WARNING:
        return Session.AppState.WARNING
    if requested_state == Session.AppState.OFFLINE:
        return Session.AppState.OFFLINE

    if has_live_customer_call:
        return Session.AppState.FOREGROUND
    if not _session_has_recent_verified_activity(session, now):
        return Session.AppState.OFFLINE

    last_interaction_at = now if interaction else (session.last_interaction_at or now)
    idle_seconds = max(0, int((now - last_interaction_at).total_seconds()))

    if session.last_known_state in {Session.AppState.WARNING, Session.AppState.OFFLINE} and not interaction:
        return session.last_known_state
    if idle_seconds >= IDLE_OFFLINE_AFTER_SECONDS:
        return Session.AppState.OFFLINE
    if idle_seconds >= IDLE_WARNING_AFTER_SECONDS:
        return Session.AppState.WARNING
    return Session.AppState.FOREGROUND


def _close_session(session, now, *, close_reason, auto_generated=False):
    if not session or not session.is_open:
        return session

    session.active_seconds += _active_elapsed_until(session, now)
    session.logout_time = now
    session.is_open = False
    session.last_heartbeat_at = now
    session.close_reason = close_reason
    session.save(
        update_fields=_dedupe_fields(
            [
                "active_seconds",
                "logout_time",
                "is_open",
                "last_heartbeat_at",
                "close_reason",
            ]
        )
    )
    mark_staff_seen(session.staff, now)
    _log_staff_action(
        session.staff,
        StaffAction.ActionType.SESSION_AUTO_ENDED if auto_generated else StaffAction.ActionType.SESSION_ENDED,
        session=session,
        app_state=session.last_known_state,
        metadata={"close_reason": close_reason},
    )
    return session


def reconcile_session(session, now=None):
    if not session or not session.is_open:
        return session

    now = now or timezone.now()
    state_anchor = session.state_changed_at or session.last_heartbeat_at or session.login_time
    warning_anchor = session.warning_started_at or state_anchor
    has_live_customer_call = session.staff_id in _reconcile_open_calls(staff=session.staff, now=now)

    if (
        session.last_known_state == Session.AppState.FOREGROUND
        and not has_live_customer_call
        and not _session_has_recent_verified_activity(session, now)
    ):
        session.last_known_state = Session.AppState.OFFLINE
        session.state_changed_at = now
        session.warning_started_at = None
        session.save(
            update_fields=_dedupe_fields(
                ["last_known_state", "state_changed_at", "warning_started_at"]
            )
        )
        _log_staff_action(
            session.staff,
            StaffAction.ActionType.MARKED_OFFLINE,
            session=session,
            app_state=Session.AppState.OFFLINE,
            metadata={"reason": "verified_call_timeout"},
        )
        return session

    if (
        session.last_known_state == Session.AppState.BACKGROUND
        and state_anchor
        and (now - state_anchor).total_seconds() >= BACKGROUND_TIMEOUT_SECONDS
    ):
        session.last_known_state = Session.AppState.OFFLINE
        session.state_changed_at = now
        session.warning_started_at = None
        session.save(
            update_fields=_dedupe_fields(
                ["last_known_state", "state_changed_at", "warning_started_at"]
            )
        )
        _log_staff_action(
            session.staff,
            StaffAction.ActionType.MARKED_OFFLINE,
            session=session,
            app_state=Session.AppState.OFFLINE,
            metadata={"reason": "background_timeout"},
        )
        return session

    if (
        session.last_known_state == Session.AppState.WARNING
        and warning_anchor
        and (now - warning_anchor).total_seconds() >= IDLE_WARNING_GRACE_SECONDS
    ):
        session.last_known_state = Session.AppState.OFFLINE
        session.state_changed_at = now
        session.save(update_fields=_dedupe_fields(["last_known_state", "state_changed_at"]))
        _log_staff_action(
            session.staff,
            StaffAction.ActionType.MARKED_OFFLINE,
            session=session,
            app_state=Session.AppState.OFFLINE,
            metadata={"reason": "warning_timeout"},
        )

    return session


def authenticate_staff(identifier, password, required_role=None):
    identifier = identifier.strip()
    queryset = Staff.objects.filter(is_active=True).filter(
        Q(phone=identifier) | Q(email__iexact=identifier)
    )
    if required_role:
        queryset = queryset.filter(role=required_role)

    staff = queryset.first()
    if not staff or not staff.check_password(password):
        return None
    return staff


def mark_staff_seen(staff, seen_at=None):
    staff.last_seen_at = seen_at or timezone.now()
    staff.save(update_fields=["last_seen_at", "updated_at"])


def get_open_session(staff, reconcile=False):
    session = Session.objects.filter(staff=staff, is_open=True).order_by("-login_time").first()
    if reconcile and session:
        session = reconcile_session(session)
        if session and not session.is_open:
            return None
    return session


def _touch_session_interaction(session, now, *, metadata=None):
    if not session or not session.is_open:
        return session

    has_live_customer_call = session.staff_id in _reconcile_open_calls(staff=session.staff, now=now)
    session.active_seconds += _active_elapsed_until(
        session,
        now,
        has_live_customer_call=has_live_customer_call,
    )
    session.save(update_fields=_dedupe_fields(["active_seconds"]))
    return _mark_session_foreground(session, now, metadata=metadata)


def _session_status_label(open_session, latest_session=None):
    if open_session:
        if open_session.last_known_state == Session.AppState.OFFLINE and not open_session.last_verified_call_at:
            return "Call a customer to begin working"
        return {
            Session.AppState.FOREGROUND: "Working",
            Session.AppState.BACKGROUND: "Away from app",
            Session.AppState.WARNING: "Warning shown",
            Session.AppState.OFFLINE: "Offline",
        }.get(open_session.last_known_state, "Working")

    if latest_session and latest_session.close_reason == "background_timeout":
        return "Stopped after background timeout"
    return "Stopped"


def _staff_online_label(session, active_cutoff, *, is_in_customer_call=False):
    if not session:
        return "On Call" if is_in_customer_call else "Offline"
    if is_in_customer_call:
        return "On Call"
    if session.last_known_state == Session.AppState.FOREGROUND and session.last_heartbeat_at and session.last_heartbeat_at >= active_cutoff:
        return "Online"
    if session.last_known_state == Session.AppState.WARNING:
        return "Warning"
    if session.last_known_state == Session.AppState.BACKGROUND:
        return "Away"
    return "Offline"


def _return_lead_to_queue_after_invalid_short(lead):
    if not lead:
        return None

    lead.status = Lead.Status.NEW
    lead.assigned_to = None
    lead.callback_window = ""
    lead.callback_date = None
    lead.save(
        update_fields=[
            "status",
            "assigned_to",
            "callback_window",
            "callback_date",
            "updated_at",
        ]
    )
    return lead


def _return_lead_to_same_staff_after_invalid_short(lead, staff):
    if not lead or not staff:
        return None

    lead.status = Lead.Status.NEW
    lead.assigned_to = staff
    lead.callback_window = ""
    lead.callback_date = None
    lead.save(
        update_fields=[
            "status",
            "assigned_to",
            "callback_window",
            "callback_date",
            "updated_at",
        ]
    )
    return lead


def _close_unresolved_call(call, *, reason):
    if not call or call.end_time is not None:
        return call

    session = get_open_session(call.staff, reconcile=False)
    call.end_time = call.start_time
    call.duration_seconds = 0
    call.is_qualifying = False
    call.is_verified = False
    call.status = Call.Status.INVALID_SHORT
    call.callback_window = ""
    call.callback_date = None
    call.verification_source = ""
    call.save(
        update_fields=[
            "end_time",
            "duration_seconds",
            "is_qualifying",
            "is_verified",
            "status",
            "callback_window",
            "callback_date",
            "verification_source",
            "updated_at",
        ]
    )

    call.lead.last_contacted_at = call.start_time
    call.lead.save(update_fields=["last_contacted_at", "updated_at"])
    _return_lead_to_queue_after_invalid_short(call.lead)
    auto_allocate_leads()

    _log_staff_action(
        call.staff,
        StaffAction.ActionType.CALL_ENDED,
        session=session,
        call=call,
        lead=call.lead,
        app_state=session.last_known_state if session else None,
        metadata={
            "duration_seconds": 0,
            "is_qualifying": False,
            "status": call.status,
            "source": reason,
            "ended_at": call.end_time.isoformat(),
        },
    )
    return call


def _reconcile_open_calls(*, staff=None, now=None):
    now = now or timezone.now()
    queryset = Call.objects.filter(end_time__isnull=True).select_related("staff", "lead")
    if staff is not None:
        queryset = queryset.filter(staff=staff)

    open_calls = queryset.order_by("staff_id", "-start_time", "-created_at")
    live_staff_ids = set()
    seen_staff_ids = set()

    for call in open_calls:
        if call.staff_id in seen_staff_ids:
            _close_unresolved_call(call, reason="superseded_open_call")
            continue

        seen_staff_ids.add(call.staff_id)
        call_age_seconds = max(0, int((now - call.start_time).total_seconds()))
        if call_age_seconds >= LIVE_CALL_STALE_SECONDS:
            _close_unresolved_call(call, reason="stale_open_call")
            continue

        live_staff_ids.add(call.staff_id)

    return live_staff_ids


def get_recoverable_open_call(staff, *, now=None):
    _reconcile_open_calls(staff=staff, now=now)
    return (
        Call.objects.filter(
            staff=staff,
            status=Call.Status.STARTED,
            end_time__isnull=True,
        )
        .select_related("lead")
        .order_by("-start_time", "-created_at")
        .first()
    )


def _staff_queryset():
    return Staff.objects.filter(role=Staff.Role.STAFF).order_by("name")


def _normalize_phone(phone_value):
    digits = re.sub(r"\D+", "", str(phone_value or "")).strip()
    if len(digits) > 10:
        digits = digits[-10:]
    return digits


def _allocatable_lead_ids(candidate_lead_ids):
    valid_ids = [lead_id for lead_id in (candidate_lead_ids or []) if lead_id]
    if not valid_ids:
        return set()

    candidate_rows = list(
        Lead.objects.filter(id__in=valid_ids)
        .values("id", "phone", "created_at")
    )
    if not candidate_rows:
        return set()

    candidate_phone_map = {}
    phone_values = set()
    for row in candidate_rows:
        normalized_phone = _normalize_phone(row["phone"])
        if not normalized_phone:
            continue
        candidate_phone_map[row["id"]] = normalized_phone
        phone_values.add(normalized_phone)

    if not phone_values:
        return set()

    canonical_ids_by_phone = {}
    global_rows = (
        Lead.objects.filter(phone__in=phone_values)
        .order_by("phone", "created_at", "id")
        .values("id", "phone")
    )
    for row in global_rows:
        normalized_phone = _normalize_phone(row["phone"])
        if normalized_phone and normalized_phone not in canonical_ids_by_phone:
            canonical_ids_by_phone[normalized_phone] = row["id"]

    return {
        lead_id
        for lead_id, normalized_phone in candidate_phone_map.items()
        if canonical_ids_by_phone.get(normalized_phone) == lead_id
    }


def _normalize_column_name(value):
    text = re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower())
    return " ".join(text.split())


def _lead_queue_queryset():
    return Lead.objects.filter(status__in=ACTIVE_QUEUE_STATUSES)


def _is_active_queue_status(status):
    return status in ACTIVE_QUEUE_STATUSES


def _follow_up_queryset():
    return Lead.objects.filter(
        Q(status__in=FOLLOW_UP_STATUSES) | Q(followup_moved_back_at__isnull=False)
    )


def _followup_expiry_settings(*, company_profile=None):
    profile = company_profile or get_company_profile()
    gate_mode = str(
        getattr(
            profile,
            "followup_sla_gate_mode",
            CompanyProfile.FollowupSlaGateMode.ALLOW_NORMAL_CALLS,
        )
        or CompanyProfile.FollowupSlaGateMode.ALLOW_NORMAL_CALLS
    ).strip()
    valid_gate_modes = {choice for choice, _label in CompanyProfile.FollowupSlaGateMode.choices}
    if gate_mode not in valid_gate_modes:
        gate_mode = CompanyProfile.FollowupSlaGateMode.ALLOW_NORMAL_CALLS
    return {
        "enabled": bool(getattr(profile, "followup_auto_expire_enabled", True)),
        "expiry_days": max(
            1,
            int(getattr(profile, "followup_auto_expire_days", FOLLOWUP_STALE_EXPIRY_DAYS) or FOLLOWUP_STALE_EXPIRY_DAYS),
        ),
        "warning_days": max(
            1,
            int(
                getattr(profile, "followup_staff_warning_days", FOLLOWUP_STAFF_WARNING_DAYS)
                or FOLLOWUP_STAFF_WARNING_DAYS
            ),
        ),
        "sla_gate_enabled": bool(getattr(profile, "followup_sla_gate_enabled", False)),
        "sla_gate_mode": gate_mode,
        "uncalled_alert_enabled": bool(getattr(profile, "followup_uncalled_alert_enabled", True)),
        "uncalled_alert_hours": max(
            1,
            int(
                getattr(profile, "followup_uncalled_alert_hours", FOLLOWUP_UNCALLED_ALERT_HOURS)
                or FOLLOWUP_UNCALLED_ALERT_HOURS
            ),
        ),
    }


def expire_stale_followups(*, now=None, expiry_days=None, enabled=None, company_profile=None):
    now = now or timezone.now()
    settings = _followup_expiry_settings(company_profile=company_profile)
    is_enabled = settings["enabled"] if enabled is None else bool(enabled)
    resolved_expiry_days = settings["expiry_days"] if expiry_days is None else max(1, int(expiry_days))

    if not is_enabled:
        return {
            "expired_count": 0,
            "cutoff": None,
            "enabled": False,
            "expiry_days": resolved_expiry_days,
        }

    cutoff = timezone.localtime(now - timedelta(days=resolved_expiry_days))
    stale_followups = list(
        Lead.objects.select_related("assigned_to", "interested_detail__staff").filter(
            status__in=FOLLOW_UP_AUTO_EXPIRE_STATUSES
        )
    )
    expired_count = 0
    for lead in stale_followups:
        stale_anchor = _followup_activity_anchor(lead)
        if stale_anchor is None:
            continue
        stale_anchor = timezone.localtime(stale_anchor)
        if stale_anchor >= cutoff:
            continue
        actor_staff = lead.assigned_to or (getattr(getattr(lead, "interested_detail", None), "staff", None))
        lead.status = Lead.Status.EXPIRED_FOLLOWUP
        lead.callback_window = ""
        lead.callback_date = None
        lead.save(
            update_fields=[
                "status",
                "callback_window",
                "callback_date",
                "updated_at",
            ]
        )
        if actor_staff:
            _log_staff_action(
                actor_staff,
                StaffAction.ActionType.CALL_STATUS_UPDATED,
                lead=lead,
                metadata={
                    "source": "followup_expiry",
                    "status": Lead.Status.EXPIRED_FOLLOWUP,
                    "expiry_days": resolved_expiry_days,
                    "idle_days": max(0, (timezone.localdate(now) - timezone.localdate(stale_anchor)).days),
                },
            )
        expired_count += 1

    return {
        "expired_count": expired_count,
        "cutoff": cutoff,
        "enabled": True,
        "expiry_days": resolved_expiry_days,
    }


def _callback_tracking_queryset():
    return Lead.objects.filter(
        status=Lead.Status.INTERESTED,
        callback_date__isnull=False,
    ).exclude(callback_window="")


def build_staff_followup_sla_gate_status(staff, *, now=None, company_profile=None):
    profile = company_profile or get_company_profile()
    settings = _followup_expiry_settings(company_profile=profile)
    reference = now or timezone.now()
    warning_days = max(1, int(settings["warning_days"] or FOLLOWUP_STAFF_WARNING_DAYS))
    cutoff = timezone.localtime(reference - timedelta(days=warning_days))
    crossed_sla_count = 0
    for lead in Lead.objects.select_related("assigned_to", "interested_detail__staff").filter(
        assigned_to=staff,
        status=Lead.Status.CALL_BACK,
    ):
        activity_anchor = _followup_activity_anchor(lead)
        if activity_anchor is None:
            continue
        if timezone.localtime(activity_anchor) < cutoff:
            crossed_sla_count += 1

    _today, start, end = _today_range(reference)
    completed_followup_call_statuses = {
        Call.Status.INTERESTED,
        Call.Status.CALL_BACK,
        Call.Status.NO_ANSWER,
        Call.Status.NOT_INTERESTED,
        Call.Status.CONVERTED,
        Call.Status.INVALID_SHORT,
    }
    followup_calls_today = Call.objects.filter(
        staff=staff,
        start_time__range=(start, end),
        status__in=completed_followup_call_statuses,
        lead__assigned_to=staff,
        lead__status=Lead.Status.CALL_BACK,
    ).count()

    gate_enabled = bool(settings["sla_gate_enabled"])
    gate_mode = str(settings["sla_gate_mode"] or CompanyProfile.FollowupSlaGateMode.ALLOW_NORMAL_CALLS)
    requires_call_before_new = (
        gate_enabled
        and gate_mode == CompanyProfile.FollowupSlaGateMode.REQUIRE_ONE_FOLLOWUP_CALL
        and crossed_sla_count > 0
    )
    normal_lead_calls_allowed = not (requires_call_before_new and followup_calls_today < 1)

    if normal_lead_calls_allowed:
        gate_message = ""
    else:
        gate_message = (
            f"{crossed_sla_count} follow-up lead(s) crossed SLA. "
            "Complete at least one follow-up call before calling a New lead."
        )

    return {
        "warning_days": warning_days,
        "crossed_sla_count": crossed_sla_count,
        "gate_enabled": gate_enabled,
        "gate_mode": gate_mode,
        "requires_call_before_new": requires_call_before_new,
        "followup_calls_today": followup_calls_today,
        "normal_lead_calls_allowed": normal_lead_calls_allowed,
        "block_message": gate_message,
    }


def _staff_call_queue_queryset(queryset=None):
    base_queryset = queryset if queryset is not None else Lead.objects.all()
    return _visible_staff_lead_queryset(base_queryset)


def _active_started_call_lead_ids(*, target_staff=None):
    queryset = Call.objects.filter(status=Call.Status.STARTED, end_time__isnull=True)
    if target_staff is not None:
        queryset = queryset.filter(staff=target_staff)
    return set(queryset.values_list("lead_id", flat=True))


def _recovery_lead_queryset():
    return Lead.objects.filter(status__in=RECOVERY_LEAD_STATUSES)


def _lead_management_queryset():
    return Lead.objects.exclude(status__in=RECOVERY_LEAD_STATUSES)


def _oldest_manageable_leads_queryset():
    return _lead_management_queryset().filter(calls__isnull=False).distinct().order_by("created_at", "id")


def delete_oldest_manageable_leads(*, older_than_days=None, oldest_count=None):
    if older_than_days is None and oldest_count is None:
        raise ValueError("Choose delete by age or delete by oldest count.")

    queryset = _oldest_manageable_leads_queryset()
    mode = ""
    if older_than_days is not None:
        days_value = int(older_than_days)
        if days_value < 1:
            raise ValueError("Delete age must be at least 1 day.")
        cutoff = timezone.now() - timedelta(days=days_value)
        queryset = queryset.filter(created_at__lt=cutoff)
        mode = "age_days"
    else:
        count_value = int(oldest_count)
        if count_value < 1:
            raise ValueError("Delete count must be at least 1.")
        queryset = queryset[:count_value]
        mode = "oldest_count"

    selected_ids = list(queryset.values_list("id", flat=True))
    if not selected_ids:
        return {
            "deleted_count": 0,
            "mode": mode,
        }

    deleted_count = Lead.objects.filter(id__in=selected_ids).count()
    Lead.objects.filter(id__in=selected_ids).delete()
    auto_allocate_leads()
    return {
        "deleted_count": deleted_count,
        "mode": mode,
    }


def run_automatic_lead_cleanup_if_due(*, reference_date=None):
    company_profile = get_company_profile()
    if not company_profile.lead_auto_delete_enabled:
        return {"ran": False, "deleted_count": 0, "reason": "disabled"}

    today = reference_date or timezone.localdate()
    if company_profile.lead_auto_delete_last_run_on == today:
        return {"ran": False, "deleted_count": 0, "reason": "already_ran_today"}

    if company_profile.lead_auto_delete_mode == CompanyProfile.LeadAutoDeleteMode.OLDEST_COUNT:
        summary = delete_oldest_manageable_leads(oldest_count=company_profile.lead_auto_delete_count)
    else:
        summary = delete_oldest_manageable_leads(older_than_days=company_profile.lead_auto_delete_days)

    company_profile.lead_auto_delete_last_run_on = today
    company_profile.save(update_fields=["lead_auto_delete_last_run_on", "updated_at"])
    return {
        "ran": True,
        "deleted_count": summary["deleted_count"],
        "reason": "executed",
        "mode": summary["mode"],
    }


def _normalize_status_value(value):
    normalized = _normalize_column_name(value)
    status_map = {
        "new": Lead.Status.NEW,
        "follow up": Lead.Status.INTERESTED,
        "followup": Lead.Status.INTERESTED,
        "interested": Lead.Status.INTERESTED,
        "rejected": Lead.Status.NOT_INTERESTED,
        "not interested": Lead.Status.NOT_INTERESTED,
        "no response": Lead.Status.NO_ANSWER,
        "no answer": Lead.Status.NO_ANSWER,
        "expired follow up": Lead.Status.EXPIRED_FOLLOWUP,
        "expired_followup": Lead.Status.EXPIRED_FOLLOWUP,
        "callback": Lead.Status.INTERESTED,
        "call back": Lead.Status.INTERESTED,
        "converted": Lead.Status.CONVERTED,
    }
    return status_map.get(normalized, "")


def _normalize_handover_status_value(value):
    normalized = _normalize_column_name(value)
    handover_map = {
        "": "",
        "not sent": Lead.HandoverStatus.NOT_SENT,
        "not_sent": Lead.HandoverStatus.NOT_SENT,
        "pending": Lead.HandoverStatus.NOT_SENT,
        "sent": Lead.HandoverStatus.SENT,
        "sent to client": Lead.HandoverStatus.SENT,
        "accepted": Lead.HandoverStatus.ACCEPTED,
        "rejected": Lead.HandoverStatus.REJECTED,
        "completed": Lead.HandoverStatus.COMPLETED,
    }
    return handover_map.get(normalized, "")


def _normalize_callback_window_value(value):
    normalized = _normalize_column_name(value)
    callback_map = {
        "": "",
        "noon": Lead.CallbackWindow.NOON,
        "afternoon": Lead.CallbackWindow.NOON,
        "evening": Lead.CallbackWindow.EVENING,
        "night": Lead.CallbackWindow.NIGHT,
    }
    return callback_map.get(normalized, "")


def _format_callback_date_label(value):
    if not value:
        return ""
    return value.strftime("%d %b %Y")


def _format_callback_schedule_label(callback_date=None, callback_window=""):
    parts = []
    if callback_date:
        parts.append(_format_callback_date_label(callback_date))
    if callback_window:
        parts.append(dict(Lead.CallbackWindow.choices).get(callback_window, callback_window))
    return " • ".join(parts)


def _current_callback_window(now=None):
    local_now = timezone.localtime(now or timezone.now())
    hour = local_now.hour
    if hour in CALLBACK_NOON_HOURS:
        return Lead.CallbackWindow.NOON
    if hour in CALLBACK_EVENING_HOURS:
        return Lead.CallbackWindow.EVENING
    if hour in CALLBACK_NIGHT_HOURS:
        return Lead.CallbackWindow.NIGHT
    return ""


def _is_callback_due(callback_date, callback_window, *, now=None):
    if not callback_date:
        return False
    reference = now or timezone.now()
    return timezone.localdate(reference) >= callback_date


def _callback_window_start_hour(callback_window):
    return {
        Lead.CallbackWindow.NOON: 12,
        Lead.CallbackWindow.EVENING: 16,
        Lead.CallbackWindow.NIGHT: 20,
    }.get(callback_window)


def _callback_due_at(callback_date, callback_window):
    if not callback_date or not callback_window:
        return None
    start_hour = _callback_window_start_hour(callback_window)
    if start_hour is None:
        return None
    return timezone.make_aware(
        datetime.combine(callback_date, time(hour=start_hour)),
        timezone.get_current_timezone(),
    )


def _is_followup_highlighted(lead, *, now=None):
    if not lead.callback_date:
        return False
    reference = timezone.localtime(now or timezone.now())
    if timezone.localdate(reference) < lead.callback_date:
        return False
    if lead.last_contacted_at is None:
        return True
    return timezone.localdate(timezone.localtime(lead.last_contacted_at)) < lead.callback_date


def _normalize_followup_status(status):
    if status == Lead.Status.CALL_BACK:
        return Lead.Status.INTERESTED
    if status == Call.Status.CALL_BACK:
        return Call.Status.INTERESTED
    return status


def _is_followup_status(status):
    return status in {
        Lead.Status.INTERESTED,
        Lead.Status.CALL_BACK,
        Call.Status.INTERESTED,
        Call.Status.CALL_BACK,
    }


def _is_scheduled_followup(lead):
    return bool(_is_followup_status(getattr(lead, "status", "")) and lead.callback_date and lead.callback_window)


def _call_activity_stamp(call_row):
    if not call_row:
        return None
    return call_row.get("end_time") or call_row.get("start_time") or call_row.get("created_at")


def _followup_activity_anchor(lead):
    stamps = [
        getattr(lead, "followup_moved_back_at", None),
        getattr(lead, "last_contacted_at", None),
        getattr(lead, "updated_at", None),
        getattr(lead, "created_at", None),
    ]
    normalized_stamps = [timezone.localtime(stamp) for stamp in stamps if stamp]
    if not normalized_stamps:
        return None
    return max(normalized_stamps)


def _lead_route_status_meta(status):
    normalized = str(status or "").strip()
    if normalized == Lead.Status.CONVERTED:
        return {
            "label": "Converted",
            "tone": "success",
            "icon": "check2-circle-fill",
        }
    if normalized in {Lead.Status.NOT_INTERESTED}:
        return {
            "label": "Rejected",
            "tone": "danger",
            "icon": "x-circle-fill",
        }
    if normalized in {Lead.Status.NO_ANSWER}:
        return {
            "label": "No Response",
            "tone": "primary",
            "icon": "skip-forward-circle-fill",
        }
    if normalized in {Lead.Status.EXPIRED_FOLLOWUP}:
        return {
            "label": "Expired Follow Up",
            "tone": "warning",
            "icon": "hourglass-split",
        }
    if normalized == Lead.Status.INTERESTED:
        return {
            "label": "Interested",
            "tone": "warning",
            "icon": "hand-thumbs-up-fill",
        }
    if normalized in {Lead.Status.CALL_BACK}:
        return {
            "label": "Follow Up",
            "tone": "warning",
            "icon": "arrow-repeat",
        }
    if normalized == Lead.Status.NEW:
        return {
            "label": "New",
            "tone": "muted",
            "icon": "plus-circle-fill",
        }
    return {
        "label": str(status or "Lead"),
        "tone": "muted",
        "icon": "circle",
    }


def _lead_loan_stage_meta(lead):
    stage = str(getattr(lead, "loan_stage", "") or "").strip()
    lead_status = str(getattr(lead, "status", "") or "").strip()
    if stage:
        label = dict(Lead.LoanStage.choices).get(stage, stage.replace("_", " ").title())
        tone = {
            Lead.LoanStage.OFFICE_REVIEW: "warning",
            Lead.LoanStage.DOCUMENTS_PENDING: "primary",
            Lead.LoanStage.VERIFICATION: "primary",
            Lead.LoanStage.APPROVAL: "info",
            Lead.LoanStage.DISBURSEMENT: "success",
            Lead.LoanStage.SUCCESSFUL: "success",
            Lead.LoanStage.UNSUCCESSFUL: "danger",
        }.get(stage, "muted")
    elif lead_status == Lead.Status.INTERESTED:
        label = "Office Review"
        tone = "warning"
    elif lead_status == Lead.Status.CONVERTED:
        label = "Successful"
        tone = "success"
    elif lead_status in {Lead.Status.NOT_INTERESTED, Lead.Status.NO_ANSWER, Lead.Status.EXPIRED_FOLLOWUP}:
        label = "Unsuccessful"
        tone = "danger"
    else:
        label = "Not set"
        tone = "muted"
    return {
        "value": stage,
        "label": label,
        "tone": tone,
    }


def _lead_route_change_labels(metadata):
    metadata = metadata or {}
    changes = metadata.get("changes") or {}
    if not isinstance(changes, dict):
        changes = {}

    labels = []
    for key, field_label in (
        ("name", "Name"),
        ("phone", "Phone"),
        ("status", "Status"),
        ("loan_stage", "Loan stage"),
        ("assigned_to", "Owner"),
        ("callback_date", "Callback date"),
        ("callback_window", "Callback slot"),
        ("notes", "Notes"),
        ("handover_status", "Handover"),
    ):
        change_value = changes.get(key)
        if not isinstance(change_value, dict):
            continue
        before = change_value.get("from", "")
        after = change_value.get("to", "")
        if before == after:
            continue
        before_label = str(before or "—")
        after_label = str(after or "—")
        labels.append(f"{field_label}: {before_label} -> {after_label}")
    return labels


def _lead_route_event_tone_from_status(status):
    return _lead_route_status_meta(status)["tone"]


def _lead_route_event_icon_from_status(status):
    return _lead_route_status_meta(status)["icon"]


def _followup_no_answer_attempt_count(lead, *, staff=None, exclude_call_id=None):
    return _followup_no_response_progress(
        lead,
        staff=staff,
        exclude_call_id=exclude_call_id,
    )["attempt_count"]


def _followup_no_response_progress_from_rows(rows, *, include_attempt_at=None):
    attempt_times = []
    for row in rows:
        status = row["status"]
        if status == Call.Status.NO_ANSWER:
            stamp = _call_activity_stamp(row)
            if stamp:
                attempt_times.append(timezone.localtime(stamp))
            continue
        if status == Call.Status.INVALID_SHORT:
            continue
        break

    if include_attempt_at is not None:
        attempt_times.append(timezone.localtime(include_attempt_at))

    unique_dates = {stamp.date().isoformat() for stamp in attempt_times}
    unique_times = {stamp.strftime("%H:%M") for stamp in attempt_times}
    attempt_count = len(attempt_times)
    date_count = len(unique_dates)
    time_count = len(unique_times)
    ready_threshold = FOLLOWUP_NO_RESPONSE_LIMIT
    can_close = attempt_count >= ready_threshold
    remaining = max(0, ready_threshold - attempt_count)
    return {
        "attempt_count": attempt_count,
        "unique_date_count": date_count,
        "unique_time_count": time_count,
        "can_close": can_close,
        "remaining": remaining,
    }


def _followup_no_response_progress(
    lead,
    *,
    staff=None,
    exclude_call_id=None,
    include_attempt_at=None,
    preloaded_rows=None,
):
    if preloaded_rows is None:
        queryset = lead.calls.all()
        if staff is not None:
            queryset = queryset.filter(staff=staff)
        if exclude_call_id:
            queryset = queryset.exclude(id=exclude_call_id)
        rows = list(
            queryset.order_by("-start_time", "-created_at").values(
                "id",
                "staff_id",
                "status",
                "end_time",
                "start_time",
                "created_at",
            )
        )
    else:
        rows = list(preloaded_rows)
        if staff is not None:
            rows = [row for row in rows if row.get("staff_id") == getattr(staff, "id", staff)]
        if exclude_call_id:
            rows = [row for row in rows if str(row.get("id")) != str(exclude_call_id)]
    return _followup_no_response_progress_from_rows(rows, include_attempt_at=include_attempt_at)


def _quality_tone(score):
    if score >= 85:
        return "success"
    if score >= 70:
        return "primary"
    if score >= 55:
        return "warning"
    return "muted"


def _quality_label(score, *, has_activity):
    if not has_activity:
        return "No Recent Activity"
    if score >= 85:
        return "Excellent"
    if score >= 70:
        return "Strong"
    if score >= 55:
        return "Needs Attention"
    return "Review Needed"


def _build_quality_note(
    *,
    missed_callback_count,
    expired_followup_count,
    suspicious_block_count,
    suspicious_attempt_count,
    zero_only_block_count,
    long_away_count,
    real_call_count,
    verified_attempt_count,
    attempt_threshold,
):
    if verified_attempt_count >= attempt_threshold and real_call_count == 0:
        return (
            f"No real conversations were recorded across {verified_attempt_count} verified attempts, "
            "so this calling pattern needs review."
        )
    if (
        verified_attempt_count >= attempt_threshold
        and real_call_count > 0
        and (Decimal(real_call_count) / Decimal(verified_attempt_count)) < Decimal("0.20")
    ):
        return (
            f"Only {real_call_count} real conversations were recorded across {verified_attempt_count} verified attempts, "
            "so this calling pattern needs review."
        )
    if suspicious_block_count:
        return (
            f"Review calling pattern: {suspicious_attempt_count} attempts across "
            f"{suspicious_block_count} block(s) had too few real conversations."
        )
    if zero_only_block_count:
        return (
            f"{zero_only_block_count} calling block(s) had only unanswered attempts, "
            "so no work hours were added for those periods."
        )
    if missed_callback_count:
        return f"{missed_callback_count} scheduled follow-up lead(s) need review."
    if expired_followup_count:
        return (
            f"{expired_followup_count} follow-up lead(s) auto-expired in this review period, "
            "so response speed should be improved."
        )
    return "Build more recent call activity for a fuller review."


def _build_staff_quality_metrics(staff_ids, *, now=None, range_start=None, range_end=None, rules=None):
    if not staff_ids:
        return {}
    active_rules = rules or _work_review_rules()
    attempt_threshold = int(active_rules.get("attempt_threshold", MIN_REAL_CALLS_PER_ATTEMPT_BLOCK) or 1)
    followup_expired_penalty_points = max(
        0,
        int(active_rules.get("followup_expired_penalty_points", FOLLOWUP_EXPIRED_SCORE_PENALTY_POINTS) or 0),
    )
    followup_expired_penalty_cap = max(
        0,
        int(active_rules.get("followup_expired_penalty_cap", FOLLOWUP_EXPIRED_SCORE_PENALTY_CAP) or 0),
    )

    current_time = timezone.localtime(range_end or now or timezone.now())
    if range_start:
        lookback_start = timezone.localtime(range_start)
    else:
        lookback_start = timezone.make_aware(
            timezone.datetime.combine(_month_start(current_time.date()), timezone.datetime.min.time())
        )
    missed_callback_cutoff = current_time - timedelta(hours=MISSED_CALLBACK_AFTER_HOURS)
    lookback_days = max((current_time.date() - lookback_start.date()).days + 1, 1)
    period_label = lookback_start.strftime("%b %Y")

    recent_calls = Call.objects.filter(
        staff_id__in=staff_ids,
        start_time__gte=lookback_start,
    )
    if range_end:
        recent_calls = recent_calls.filter(start_time__lte=range_end)
    recent_calls = recent_calls.annotate(activity_day=TruncDate("start_time")).values(
        "staff_id",
        "lead_id",
        "status",
        "is_verified",
        "duration_seconds",
        "activity_day",
        "start_time",
        "end_time",
    )

    callback_rows = Lead.objects.filter(
        assigned_to_id__in=staff_ids,
        status=Lead.Status.INTERESTED,
        callback_date__isnull=False,
    ).exclude(callback_window="").values("assigned_to_id", "last_contacted_at")
    expired_followup_rows = Lead.objects.filter(
        assigned_to_id__in=staff_ids,
        status=Lead.Status.EXPIRED_FOLLOWUP,
        updated_at__gte=lookback_start,
    )
    if range_end:
        expired_followup_rows = expired_followup_rows.filter(updated_at__lte=range_end)
    expired_followup_rows = expired_followup_rows.values("assigned_to_id")

    metrics = {
        staff_id: {
            "total_completed_calls": 0,
            "invalid_short_calls": 0,
            "verified_resolved_calls": 0,
            "followup_started_leads": set(),
            "followup_closed_leads": set(),
            "callback_total": 0,
            "missed_callbacks": 0,
            "verified_attempt_count": 0,
            "real_call_count": 0,
            "zero_second_attempt_count": 0,
            "suspicious_block_count": 0,
            "zero_only_block_count": 0,
            "suspicious_attempt_count": 0,
            "long_away_count": 0,
            "expired_followup_count": 0,
        }
        for staff_id in staff_ids
    }
    calls_by_staff_day = defaultdict(list)

    for row in recent_calls:
        staff_id = row["staff_id"]
        if staff_id not in metrics:
            continue
        status = row["status"]
        if status != Call.Status.STARTED:
            metrics[staff_id]["total_completed_calls"] += 1
        if status == Call.Status.INVALID_SHORT:
            metrics[staff_id]["invalid_short_calls"] += 1
        elif status != Call.Status.STARTED and row["is_verified"]:
            metrics[staff_id]["verified_resolved_calls"] += 1
        if status in {Call.Status.INTERESTED, Call.Status.CALL_BACK}:
            metrics[staff_id]["followup_started_leads"].add(str(row["lead_id"]))
        if status in {Call.Status.CONVERTED, Call.Status.NOT_INTERESTED}:
            metrics[staff_id]["followup_closed_leads"].add(str(row["lead_id"]))
        if row["is_verified"] and status != Call.Status.STARTED:
            metrics[staff_id]["verified_attempt_count"] += 1
            duration_seconds = max(0, int(row["duration_seconds"] or 0))
            if duration_seconds > 0:
                metrics[staff_id]["real_call_count"] += 1
            else:
                metrics[staff_id]["zero_second_attempt_count"] += 1
            calls_by_staff_day[(staff_id, row["activity_day"])].append(
                {
                    "start_time": row["start_time"],
                    "end_time": row["end_time"],
                    "duration_seconds": duration_seconds,
                }
            )

    for row in callback_rows:
        staff_id = row["assigned_to_id"]
        if staff_id not in metrics:
            continue
        metrics[staff_id]["callback_total"] += 1
        last_contacted_at = row["last_contacted_at"]
        if not last_contacted_at or timezone.localtime(last_contacted_at) <= missed_callback_cutoff:
            metrics[staff_id]["missed_callbacks"] += 1

    for row in expired_followup_rows:
        staff_id = row["assigned_to_id"]
        if staff_id not in metrics:
            continue
        metrics[staff_id]["expired_followup_count"] += 1

    away_end_actions = {
        StaffAction.ActionType.APP_FOREGROUNDED,
        StaffAction.ActionType.RETURNED_ONLINE,
        StaffAction.ActionType.MARKED_OFFLINE,
        StaffAction.ActionType.SESSION_ENDED,
        StaffAction.ActionType.SESSION_AUTO_ENDED,
    }
    away_rows = (
        StaffAction.objects.filter(
            staff_id__in=staff_ids,
            created_at__gte=lookback_start - timedelta(days=1),
            action_type__in={StaffAction.ActionType.APP_BACKGROUNDED, *away_end_actions},
        )
        .values("staff_id", "action_type", "created_at")
        .order_by("staff_id", "created_at")
    )
    if range_end:
        away_rows = away_rows.filter(created_at__lte=range_end)
    open_away_since = {}
    for row in away_rows:
        staff_id = row["staff_id"]
        if staff_id not in metrics:
            continue
        action_type = row["action_type"]
        action_time = row["created_at"]
        if action_type == StaffAction.ActionType.APP_BACKGROUNDED:
            open_away_since[staff_id] = action_time
            continue
        away_started_at = open_away_since.pop(staff_id, None)
        if not away_started_at:
            continue
        overlap_start = max(away_started_at, lookback_start)
        away_seconds = max(0, int((action_time - overlap_start).total_seconds()))
        if away_seconds >= 10 * 60:
            metrics[staff_id]["long_away_count"] += 1

    for staff_id, away_started_at in open_away_since.items():
        if staff_id not in metrics:
            continue
        overlap_start = max(away_started_at, lookback_start)
        away_seconds = max(0, int((current_time - overlap_start).total_seconds()))
        if away_seconds >= 10 * 60:
            metrics[staff_id]["long_away_count"] += 1

    for (staff_id, _day), day_calls in calls_by_staff_day.items():
        summary = _call_activity_block_summary(day_calls, rules=active_rules)
        metrics[staff_id]["suspicious_block_count"] += summary["suspicious_block_count"]
        metrics[staff_id]["zero_only_block_count"] += summary["zero_only_block_count"]
        metrics[staff_id]["suspicious_attempt_count"] += summary["suspicious_attempt_count"]

    quality_payload = {}
    for staff_id, staff_metrics in metrics.items():
        total_completed_calls = staff_metrics["total_completed_calls"]
        invalid_short_calls = staff_metrics["invalid_short_calls"]
        resolved_calls = max(total_completed_calls - invalid_short_calls, 0)
        verified_resolved_calls = min(staff_metrics["verified_resolved_calls"], resolved_calls)
        callback_total = staff_metrics["callback_total"]
        missed_callback_count = staff_metrics["missed_callbacks"]
        verified_attempt_count = staff_metrics["verified_attempt_count"]
        real_call_count = staff_metrics["real_call_count"]
        zero_second_attempt_count = staff_metrics["zero_second_attempt_count"]
        suspicious_block_count = staff_metrics["suspicious_block_count"]
        zero_only_block_count = staff_metrics["zero_only_block_count"]
        suspicious_attempt_count = staff_metrics["suspicious_attempt_count"]
        long_away_count = staff_metrics["long_away_count"]
        expired_followup_count = staff_metrics["expired_followup_count"]
        real_call_ratio = (
            Decimal(real_call_count) / Decimal(verified_attempt_count)
            if verified_attempt_count > 0
            else None
        )

        weighted_total = Decimal("0")
        total_weight = Decimal("0")

        outcome_score = None
        if total_completed_calls > 0:
            resolved_ratio = Decimal(resolved_calls) / Decimal(total_completed_calls)
            verified_ratio = (
                Decimal(verified_resolved_calls) / Decimal(resolved_calls)
                if resolved_calls
                else Decimal("0")
            )
            outcome_score = ((resolved_ratio * Decimal("0.6")) + (verified_ratio * Decimal("0.4"))) * Decimal("100")
            weighted_total += outcome_score * Decimal("0.70")
            total_weight += Decimal("0.70")

        callback_score = None
        if callback_total > 0:
            callback_score = max(
                Decimal("0"),
                (Decimal(callback_total - missed_callback_count) / Decimal(callback_total)) * Decimal("100"),
            )
            weighted_total += callback_score * Decimal("0.30")
            total_weight += Decimal("0.30")

        if total_weight > 0:
            overall_score = int((weighted_total / total_weight).quantize(Decimal("1")))
        else:
            overall_score = 0

        pattern_penalty_points = min((suspicious_block_count * 12) + (zero_only_block_count * 10), 40)
        expired_followup_penalty_points = 0
        if followup_expired_penalty_points > 0 and expired_followup_count > 0:
            expired_followup_penalty_points = expired_followup_count * followup_expired_penalty_points
            if followup_expired_penalty_cap > 0:
                expired_followup_penalty_points = min(
                    expired_followup_penalty_points,
                    followup_expired_penalty_cap,
                )
        total_penalty_points = pattern_penalty_points + expired_followup_penalty_points
        overall_score = max(overall_score - total_penalty_points, 0)
        if verified_attempt_count >= attempt_threshold:
            if real_call_count == 0:
                overall_score = min(overall_score, 25)
            elif real_call_ratio is not None and real_call_ratio < Decimal("0.20"):
                overall_score = min(overall_score, 45)
            elif (
                real_call_ratio is not None
                and real_call_ratio < Decimal("0.35")
                and (suspicious_block_count > 0 or zero_only_block_count > 0)
            ):
                overall_score = min(overall_score, 54)

        has_activity = total_completed_calls > 0 or verified_attempt_count > 0 or callback_total > 0

        outcome_value = int(outcome_score.quantize(Decimal("1"))) if outcome_score is not None else None
        callback_value = int(callback_score.quantize(Decimal("1"))) if callback_score is not None else None
        if verified_attempt_count > 0:
            attempt_review_label = f"{real_call_count} real from {verified_attempt_count} attempts"
        else:
            attempt_review_label = "--"
        if long_away_count > 0:
            away_review_label = f"{long_away_count} period(s) recorded"
        else:
            away_review_label = "No long away periods"

        quality_payload[staff_id] = {
            "score": overall_score,
            "label": _quality_label(overall_score, has_activity=has_activity),
            "tone": _quality_tone(overall_score) if has_activity else "muted",
            "note": _build_quality_note(
                missed_callback_count=missed_callback_count,
                expired_followup_count=expired_followup_count,
                suspicious_block_count=suspicious_block_count,
                suspicious_attempt_count=suspicious_attempt_count,
                zero_only_block_count=zero_only_block_count,
                long_away_count=long_away_count,
                real_call_count=real_call_count,
                verified_attempt_count=verified_attempt_count,
                attempt_threshold=attempt_threshold,
            ),
            "has_activity": has_activity,
            "invalid_short_count": invalid_short_calls,
            "outcome_consistency": outcome_value,
            "outcome_consistency_label": f"{outcome_value}%" if outcome_value is not None else "--",
            "callback_discipline": callback_value,
            "callback_discipline_label": f"{callback_value}%" if callback_value is not None else "--",
            "callback_total": callback_total,
            "missed_callbacks": missed_callback_count,
            "verified_attempt_count": verified_attempt_count,
            "real_call_count": real_call_count,
            "zero_second_attempt_count": zero_second_attempt_count,
            "suspicious_block_count": suspicious_block_count,
            "zero_only_block_count": zero_only_block_count,
            "suspicious_attempt_count": suspicious_attempt_count,
            "expired_followup_count": expired_followup_count,
            "pattern_penalty_points": pattern_penalty_points,
            "expired_followup_penalty_points": expired_followup_penalty_points,
            "total_penalty_points": total_penalty_points,
            "attempt_review_label": attempt_review_label,
            "away_review_label": away_review_label,
            "long_away_count": long_away_count,
            "lookback_days": lookback_days,
            "period_label": period_label,
            "period_start": lookback_start.date().isoformat(),
        }

    return quality_payload


def build_staff_quality_history(staff, *, months=6, now=None):
    if not staff:
        return []
    reference_date = timezone.localdate()
    history_rows = []
    for offset in range(max(int(months or 0), 0)):
        month_date = _shift_month(reference_date.replace(day=1), -offset)
        range_start, range_end = _month_range_for_reference(
            month_date,
            end_at=timezone.localtime(now or timezone.now()),
        )
        quality = _build_staff_quality_metrics(
            [staff.id],
            range_start=range_start,
            range_end=range_end,
        ).get(staff.id, {})
        history_rows.append(
            {
                "month_label": month_date.strftime("%b %Y"),
                "score": quality.get("score", 0),
                "label": quality.get("label", "No Recent Activity"),
                "tone": quality.get("tone", "muted"),
                "note": quality.get("note", "Build more recent call activity for a fuller review."),
            }
        )
    return history_rows


def _staff_review_call_rows(staff, *, now=None):
    reference = now or timezone.now()
    lookback_start = reference - timedelta(days=QUALITY_SCORE_LOOKBACK_DAYS)
    recent_calls = (
        Call.objects.filter(
            staff=staff,
            start_time__gte=lookback_start,
        )
        .select_related("lead")
        .order_by("-start_time", "-created_at")
    )

    latest_by_lead = {}
    for call in recent_calls:
        latest_by_lead.setdefault(call.lead_id, call)

    review_rows = []
    for call in latest_by_lead.values():
        if call.status == Call.Status.STARTED:
            continue
        if not call.is_verified or call.is_qualifying:
            continue

        lead = call.lead
        review_rows.append(
            {
                "lead_id": str(lead.id),
                "call_id": str(call.id),
                "lead_name": lead.name,
                "lead_phone": lead.phone,
                "call_status": call.status,
                "call_status_label": call.get_status_display(),
                "duration_label": _format_duration(call.duration_seconds),
                "start_time": _format_datetime(call.start_time),
                "end_time": _format_datetime(call.end_time),
                "current_lead_status": lead.status,
                "current_lead_status_label": lead.get_status_display(),
                "assigned_to_staff": bool(lead.assigned_to_id == staff.id),
                "is_invalid_short": call.status == Call.Status.INVALID_SHORT,
                "search_text": " ".join(
                    [
                        lead.name.lower(),
                        lead.phone.lower(),
                        call.get_status_display().lower(),
                        lead.get_status_display().lower(),
                        _format_duration(call.duration_seconds).lower(),
                    ]
                ),
            }
        )

    review_rows.sort(
        key=lambda row: (
            0 if row["is_invalid_short"] else 1,
            row["lead_name"].lower(),
            row["lead_phone"],
        )
    )
    return review_rows


def _review_lead_ids_for_staff(staff, *, now=None):
    return [uuid.UUID(row["lead_id"]) for row in _staff_review_call_rows(staff, now=now)]


def _build_work_review_day_previews(staff_ids, *, now=None, preview_days=7, preview_limit=2, rules=None):
    if not staff_ids:
        return {}
    active_rules = rules or _work_review_rules()
    attempt_threshold = int(active_rules.get("attempt_threshold", MIN_REAL_CALLS_PER_ATTEMPT_BLOCK) or 1)

    current_time = timezone.localtime(now or timezone.now())
    preview_start_date = current_time.date() - timedelta(days=max(preview_days - 1, 0))
    preview_start = timezone.make_aware(
        timezone.datetime.combine(preview_start_date, timezone.datetime.min.time())
    )

    recent_calls = (
        Call.objects.filter(
            staff_id__in=staff_ids,
            start_time__gte=preview_start,
            is_verified=True,
        )
        .annotate(activity_day=TruncDate("start_time"))
        .values(
            "staff_id",
            "status",
            "duration_seconds",
            "activity_day",
            "start_time",
            "end_time",
            "created_at",
        )
    )

    calls_by_staff_day = defaultdict(list)
    invalid_short_totals = defaultdict(int)
    for row in recent_calls:
        if row["status"] == Call.Status.STARTED or not row["activity_day"]:
            continue
        key = (row["staff_id"], row["activity_day"])
        calls_by_staff_day[key].append(
            {
                "start_time": row["start_time"],
                "end_time": row["end_time"],
                "created_at": row["created_at"],
                "duration_seconds": max(0, int(row["duration_seconds"] or 0)),
            }
        )
        if row["status"] == Call.Status.INVALID_SHORT:
            invalid_short_totals[key] += 1

    previews_by_staff = defaultdict(list)
    for (staff_id, activity_day), day_calls in calls_by_staff_day.items():
        summary = _call_activity_block_summary(day_calls, rules=active_rules)
        attempt_count = summary["attempt_count"]
        real_call_count = summary["real_call_count"]
        zero_second_attempt_count = summary["zero_second_attempt_count"]
        suspicious_block_count = summary["suspicious_block_count"]
        zero_only_block_count = summary["zero_only_block_count"]
        invalid_short_count = invalid_short_totals[(staff_id, activity_day)]

        review_state = "stable"
        review_label = "Stable day"
        review_tone = "success"
        if attempt_count >= attempt_threshold and real_call_count == 0:
            review_state = "review"
            review_label = "No real conversations"
            review_tone = "danger"
        elif suspicious_block_count or zero_only_block_count:
            review_state = "review"
            review_label = "Call pattern needs review"
            review_tone = "danger"
        elif invalid_short_count or zero_second_attempt_count:
            review_state = "attention"
            review_label = "Some empty call attempts"
            review_tone = "warning"

        if review_state == "stable":
            continue

        previews_by_staff[staff_id].append(
            {
                "date": activity_day,
                "date_label": activity_day.strftime("%d %b %Y"),
                "day_label": activity_day.strftime("%a"),
                "attempt_count": attempt_count,
                "real_call_count": real_call_count,
                "zero_second_attempt_count": zero_second_attempt_count,
                "invalid_short_count": invalid_short_count,
                "review_state": review_state,
                "review_label": review_label,
                "review_tone": review_tone,
            }
        )

    preview_payload = {}
    for staff_id, rows in previews_by_staff.items():
        ordered_rows = sorted(rows, key=lambda row: row["date"], reverse=True)
        preview_payload[staff_id] = {
            "count": len(ordered_rows),
            "rows": ordered_rows[:preview_limit],
        }
    return preview_payload


def _zero_talk_block_details_by_staff(
    staff_ids,
    *,
    range_start,
    range_end,
    block_limit=3,
    include_unverified=True,
    include_invalid_short=True,
    rules=None,
):
    staff_ids = [staff_id for staff_id in (staff_ids or []) if staff_id]
    if not staff_ids:
        return {}
    active_rules = rules or _work_review_rules()
    attempt_threshold = int(active_rules.get("attempt_threshold", MIN_REAL_CALLS_PER_ATTEMPT_BLOCK) or 1)

    call_queryset = Call.objects.filter(staff_id__in=staff_ids, start_time__range=(range_start, range_end))
    if not include_unverified:
        call_queryset = call_queryset.filter(is_verified=True)
    if not include_invalid_short:
        call_queryset = call_queryset.exclude(status=Call.Status.INVALID_SHORT)

    calls = (
        call_queryset
        .select_related("lead")
        .order_by("staff_id", "start_time", "end_time", "id")
    )
    calls_by_staff = defaultdict(list)
    for call in calls:
        calls_by_staff[call.staff_id].append(call)

    block_payload = {}
    for staff_id, staff_calls in calls_by_staff.items():
        zero_blocks = []
        all_streaks = []
        blocks = _call_activity_blocks_with_stats(staff_calls, rules=active_rules)
        zero_streak = []
        zero_streak_attempts = 0
        pending_gap_label = "0s"
        pending_next_call_time = "--"

        def flush_zero_streak():
            nonlocal zero_blocks, all_streaks, zero_streak, zero_streak_attempts, pending_gap_label, pending_next_call_time
            if not zero_streak:
                return
            merged_calls = []
            total_zero_seconds = 0
            streak_start = None
            streak_end = None
            streak_gap_seconds = 0
            previous_block_end = None
            for block in zero_streak:
                merged_calls.extend(block["calls"])
                total_zero_seconds += block["zero_seconds_in_block"]
                if block["block_start"]:
                    streak_start = block["block_start"] if not streak_start else min(streak_start, block["block_start"])
                if block["block_end"]:
                    streak_end = block["block_end"] if not streak_end else max(streak_end, block["block_end"])
                if previous_block_end and block["block_start"]:
                    gap_seconds = max(0, int((block["block_start"] - previous_block_end).total_seconds()))
                    streak_gap_seconds += gap_seconds
                previous_block_end = block["block_end"]

            local_start = timezone.localtime(streak_start) if streak_start else None
            local_end = timezone.localtime(streak_end) if streak_end else None
            date_label = local_start.strftime("%d %b %Y") if local_start else "Unknown date"
            time_range = (
                f"{local_start.strftime('%I:%M %p').lstrip('0')} - {local_end.strftime('%I:%M %p').lstrip('0')}"
                if local_start and local_end
                else "--"
            )
            block_seconds = max(0, int((streak_end - streak_start).total_seconds())) if streak_start and streak_end else 0
            call_rows = []
            for call in merged_calls:
                call_rows.append(
                    {
                        "lead_name": call.lead.name,
                        "lead_phone": call.lead.phone,
                        "start_time": _format_datetime(call.start_time),
                        "duration_label": _format_duration(call.duration_seconds),
                        "status_label": call.get_status_display(),
                    }
                )

            is_zero_talk = zero_streak_attempts >= attempt_threshold
            streak_payload = {
                "date_label": date_label,
                "time_range": time_range,
                "attempt_count": len(merged_calls),
                "zero_second_count": total_zero_seconds,
                "duration_label": _format_duration(block_seconds),
                "calls": call_rows,
                "call_count": len(call_rows),
                "extra_calls": 0,
                "streak_note": (
                    f"Zero-only streak ({attempt_threshold}+ attempts reached)"
                    if is_zero_talk
                    else f"Zero-only streak below {attempt_threshold} attempts"
                ),
                "next_call_gap_label": pending_gap_label,
                "next_connected_call_time": pending_next_call_time or "No connected call yet",
                "streak_gap_label": _format_duration(streak_gap_seconds),
                "is_zero_talk": is_zero_talk,
            }

            all_streaks.append(streak_payload)
            if is_zero_talk:
                zero_blocks.append(streak_payload)
            zero_streak = []
            zero_streak_attempts = 0
            pending_gap_label = "0s"
            pending_next_call_time = "--"

        for block in blocks:
            if block["real_calls_in_block"] <= 0:
                zero_streak.append(block)
                zero_streak_attempts += len(block["calls"])
                continue

            if zero_streak:
                gap_seconds = 0
                if zero_streak[-1]["block_end"] and block["block_start"]:
                    gap_seconds = max(
                        0,
                        int((block["block_start"] - zero_streak[-1]["block_end"]).total_seconds()),
                    )
                pending_gap_label = _format_duration(gap_seconds)
                if block.get("first_real_call_start"):
                    pending_next_call_time = _format_datetime(block["first_real_call_start"])
                else:
                    pending_next_call_time = _format_datetime(block.get("block_start"))

            flush_zero_streak()

        flush_zero_streak()

        if zero_blocks:
            total_blocks = len(zero_blocks)
            kept_blocks = zero_blocks[-block_limit:]
            block_payload[staff_id] = {
                "blocks": kept_blocks,
                "extra_count": max(total_blocks - len(kept_blocks), 0),
                "all_streaks": all_streaks,
            }
        elif all_streaks:
            block_payload[staff_id] = {
                "blocks": [],
                "extra_count": 0,
                "all_streaks": all_streaks,
            }

    return block_payload


def reassign_staff_review_leads(staff):
    review_lead_ids = _review_lead_ids_for_staff(staff)
    if not review_lead_ids:
        return {
            "review_count": 0,
            "released_count": 0,
            "assigned_count": 0,
            "waiting_count": 0,
        }

    released_count = release_staff_queue(staff)
    Lead.objects.filter(id__in=review_lead_ids).update(
        status=Lead.Status.NEW,
        assigned_to=staff,
        callback_window="",
        callback_date=None,
        updated_at=timezone.now(),
    )
    auto_allocate_leads(target_staff=staff, prioritized_lead_ids=review_lead_ids)
    assigned_count = Lead.objects.filter(id__in=review_lead_ids, assigned_to=staff).count()
    waiting_count = max(len(review_lead_ids) - assigned_count, 0)
    return {
        "review_count": len(review_lead_ids),
        "released_count": released_count,
        "assigned_count": assigned_count,
        "waiting_count": waiting_count,
    }


def reset_staff_review_leads_to_new_queue(staff):
    review_lead_ids = _review_lead_ids_for_staff(staff)
    if not review_lead_ids:
        return {"review_count": 0, "reopened_count": 0}

    reopened_count = Lead.objects.filter(id__in=review_lead_ids).update(
        status=Lead.Status.NEW,
        assigned_to=None,
        callback_window="",
        callback_date=None,
        updated_at=timezone.now(),
    )
    auto_allocate_leads(prioritized_lead_ids=review_lead_ids)
    return {"review_count": len(review_lead_ids), "reopened_count": reopened_count}


def _with_lead_priority(queryset, *, now=None, prioritized_lead_ids=None):
    callback_date = timezone.localdate(now or timezone.now())
    current_slot = _current_callback_window(now)
    if current_slot:
        queue_priority = Case(
            When(
                status=Lead.Status.CALL_BACK,
                callback_date=callback_date,
                callback_window=current_slot,
                then=Value(0),
            ),
            When(status=Lead.Status.INTERESTED, then=Value(1)),
            When(status=Lead.Status.CALL_BACK, then=Value(2)),
            When(status=Lead.Status.NEW, then=Value(3)),
            default=Value(4),
            output_field=IntegerField(),
        )
    else:
        queue_priority = Case(
            When(status=Lead.Status.INTERESTED, then=Value(0)),
            When(status=Lead.Status.CALL_BACK, then=Value(1)),
            When(status=Lead.Status.NEW, then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ordered_queryset = queryset.annotate(queue_priority=queue_priority)
    if prioritized_lead_ids:
        ordered_queryset = ordered_queryset.annotate(
            manual_priority=Case(
                When(id__in=list(prioritized_lead_ids), then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        )
    return ordered_queryset


def _ordered_lead_queryset(queryset, *, now=None, include_assignee=False, prioritized_lead_ids=None):
    ordered_queryset = _with_lead_priority(
        queryset,
        now=now,
        prioritized_lead_ids=prioritized_lead_ids,
    ).annotate(lead_sequence_anchor=Coalesce("last_contacted_at", "created_at"))
    order_fields = []
    if include_assignee:
        order_fields.append("assigned_to_id")
    if prioritized_lead_ids:
        order_fields.append("manual_priority")
    order_fields.extend(["queue_priority", "lead_sequence_anchor", "created_at", "id"])
    return ordered_queryset.order_by(*order_fields)


def _visible_staff_lead_queryset(queryset, *, now=None):
    return queryset.filter(status=Lead.Status.NEW)


def is_staff_lead_visible_now(lead, *, now=None):
    if lead.status != Lead.Status.INTERESTED:
        return True
    if not _is_scheduled_followup(lead):
        return False
    return _is_callback_due(lead.callback_date, lead.callback_window, now=now)


def _daily_completed_call_counts():
    _, start, end = _today_range()
    qualifying_statuses = [
        Call.Status.INTERESTED,
        Call.Status.NOT_INTERESTED,
        Call.Status.NO_ANSWER,
        Call.Status.CALL_BACK,
        Call.Status.CONVERTED,
    ]
    return {
        row["staff_id"]: row["count"]
        for row in Call.objects.filter(start_time__range=(start, end), status__in=qualifying_statuses)
        .values("staff_id")
        .annotate(count=Count("id"))
    }


def release_staff_queue(staff):
    if not staff:
        return 0
    released = Lead.objects.filter(
        assigned_to=staff,
        status__in=ACTIVE_QUEUE_STATUSES,
    ).update(assigned_to=None, updated_at=timezone.now())
    return released


def _normalize_active_queue_assignments(*, target_staff=None, prioritized_lead_ids=None):
    now = timezone.now()
    queue_limit = get_lead_queue_limit()
    protected_lead_ids = _active_started_call_lead_ids(target_staff=target_staff)
    queue_queryset = _staff_call_queue_queryset(
        _lead_queue_queryset().select_related("assigned_to")
    ).exclude(assigned_to=None)
    if target_staff is not None:
        queue_queryset = queue_queryset.filter(assigned_to=target_staff)

    allocatable_queue_ids = _allocatable_lead_ids(queue_queryset.values_list("id", flat=True))
    release_ids = []
    kept_counts = defaultdict(int)
    for lead in _ordered_lead_queryset(
        queue_queryset,
        now=now,
        include_assignee=True,
        prioritized_lead_ids=prioritized_lead_ids,
    ):
        assigned_staff = lead.assigned_to
        if not assigned_staff or assigned_staff.role != Staff.Role.STAFF or not assigned_staff.is_active:
            release_ids.append(lead.id)
            continue

        if not assigned_staff.receives_new_leads:
            kept_counts[assigned_staff.id] += 1
            continue

        if lead.id not in allocatable_queue_ids:
            release_ids.append(lead.id)
            continue

        if lead.id in protected_lead_ids:
            kept_counts[assigned_staff.id] += 1
            continue

        kept_counts[assigned_staff.id] += 1
        if kept_counts[assigned_staff.id] > queue_limit:
            release_ids.append(lead.id)

    if not release_ids:
        return 0

    return Lead.objects.filter(id__in=release_ids).update(
        assigned_to=None,
        updated_at=timezone.now(),
    )


def auto_allocate_leads(*, target_staff=None, prioritized_lead_ids=None):
    now = timezone.now()
    queue_limit = get_lead_queue_limit()
    _normalize_active_queue_assignments(
        target_staff=target_staff,
        prioritized_lead_ids=prioritized_lead_ids,
    )
    staff_queryset = _staff_queryset().filter(is_active=True, receives_new_leads=True)
    if target_staff is not None:
        staff_queryset = staff_queryset.filter(id=target_staff.id)

    staff_members = list(staff_queryset)
    if not staff_members:
        return {
            "assigned_count": 0,
            "remaining_unassigned_count": _staff_call_queue_queryset(
                _lead_queue_queryset().filter(assigned_to=None)
            ).count(),
        }

    active_counts = {
        row["assigned_to"]: row["count"]
        for row in _staff_call_queue_queryset(_lead_queue_queryset())
        .exclude(assigned_to=None)
        .values("assigned_to")
        .annotate(count=Count("id"))
    }
    completed_today_counts = _daily_completed_call_counts()
    allocatable_open_ids = _allocatable_lead_ids(
        _staff_call_queue_queryset(_lead_queue_queryset().filter(assigned_to=None)).values_list("id", flat=True)
    )
    open_leads = list(
        _ordered_lead_queryset(
            _staff_call_queue_queryset(
                _lead_queue_queryset().filter(assigned_to=None, id__in=allocatable_open_ids)
            ),
            now=now,
            prioritized_lead_ids=prioritized_lead_ids,
        ).values("id", "status", "callback_window")
    )
    if not open_leads:
        return {"assigned_count": 0, "remaining_unassigned_count": 0}

    staff_slots = [
        {
            "staff_id": staff.id,
            "name": staff.name.lower(),
            "active_count": active_counts.get(staff.id, 0),
            "completed_today": completed_today_counts.get(staff.id, 0),
        }
        for staff in staff_members
    ]

    assigned_by_staff = defaultdict(list)
    unassigned_lead_ids = []
    for lead in open_leads:
        eligible_slots = staff_slots

        eligible_slots.sort(
            key=lambda item: (
                item["active_count"],
                -item["completed_today"],
                item["name"],
            )
        )
        chosen_slot = next(
            (slot for slot in eligible_slots if slot["active_count"] < queue_limit),
            None,
        )
        if not chosen_slot:
            unassigned_lead_ids.append(lead["id"])
            continue
        assigned_by_staff[chosen_slot["staff_id"]].append(lead["id"])
        chosen_slot["active_count"] += 1

    assigned_count = 0
    assigned_at = timezone.now()
    for staff_id, lead_ids in assigned_by_staff.items():
        assigned_count += len(lead_ids)
        Lead.objects.filter(id__in=lead_ids).update(
            assigned_to_id=staff_id,
            updated_at=assigned_at,
        )

    return {
        "assigned_count": assigned_count,
        "remaining_unassigned_count": len(unassigned_lead_ids),
    }


def assign_imported_leads_to_staff(*, imported_lead_ids, selected_staff):
    selected_staff = [
        staff
        for staff in selected_staff
        if staff and staff.is_active and staff.receives_new_leads and staff.role == Staff.Role.STAFF
    ]
    if not imported_lead_ids or not selected_staff:
        return {
            "assigned_count": 0,
            "remaining_unassigned_count": len(imported_lead_ids or []),
            "released_count": 0,
        }

    queue_limit = get_lead_queue_limit()
    released_count = 0
    for staff in selected_staff:
        released_count += release_staff_queue(staff)

    staff_slots = [
        {
            "staff_id": staff.id,
            "name": staff.name.lower(),
            "assigned_count": 0,
        }
        for staff in selected_staff
    ]
    allocatable_imported_ids = _allocatable_lead_ids(imported_lead_ids)
    ordered_imported_ids = list(
        Lead.objects.filter(id__in=list(allocatable_imported_ids)).order_by("created_at", "id").values_list("id", flat=True)
    )
    assigned_by_staff = defaultdict(list)
    remaining_unassigned = []
    for lead_id in ordered_imported_ids:
        eligible_slots = [slot for slot in staff_slots if slot["assigned_count"] < queue_limit]
        eligible_slots.sort(key=lambda item: (item["assigned_count"], item["name"], str(item["staff_id"])))
        chosen_slot = eligible_slots[0] if eligible_slots else None
        if not chosen_slot:
            remaining_unassigned.append(lead_id)
            continue
        assigned_by_staff[chosen_slot["staff_id"]].append(lead_id)
        chosen_slot["assigned_count"] += 1

    assigned_at = timezone.now()
    assigned_count = 0
    for staff_id, lead_ids in assigned_by_staff.items():
        assigned_count += len(lead_ids)
        Lead.objects.filter(id__in=lead_ids).update(assigned_to_id=staff_id, updated_at=assigned_at)

    return {
        "assigned_count": assigned_count,
        "remaining_unassigned_count": len(remaining_unassigned),
        "released_count": released_count,
    }


def assign_selected_leads_to_staff_queue(*, selected_lead_ids, target_staff):
    if (
        not selected_lead_ids
        or not target_staff
        or not target_staff.is_active
        or not target_staff.receives_new_leads
        or target_staff.role != Staff.Role.STAFF
    ):
        return {
            "selected_count": 0,
            "eligible_count": 0,
            "assigned_count": 0,
            "waiting_count": 0,
            "skipped_count": 0,
            "displaced_count": 0,
        }

    ordered_selected_ids = list(
        Lead.objects.filter(id__in=list(selected_lead_ids))
        .order_by("created_at", "id")
        .values_list("id", flat=True)
    )
    selected_leads = list(
        Lead.objects.filter(id__in=ordered_selected_ids).select_related("assigned_to")
    )
    if not selected_leads:
        return {
            "selected_count": 0,
            "eligible_count": 0,
            "assigned_count": 0,
            "waiting_count": 0,
            "skipped_count": 0,
            "displaced_count": 0,
        }

    allocatable_selected_ids = _allocatable_lead_ids(lead.id for lead in selected_leads)
    eligible_lead_ids = []
    previous_staff_ids = set()
    for lead in selected_leads:
        if not _is_active_queue_status(lead.status):
            continue
        if lead.id not in allocatable_selected_ids:
            continue
        eligible_lead_ids.append(lead.id)
        if lead.assigned_to_id and lead.assigned_to_id != target_staff.id:
            previous_staff_ids.add(lead.assigned_to_id)

    skipped_count = max(len(selected_leads) - len(eligible_lead_ids), 0)
    if not eligible_lead_ids:
        return {
            "selected_count": len(selected_leads),
            "eligible_count": 0,
            "assigned_count": 0,
            "waiting_count": 0,
            "skipped_count": skipped_count,
            "displaced_count": 0,
        }

    existing_target_queue_ids = set(
        _staff_call_queue_queryset(
            Lead.objects.filter(
                assigned_to=target_staff,
                status__in=ACTIVE_QUEUE_STATUSES,
            ).exclude(id__in=eligible_lead_ids)
        ).values_list("id", flat=True)
    )

    assigned_at = timezone.now()
    Lead.objects.filter(id__in=eligible_lead_ids).update(
        assigned_to=target_staff,
        updated_at=assigned_at,
    )
    auto_allocate_leads(
        target_staff=target_staff,
        prioritized_lead_ids=eligible_lead_ids,
    )

    for previous_staff in Staff.objects.filter(
        id__in=list(previous_staff_ids),
        role=Staff.Role.STAFF,
        is_active=True,
    ).exclude(id=target_staff.id):
        auto_allocate_leads(target_staff=previous_staff)

    assigned_count = Lead.objects.filter(
        id__in=eligible_lead_ids,
        assigned_to=target_staff,
    ).count()
    waiting_count = max(len(eligible_lead_ids) - assigned_count, 0)
    displaced_count = 0
    if existing_target_queue_ids:
        displaced_count = Lead.objects.filter(
            id__in=list(existing_target_queue_ids)
        ).exclude(assigned_to=target_staff).count()

    return {
        "selected_count": len(selected_leads),
        "eligible_count": len(eligible_lead_ids),
        "assigned_count": assigned_count,
        "waiting_count": waiting_count,
        "skipped_count": skipped_count,
        "displaced_count": displaced_count,
    }


def _decode_csv_bytes(file_bytes):
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")


def _decode_vcard_value(value):
    decoded = str(value or "").replace("\\n", " ").replace("\\N", " ")
    decoded = decoded.replace("\\,", ",").replace("\\;", ";").replace("\\\\", "\\")
    return " ".join(decoded.split()).strip()


def _structured_vcard_name(value):
    parts = [_decode_vcard_value(part) for part in str(value or "").split(";")]
    ordered_parts = [part for part in (parts[1:3] + parts[:1] + parts[3:]) if part]
    return " ".join(ordered_parts).strip()


def _unfold_vcard_lines(content):
    unfolded_lines = []
    for raw_line in str(content or "").splitlines():
        line = raw_line.rstrip("\r")
        if not line:
            unfolded_lines.append("")
            continue
        if line.startswith((" ", "\t")) and unfolded_lines:
            unfolded_lines[-1] = f"{unfolded_lines[-1]}{line[1:]}"
            continue
        unfolded_lines.append(line)
    return unfolded_lines


def _read_vcard_rows(uploaded_file):
    content = _decode_csv_bytes(uploaded_file.read())
    lines = _unfold_vcard_lines(content)
    rows = [["Name", "Phone"]]
    current_name = ""
    current_structured_name = ""
    current_phones = []

    def flush_contact():
        nonlocal current_name, current_structured_name, current_phones
        resolved_name = current_name or current_structured_name
        for phone in current_phones:
            contact_name = resolved_name or f"Contact {phone[-4:]}" if len(phone) >= 4 else "Imported Contact"
            rows.append([contact_name, phone])
        current_name = ""
        current_structured_name = ""
        current_phones = []

    for line in lines:
        upper_line = line.upper()
        if upper_line == "BEGIN:VCARD":
            current_name = ""
            current_structured_name = ""
            current_phones = []
            continue
        if upper_line == "END:VCARD":
            flush_contact()
            continue
        if ":" not in line:
            continue
        field_meta, raw_value = line.split(":", 1)
        field_name = field_meta.split(";", 1)[0].upper()
        value = _decode_vcard_value(raw_value)
        if field_name == "FN" and value:
            current_name = value
        elif field_name == "N" and value and not current_name:
            current_structured_name = _structured_vcard_name(raw_value)
        elif field_name == "TEL" and value:
            phone_value = value[4:] if value.lower().startswith("tel:") else value
            if phone_value:
                current_phones.append(phone_value)

    if current_name or current_structured_name or current_phones:
        flush_contact()

    return [row for row in rows if any(str(cell or "").strip() for cell in row)]


def _read_lead_rows_from_upload(uploaded_file):
    file_name = str(getattr(uploaded_file, "name", "")).lower()
    uploaded_file.seek(0)

    if file_name.endswith(".csv"):
        content = _decode_csv_bytes(uploaded_file.read())
        return [row for row in csv.reader(io.StringIO(content)) if any(str(cell or "").strip() for cell in row)]

    if file_name.endswith(".xlsx") or file_name.endswith(".xlsm"):
        if load_workbook is None:
            raise ValueError("Excel import support is not available right now.")
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            normalized_row = ["" if cell is None else str(cell).strip() for cell in row]
            if any(normalized_row):
                rows.append(normalized_row)
        return rows

    if file_name.endswith(".vcf"):
        return _read_vcard_rows(uploaded_file)

    raise ValueError("Upload a CSV, Excel, or VCF file.")


def _detect_lead_column_indexes(rows):
    if not rows:
        raise ValueError("The uploaded file is empty.")

    header = rows[0]
    normalized_header = [_normalize_column_name(cell) for cell in header]
    name_index = None
    phone_index = None

    for index, value in enumerate(normalized_header):
        if name_index is None and value in NAME_COLUMN_ALIASES:
            name_index = index
        if phone_index is None and value in PHONE_COLUMN_ALIASES:
            phone_index = index

    has_header_match = name_index is not None or phone_index is not None
    if not has_header_match:
        if len(header) < 2:
            raise ValueError("The file must contain name and phone columns.")
        return 0, 1, rows

    if name_index is None or phone_index is None:
        raise ValueError("The file must include both name and phone columns.")

    return name_index, phone_index, rows[1:]


def import_leads_from_upload(uploaded_file, *, assignment_mode="auto", assigned_staff=None):
    rows = _read_lead_rows_from_upload(uploaded_file)
    name_index, phone_index, data_rows = _detect_lead_column_indexes(rows)

    existing_numbers = {
        _normalize_phone(phone)
        for phone in Lead.objects.values_list("phone", flat=True)
    }
    created_leads = []
    batch_numbers = set()
    skipped_rows = 0

    for row in data_rows:
        row_values = list(row)
        if max(name_index, phone_index) >= len(row_values):
            skipped_rows += 1
            continue

        name = str(row_values[name_index] or "").strip()
        phone = _normalize_phone(row_values[phone_index])
        if not name or len(phone) < 7:
            skipped_rows += 1
            continue
        if phone in existing_numbers or phone in batch_numbers:
            skipped_rows += 1
            continue

        batch_numbers.add(phone)
        created_leads.append(
            Lead(
                name=name[:150],
                phone=phone,
                status=Lead.Status.NEW,
            )
        )

    if created_leads:
        Lead.objects.bulk_create(created_leads)

    selected_staff = list(assigned_staff or [])
    if assignment_mode == "selected_staff" and selected_staff:
        allocation = assign_imported_leads_to_staff(
            imported_lead_ids=[lead.id for lead in created_leads if getattr(lead, "id", None)],
            selected_staff=selected_staff,
        )
    else:
        allocation = auto_allocate_leads()

    return {
        "created_count": len(created_leads),
        "skipped_count": skipped_rows,
        "assigned_count": allocation["assigned_count"],
        "remaining_unassigned_count": allocation["remaining_unassigned_count"],
        "queue_limit": get_lead_queue_limit(),
        "assignment_mode": assignment_mode,
        "selected_staff_count": len(selected_staff),
        "selected_staff_names": [staff.name for staff in selected_staff],
        "released_count": allocation.get("released_count", 0),
    }


def _detect_followup_update_column_indexes(rows):
    if not rows:
        raise ValueError("The uploaded file is empty.")

    header = [_normalize_column_name(cell) for cell in rows[0]]
    indexes = {
        "lead_id": None,
        "phone": None,
        "status": None,
        "callback_window": None,
        "notes": None,
        "handover_status": None,
        "assigned_staff_phone": None,
        "assigned_staff_name": None,
    }

    for index, value in enumerate(header):
        if indexes["lead_id"] is None and value in LEAD_ID_COLUMN_ALIASES:
            indexes["lead_id"] = index
        if indexes["phone"] is None and value in PHONE_COLUMN_ALIASES:
            indexes["phone"] = index
        if indexes["status"] is None and value in STATUS_COLUMN_ALIASES:
            indexes["status"] = index
        if indexes["callback_window"] is None and value in CALLBACK_WINDOW_COLUMN_ALIASES:
            indexes["callback_window"] = index
        if indexes["notes"] is None and value in NOTES_COLUMN_ALIASES:
            indexes["notes"] = index
        if indexes["handover_status"] is None and value in HANDOVER_STATUS_COLUMN_ALIASES:
            indexes["handover_status"] = index
        if indexes["assigned_staff_phone"] is None and value in ASSIGNED_STAFF_PHONE_COLUMN_ALIASES:
            indexes["assigned_staff_phone"] = index
        if indexes["assigned_staff_name"] is None and value in ASSIGNED_STAFF_NAME_COLUMN_ALIASES:
            indexes["assigned_staff_name"] = index

    if indexes["lead_id"] is None and indexes["phone"] is None:
        raise ValueError("The update file must include either lead id or phone number.")

    editable_columns = (
        indexes["status"],
        indexes["callback_window"],
        indexes["notes"],
        indexes["handover_status"],
        indexes["assigned_staff_phone"],
        indexes["assigned_staff_name"],
    )
    if all(index is None for index in editable_columns):
        raise ValueError(
            "Include at least one update column: status, handover status, callback window, notes, or assigned staff."
        )

    return indexes, rows[1:]


def _cell_value(row_values, index):
    if index is None or index >= len(row_values):
        return ""
    return str(row_values[index] or "").strip()


def update_followups_from_upload(uploaded_file):
    rows = _read_lead_rows_from_upload(uploaded_file)
    indexes, data_rows = _detect_followup_update_column_indexes(rows)

    staff_by_phone = {
        _normalize_phone(staff.phone): staff
        for staff in _staff_queryset().filter(is_active=True)
    }
    staff_by_name = {
        _normalize_column_name(staff.name): staff
        for staff in _staff_queryset().filter(is_active=True)
    }

    updated_count = 0
    skipped_count = 0
    missing_count = 0
    error_messages = []

    for row_number, row in enumerate(data_rows, start=2):
        row_values = list(row)
        lead_id = _cell_value(row_values, indexes["lead_id"])
        phone = _normalize_phone(_cell_value(row_values, indexes["phone"]))
        if not lead_id and not phone:
            skipped_count += 1
            continue

        lead = None
        if lead_id:
            lead = Lead.objects.filter(id=lead_id).first()
        if lead is None and phone:
            lead = Lead.objects.filter(phone=phone).order_by("-updated_at").first()

        if lead is None:
            missing_count += 1
            continue

        payload = {}
        status_value = _cell_value(row_values, indexes["status"])
        if status_value:
            normalized_status = _normalize_status_value(status_value)
            if not normalized_status:
                error_messages.append(f"Row {row_number}: unknown status '{status_value}'.")
                continue
            payload["status"] = normalized_status

        callback_value = _cell_value(row_values, indexes["callback_window"])
        if indexes["callback_window"] is not None:
            normalized_callback = _normalize_callback_window_value(callback_value)
            if callback_value and not normalized_callback:
                error_messages.append(
                    f"Row {row_number}: callback slot must be Noon, Evening, or Night."
                )
                continue
            payload["callback_window"] = normalized_callback

        if indexes["notes"] is not None:
            payload["notes"] = _cell_value(row_values, indexes["notes"])

        handover_value = _cell_value(row_values, indexes["handover_status"])
        if indexes["handover_status"] is not None:
            normalized_handover = _normalize_handover_status_value(handover_value)
            if handover_value and not normalized_handover:
                error_messages.append(
                    f"Row {row_number}: unknown handover status '{handover_value}'."
                )
                continue
            if normalized_handover:
                payload["handover_status"] = normalized_handover

        assigned_staff_phone = _normalize_phone(
            _cell_value(row_values, indexes["assigned_staff_phone"])
        )
        assigned_staff_name = _normalize_column_name(
            _cell_value(row_values, indexes["assigned_staff_name"])
        )
        if indexes["assigned_staff_phone"] is not None or indexes["assigned_staff_name"] is not None:
            if assigned_staff_phone:
                assigned_staff = staff_by_phone.get(assigned_staff_phone)
                if not assigned_staff:
                    error_messages.append(
                        f"Row {row_number}: no active staff found for phone '{assigned_staff_phone}'."
                    )
                    continue
                payload["assigned_to"] = assigned_staff.id
            elif assigned_staff_name:
                assigned_staff = staff_by_name.get(assigned_staff_name)
                if not assigned_staff:
                    error_messages.append(
                        f"Row {row_number}: no active staff found for '{assigned_staff_name}'."
                    )
                    continue
                payload["assigned_to"] = assigned_staff.id
            else:
                payload["assigned_to"] = None

        if not payload:
            skipped_count += 1
            continue

        from backend.apps.telecalling.serializers import UpdateLeadSerializer

        serializer = UpdateLeadSerializer(lead, data=payload, partial=True)
        if not serializer.is_valid():
            error_messages.append(
                f"Row {row_number}: {' '.join(str(value[0]) if isinstance(value, list) else str(value) for value in serializer.errors.values())}"
            )
            continue

        serializer.save()
        updated_count += 1

    auto_allocate_leads()
    return {
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "missing_count": missing_count,
        "error_messages": error_messages[:10],
    }


def _training_lessons_queryset():
    return TrainingLesson.objects.order_by("sort_order", "-published_at", "title")


def _active_training_lessons_queryset():
    return _training_lessons_queryset().filter(is_active=True)


def _completion_map_for_staff(staff):
    return {
        completion.lesson_id: completion
        for completion in TrainingCompletion.objects.filter(staff=staff).select_related("lesson")
    }


def get_pending_mandatory_lessons(staff):
    completed_lesson_ids = TrainingCompletion.objects.filter(staff=staff).values_list("lesson_id", flat=True)
    return _active_training_lessons_queryset().filter(is_mandatory=True).exclude(id__in=completed_lesson_ids)


def build_staff_learning_payload(staff):
    lessons = list(_active_training_lessons_queryset())
    completion_map = _completion_map_for_staff(staff)
    pending_mandatory_count = 0
    next_required_title = ""
    completed_count = 0
    lesson_rows = []

    for lesson in lessons:
        completion = completion_map.get(lesson.id)
        is_completed = completion is not None
        if is_completed:
            completed_count += 1
        if lesson.is_mandatory and not is_completed:
            pending_mandatory_count += 1
            if not next_required_title:
                next_required_title = lesson.title

        lesson_rows.append(
            {
                "id": str(lesson.id),
                "title": lesson.title,
                "description": lesson.description,
                "video_url": lesson.video_url,
                "search_keywords": lesson.search_keywords,
                "is_active": lesson.is_active,
                "is_mandatory": lesson.is_mandatory,
                "is_completed": is_completed,
                "completed_at": completion.completed_at.isoformat() if completion else None,
                "published_at": lesson.published_at.isoformat() if lesson.published_at else None,
                "published_at_label": _format_datetime(lesson.published_at),
            }
        )

    return {
        "summary": {
            "total_lessons": len(lesson_rows),
            "completed_count": completed_count,
            "pending_mandatory_count": pending_mandatory_count,
            "has_pending_mandatory": pending_mandatory_count > 0,
            "next_required_title": next_required_title,
        },
        "lessons": lesson_rows,
    }


def build_learning_management_payload():
    today = timezone.localdate()
    active_staff_count = _staff_queryset().filter(is_active=True).count()
    completion_counts = {
        row["lesson_id"]: row["count"]
        for row in TrainingCompletion.objects.values("lesson_id").annotate(count=Count("id"))
    }
    lesson_rows = []
    active_count = 0
    mandatory_count = 0
    total_completions = TrainingCompletion.objects.count()

    for lesson in _training_lessons_queryset():
        completed_staff_count = completion_counts.get(lesson.id, 0)
        pending_staff_count = max(active_staff_count - completed_staff_count, 0)
        if lesson.is_active:
            active_count += 1
        if lesson.is_active and lesson.is_mandatory:
            mandatory_count += 1

        if not lesson.is_active:
            status_filter = "inactive"
            status_label = "Inactive"
        elif lesson.is_mandatory:
            status_filter = "mandatory"
            status_label = "Required"
        else:
            status_filter = "optional"
            status_label = "Optional"

        lesson_rows.append(
            {
                "id": str(lesson.id),
                "title": lesson.title,
                "description": lesson.description,
                "video_url": lesson.video_url,
                "search_keywords": lesson.search_keywords,
                "is_active": lesson.is_active,
                "is_mandatory": lesson.is_mandatory,
                "sort_order": lesson.sort_order,
                "published_at": lesson.published_at,
                "published_at_label": _format_datetime(lesson.published_at),
                "completed_staff_count": completed_staff_count,
                "pending_staff_count": pending_staff_count,
                "status_filter": status_filter,
                "status_label": status_label,
            }
        )

    return {
        "today_label": today.strftime("%A, %d %b %Y"),
        "summary": {
            "active_count": active_count,
            "mandatory_count": mandatory_count,
            "active_staff_count": active_staff_count,
            "total_completions": total_completions,
        },
        "lesson_rows": lesson_rows,
    }


def complete_training_lesson(staff, lesson):
    now = timezone.now()
    completion, created = TrainingCompletion.objects.get_or_create(
        staff=staff,
        lesson=lesson,
        defaults={"completed_at": now},
    )
    if not created and not completion.completed_at:
        completion.completed_at = now
        completion.save(update_fields=["completed_at"])

    mark_staff_seen(staff, now)
    _log_staff_action(
        staff,
        StaffAction.ActionType.TRAINING_COMPLETED,
        metadata={
            "lesson_id": str(lesson.id),
            "lesson_title": lesson.title,
            "created": created,
        },
    )
    return completion


def _open_sessions_by_staff():
    open_sessions = {}
    for session in Session.objects.select_related("staff").filter(is_open=True).order_by("-login_time"):
        session = reconcile_session(session)
        if session and session.is_open and session.staff_id not in open_sessions:
            open_sessions[session.staff_id] = session
    return open_sessions


def _currently_active_staff_ids(now=None):
    now = now or timezone.now()
    active_cutoff = now - timedelta(seconds=ONLINE_WINDOW_SECONDS)
    active_staff_ids = {
        staff_id
        for staff_id, session in _open_sessions_by_staff().items()
        if session.last_known_state == Session.AppState.FOREGROUND
        and session.last_heartbeat_at
        and session.last_heartbeat_at >= active_cutoff
    }
    active_staff_ids.update(_live_call_staff_ids())
    return active_staff_ids


def _live_call_staff_ids():
    return _reconcile_open_calls()


def _release_due_callback_leads_from_inactive_staff(*, now=None, active_staff_ids=None):
    return 0


def build_dashboard_payload():
    today, start, end = _today_range()
    staff_queryset = Staff.objects.filter(is_active=True, role=Staff.Role.STAFF)
    staff_ids = list(staff_queryset.values_list("id", flat=True))
    active_cutoff = timezone.now() - timedelta(seconds=ONLINE_WINDOW_SECONDS)
    week_start, week_end = _week_range()
    month_start, month_end = _month_range()

    open_sessions = _open_sessions_by_staff()
    live_call_staff_ids = _live_call_staff_ids()

    active_staff = sum(
        1
        for staff in staff_queryset
        if _staff_online_label(
            open_sessions.get(staff.id),
            active_cutoff,
            is_in_customer_call=staff.id in live_call_staff_ids,
        )
        in {"Online", "On Call"}
    )
    total_staff = staff_queryset.count()

    leads = Lead.objects.select_related("assigned_to")
    calls_today = Call.objects.filter(start_time__range=(start, end))
    qualifying_calls_today = calls_today.filter(is_qualifying=True)
    sessions_today = Session.objects.filter(login_time__range=(start, end))

    total_leads = leads.count()
    follow_up_count = leads.filter(status=Lead.Status.INTERESTED).count()
    converted_count = leads.filter(status=Lead.Status.CONVERTED).count()
    no_response_count = leads.filter(status=Lead.Status.NO_ANSWER).count()
    callbacks_count = leads.filter(status=Lead.Status.CALL_BACK).count()

    calls_today_count = qualifying_calls_today.count()
    converted_calls_today = qualifying_calls_today.filter(status=Call.Status.CONVERTED).count()
    conversion_rate = round((converted_calls_today / calls_today_count) * 100, 1) if calls_today_count else 0
    work_coverage_pct = round((active_staff / total_staff) * 100) if total_staff else 0
    short_calls_blocked = calls_today.filter(status=Call.Status.INVALID_SHORT).count()

    session_totals = _effective_active_seconds_map(start_at=start, end_at=end, staff_ids=staff_ids)
    active_seconds_today = sum(session_totals.values())
    salary_ready = sessions_today.values("staff_id").distinct().count()

    converted_counter = Counter(
        calls_today.filter(status=Call.Status.CONVERTED).values_list("staff_id", flat=True)
    )
    call_totals = {
        row["staff_id"]: row["total"] or 0
        for row in qualifying_calls_today.values("staff_id").annotate(total=Sum("duration_seconds"))
    }
    month_session_totals, month_call_totals, month_converted, month_bonus_summaries = _staff_period_totals(month_start, month_end)
    week_session_totals, week_call_totals, week_converted, week_bonus_summaries = _staff_period_totals(week_start, week_end)
    salary_estimate = Decimal("0.00")
    for staff in staff_queryset:
        weekly_breakdown = calculate_staff_payout(
            staff,
            active_seconds=week_session_totals.get(staff.id, 0),
            call_seconds=week_call_totals.get(staff.id, 0),
            converted_leads=week_converted.get(staff.id, 0),
            bonus_calls=week_bonus_summaries.get(staff.id, _empty_hourly_bonus_summary())["completed_bonus_calls"],
            hourly_call_bonus_summary=week_bonus_summaries.get(staff.id, _empty_hourly_bonus_summary()),
        )
        monthly_breakdown = calculate_staff_payout(
            staff,
            active_seconds=month_session_totals.get(staff.id, 0),
            call_seconds=month_call_totals.get(staff.id, 0),
            converted_leads=month_converted.get(staff.id, 0),
            bonus_calls=month_bonus_summaries.get(staff.id, _empty_hourly_bonus_summary())["completed_bonus_calls"],
            hourly_call_bonus_summary=month_bonus_summaries.get(staff.id, _empty_hourly_bonus_summary()),
        )
        current_payable, _ = _current_cycle_payout(staff, weekly_breakdown, monthly_breakdown)
        salary_estimate += current_payable

    trend_rows = (
        Call.objects.filter(start_time__date__gte=today - timedelta(days=6), start_time__date__lte=today)
        .annotate(day=TruncDate("start_time"))
        .values("day")
        .annotate(
            call_count=Count("id", filter=Q(is_qualifying=True)),
            conversion_count=Count("id", filter=Q(status=Call.Status.CONVERTED, is_qualifying=True)),
        )
        .order_by("day")
    )
    trend_map = {row["day"]: row for row in trend_rows}
    trend_labels = []
    trend_calls = []
    trend_conversions = []
    for offset in range(6, -1, -1):
        day = today - timedelta(days=offset)
        trend_labels.append(day.strftime("%a"))
        trend_calls.append(trend_map.get(day, {}).get("call_count", 0))
        trend_conversions.append(trend_map.get(day, {}).get("conversion_count", 0))

    lead_pipeline = {
        "labels": ["New", "Follow Up", "Scheduled Follow Up", "No Response", "Converted"],
        "values": [
            leads.filter(status=Lead.Status.NEW).count(),
            follow_up_count,
            callbacks_count,
            no_response_count,
            converted_count,
        ],
    }

    top_staff = []
    for staff in staff_queryset.order_by("name")[:5]:
        top_staff.append(
            {
                "label": staff.name.split()[0],
                "active_hours": round((session_totals.get(staff.id, 0) or 0) / 3600, 1),
                "call_minutes": round((call_totals.get(staff.id, 0) or 0) / 60),
            }
        )

    live_staff = []
    for staff in staff_queryset.order_by("name")[:8]:
        session = open_sessions.get(staff.id)
        is_in_customer_call = staff.id in live_call_staff_ids
        if session:
            status_label = "On customer call" if is_in_customer_call else _session_status_label(session)
            session_hours = round((session.active_seconds or 0) / 3600, 1)
            status_text = f"{status_label} - {session_hours}h active"
        else:
            status_text = "On customer call" if is_in_customer_call else "Offline - No active session"

        live_staff.append(
            {
                "name": staff.name,
                "status_text": status_text,
                "is_online": _staff_online_label(
                    session,
                    active_cutoff,
                    is_in_customer_call=is_in_customer_call,
                )
                in {"Online", "On Call"},
            }
        )

    lead_rows = []
    for lead in leads.order_by("-updated_at")[:20]:
        lead_rows.append(
            {
                "name": lead.name,
                "phone": lead.phone,
                "status": lead.status,
                "status_label": lead.get_status_display(),
                "assigned_to": lead.assigned_to.name if lead.assigned_to else "Unassigned",
                "last_call": timezone.localtime(lead.last_contacted_at).strftime("%I:%M %p") if lead.last_contacted_at else "Not called yet",
                "next_action": {
                    Lead.Status.NEW: "Assign and dial",
                    Lead.Status.CALL_BACK: "Follow up on schedule",
                    Lead.Status.INTERESTED: "Complete the follow-up",
                    Lead.Status.CONVERTED: "Move to onboarding",
                    Lead.Status.NO_ANSWER: "Retry if needed",
                    Lead.Status.NOT_INTERESTED: "Keep for reporting",
                }.get(lead.status, "Review"),
            }
        )

    salary_records = []
    for staff in staff_queryset.order_by("name")[:6]:
        staff_hours = round((session_totals.get(staff.id, 0) or 0) / 3600, 1)
        staff_call_minutes = round((call_totals.get(staff.id, 0) or 0) / 60)
        weekly_breakdown = calculate_staff_payout(
            staff,
            active_seconds=week_session_totals.get(staff.id, 0),
            call_seconds=week_call_totals.get(staff.id, 0),
            converted_leads=week_converted.get(staff.id, 0),
            bonus_calls=week_bonus_summaries.get(staff.id, _empty_hourly_bonus_summary())["completed_bonus_calls"],
            hourly_call_bonus_summary=week_bonus_summaries.get(staff.id, _empty_hourly_bonus_summary()),
        )
        monthly_breakdown = calculate_staff_payout(
            staff,
            active_seconds=month_session_totals.get(staff.id, 0),
            call_seconds=month_call_totals.get(staff.id, 0),
            converted_leads=month_converted.get(staff.id, 0),
            bonus_calls=month_bonus_summaries.get(staff.id, _empty_hourly_bonus_summary())["completed_bonus_calls"],
            hourly_call_bonus_summary=month_bonus_summaries.get(staff.id, _empty_hourly_bonus_summary()),
        )
        current_payable, cycle_label = _current_cycle_payout(staff, weekly_breakdown, monthly_breakdown)
        salary_records.append(
            {
                "name": staff.name,
                "hours": f"{staff_hours}h",
                "call_time": f"{staff_call_minutes}m",
                "bonus": _format_currency(monthly_breakdown["bonus_earnings"]),
                "final_salary": _format_currency(current_payable),
                "cycle_label": cycle_label,
            }
        )

    team_directory = []
    for staff in staff_queryset.order_by("name"):
        session = open_sessions.get(staff.id)
        team_directory.append(
            {
                "id": str(staff.id),
                "name": staff.name,
                "phone": staff.phone,
                "compensation_type": staff.compensation_type,
                "compensation_type_label": staff.get_compensation_type_display(),
                "hourly_rate": _format_currency(staff.hourly_rate),
                "weekly_salary": _format_currency(staff.weekly_salary),
                "monthly_salary": _format_currency(staff.monthly_salary),
                "target_hours_per_week": float(staff.target_hours_per_week),
                "target_hours_per_month": float(staff.target_hours_per_month),
                "call_rate": _format_currency(staff.call_rate),
                "bonus_per_conversion": _format_currency(staff.bonus_per_conversion),
                "is_active": staff.is_active,
                "online_label": _staff_online_label(
                    session,
                    active_cutoff,
                    is_in_customer_call=staff.id in live_call_staff_ids,
                ),
            }
        )

    dashboard = {
        "today_label": today.strftime("%A, %d %b %Y"),
        "staff_active": active_staff,
        "calls_today": calls_today_count,
        "conversion_rate": f"{conversion_rate:.1f}%",
        "callbacks": callbacks_count,
        "work_coverage": work_coverage_pct,
        "short_calls_blocked": short_calls_blocked,
        "salary_ready": salary_ready,
        "total_leads": total_leads,
        "no_answer": no_response_count,
        "interested": follow_up_count,
        "converted": converted_count,
        "salary_estimate": _format_currency(salary_estimate),
        "active_hours": _format_hours(active_seconds_today),
    }

    chart_payload = {
        "callVolume": {
            "labels": trend_labels,
            "calls": trend_calls,
            "conversions": trend_conversions,
        },
        "leadPipeline": lead_pipeline,
        "activityBalance": {
            "labels": [item["label"] for item in top_staff],
            "activeHours": [item["active_hours"] for item in top_staff],
            "callMinutes": [item["call_minutes"] for item in top_staff],
        },
    }

    return {
        "dashboard": dashboard,
        "chart_payload": chart_payload,
        "live_staff": live_staff,
        "lead_rows": lead_rows,
        "salary_records": salary_records,
        "team_directory": team_directory,
    }


def _current_open_call_map(*, now=None):
    current_time = now or timezone.now()
    _reconcile_open_calls(now=current_time)
    open_calls = {}
    queryset = (
        Call.objects.filter(end_time__isnull=True)
        .select_related("lead", "staff")
        .order_by("staff_id", "-start_time", "-created_at")
    )
    for call in queryset:
        if call.staff_id in open_calls:
            continue
        duration_seconds = max(0, int((current_time - call.start_time).total_seconds()))
        open_calls[call.staff_id] = {
            "call_id": str(call.id),
            "lead_id": str(call.lead_id),
            "lead_name": call.lead.name or "Lead",
            "lead_phone": call.lead.phone or "--",
            "started_at": _format_datetime(call.start_time),
            "duration_label": _format_duration(duration_seconds),
            "status_label": call.get_status_display(),
        }
    return open_calls


def build_live_monitoring_payload():
    today, start, end = _today_range()
    now = timezone.now()
    team_payload = build_team_management_payload()
    team_rows = team_payload["team_rows"]
    staff_ids = [uuid.UUID(str(row["id"])) for row in team_rows]
    gap_summary_map = _work_gap_summary_map(start_at=start, end_at=end, staff_ids=staff_ids)
    open_sessions = _open_sessions_by_staff()
    current_open_calls = _current_open_call_map(now=now)

    on_call_now = 0
    online_now = 0
    away_now = 0
    offline_now = 0
    active_hours_seconds = 0
    review_needed_now = 0
    alert_now = 0
    live_rows = []

    for row in team_rows:
        staff_id = uuid.UUID(str(row["id"]))
        is_active_account = bool(row.get("is_active", True))
        gap_summary = gap_summary_map.get(staff_id, {})
        session = open_sessions.get(staff_id)
        current_call = current_open_calls.get(staff_id)
        online_label = row.get("online_label", "Offline")
        active_seconds_today = int(row.get("active_seconds_today") or 0)
        calls_today = int(row.get("calls_today") or 0)
        quality_label = row.get("quality_label", "No Recent Activity")
        needs_review = quality_label in {"Review Needed", "Needs Attention"}
        has_worked_today = active_seconds_today > 0 or calls_today > 0
        is_live_now = bool(current_call) or online_label in {"Online", "Away", "Warning"}
        if current_call:
            on_call_now += 1
        elif online_label == "Online":
            online_now += 1
        elif online_label in {"Away", "Warning"}:
            away_now += 1
        else:
            offline_now += 1

        if quality_label == "Review Needed":
            review_needed_now += 1
        elif quality_label == "Needs Attention":
            alert_now += 1

        active_hours_seconds += active_seconds_today
        status_note = row.get("quality_note") or "Live staff activity is being monitored."
        if current_call:
            status_note = f"Talking with {current_call['lead_name']} for {current_call['duration_label']}."
        elif online_label == "Online":
            status_note = "Ready in the app and available for live calling."
        elif online_label in {"Away", "Warning"}:
            status_note = "Session is open, but the staff member is away from the app right now."
        elif has_worked_today:
            status_note = "Worked today and is currently offline from the app."

        if is_active_account and (has_worked_today or is_live_now or needs_review):
            live_rows.append(
                {
                    "id": row["id"],
                    "name": row["name"],
                    "phone": row["phone"],
                    "is_active": bool(row.get("is_active", True)),
                    "compensation_type_label": row.get("compensation_type_label", "Hourly"),
                    "online_label": online_label,
                    "status_tone": row.get("status_tone", "muted"),
                    "session_state": row.get("session_state", "stopped"),
                    "session_state_label": _session_status_label(session),
                    "active_hours_today": row.get("active_hours_today", "0.0h"),
                    "calls_today": calls_today,
                    "converted_today": int(row.get("converted_today") or 0),
                    "assigned_leads": int(row.get("assigned_leads") or 0),
                    "quality_score": int(row.get("quality_score") or 0),
                    "quality_label": quality_label,
                    "quality_tone": row.get("quality_tone", "muted"),
                    "quality_note": row.get("quality_note", "Build more recent call activity for a fuller review."),
                    "outcome_consistency_label": row.get("outcome_consistency_label", "--"),
                    "attempt_review_label": row.get("attempt_review_label", "--"),
                    "away_review_label": row.get("away_review_label", "No long away periods"),
                    "long_away_count": int(row.get("long_away_count") or 0),
                    "real_call_count": int(row.get("real_call_count") or 0),
                    "verified_attempt_count": int(row.get("verified_attempt_count") or 0),
                    "zero_second_attempt_count": int(row.get("zero_second_attempt_count") or 0),
                    "invalid_short_count": int(row.get("invalid_short_count") or 0),
                    "missed_callbacks": int(row.get("missed_callbacks") or 0),
                    "suspicious_block_count": int(row.get("suspicious_block_count") or 0),
                    "zero_only_block_count": int(row.get("zero_only_block_count") or 0),
                    "gap_count": int(gap_summary.get("gap_count") or 0),
                    "gap_total_label": gap_summary.get("gap_total_label", "0s"),
                    "gap_uncounted_label": gap_summary.get("gap_uncounted_label", "0s"),
                    "gap_buffer_label": gap_summary.get("gap_buffer_label", "0s"),
                    "call_time_label": gap_summary.get("call_time_label", "0s"),
                    "last_seen": row.get("last_seen", "--"),
                    "status_note": status_note,
                    "is_on_call": bool(current_call),
                    "worked_today": has_worked_today,
                    "needs_review": needs_review,
                    "current_call": current_call,
                }
            )

    live_rows.sort(
        key=lambda row: (
            0 if row["is_on_call"] else 1,
            0 if row["online_label"] == "Online" else 1,
            0 if row["worked_today"] else 1,
            0 if row["needs_review"] else 1,
            -row["calls_today"],
            -row["quality_score"],
            row["name"].lower(),
        )
    )

    spotlight_rows = live_rows[:6]
    return {
        "today_label": today.strftime("%A, %d %b %Y"),
        "generated_at_label": _format_datetime(now),
        "summary": {
            "total_staff": team_payload["team_summary"]["total_staff"],
            "active_accounts": team_payload["team_summary"]["active_accounts"],
            "online_now": online_now,
            "on_call_now": on_call_now,
            "away_now": away_now,
            "offline_now": offline_now,
            "review_needed_now": review_needed_now,
            "alert_now": alert_now,
            "total_assigned": team_payload["team_summary"]["total_assigned"],
            "total_calls_today": team_payload["team_summary"]["total_calls_today"],
            "total_converted_today": team_payload["team_summary"]["total_converted_today"],
            "active_hours_label": _format_hours(active_hours_seconds),
        },
        "spotlight_rows": spotlight_rows,
        "staff_rows": live_rows,
    }


def build_team_management_payload(*, quality_range_start=None, quality_range_end=None, rules=None):
    today, start, end = _today_range()
    active_cutoff = timezone.now() - timedelta(seconds=ONLINE_WINDOW_SECONDS)
    active_rules = rules or _work_review_rules()
    staff_queryset = _staff_queryset()
    staff_ids = [staff.id for staff in staff_queryset]
    open_sessions = _open_sessions_by_staff()
    live_call_staff_ids = _live_call_staff_ids()
    quality_by_staff = _build_staff_quality_metrics(
        staff_ids,
        range_start=quality_range_start,
        range_end=quality_range_end,
        rules=active_rules,
    )

    call_totals = {
        row["staff_id"]: row["count"]
        for row in Call.objects.filter(start_time__range=(start, end))
        .values("staff_id")
        .annotate(count=Count("id"))
    }
    converted_totals = {
        row["staff_id"]: row["count"]
        for row in Call.objects.filter(start_time__range=(start, end), status=Call.Status.CONVERTED)
        .values("staff_id")
        .annotate(count=Count("id"))
    }
    active_totals = _effective_active_seconds_map(
        start_at=start,
        end_at=end,
        staff_ids=staff_ids,
        rules=active_rules,
    )
    assigned_totals = {
        row["assigned_to"]: row["count"]
        for row in _lead_queue_queryset().exclude(assigned_to=None)
        .values("assigned_to")
        .annotate(count=Count("id"))
    }

    team_rows = []
    total_assigned = 0
    total_calls_today = 0
    total_converted_today = 0
    active_accounts = 0
    online_now = 0
    attention_needed = 0
    for staff in staff_queryset:
        session = open_sessions.get(staff.id)
        online_label = _staff_online_label(
            session,
            active_cutoff,
            is_in_customer_call=staff.id in live_call_staff_ids,
        )
        calls_today = call_totals.get(staff.id, 0)
        converted_today = converted_totals.get(staff.id, 0)
        assigned_leads = assigned_totals.get(staff.id, 0)
        quality = quality_by_staff.get(staff.id, {})
        is_available = staff.is_active and online_label in {"Online", "On Call"}
        status_filter = "inactive"
        status_tone = "muted"
        if staff.is_active:
            active_accounts += 1
            if not staff.receives_new_leads:
                attention_needed += 1
                status_filter = "paused"
                status_tone = "warning"
            elif online_label in {"Online", "On Call"}:
                online_now += 1
                status_filter = "online"
                status_tone = "success"
            elif online_label in {"Away", "Warning"}:
                attention_needed += 1
                status_filter = "attention"
                status_tone = "warning"
            else:
                attention_needed += 1
                status_filter = "offline"
                status_tone = "muted"

        total_assigned += assigned_leads
        total_calls_today += calls_today
        total_converted_today += converted_today
        team_rows.append(
            {
                "id": str(staff.id),
                "name": staff.name,
                "phone": staff.phone,
                "is_active": staff.is_active,
                "receives_new_leads": staff.receives_new_leads,
                "compensation_type": staff.compensation_type,
                "compensation_type_label": staff.get_compensation_type_display(),
                "hourly_rate": _format_currency(staff.hourly_rate),
                "weekly_salary": _format_currency(staff.weekly_salary),
                "monthly_salary": _format_currency(staff.monthly_salary),
                "target_hours_per_week": float(staff.target_hours_per_week),
                "target_hours_per_month": float(staff.target_hours_per_month),
                "call_rate": _format_currency(staff.call_rate),
                "bonus_per_conversion": _format_currency(staff.bonus_per_conversion),
                "online_label": online_label,
                "status_filter": status_filter,
                "status_tone": status_tone,
                "lead_intake_label": "Lead intake paused" if staff.is_active and not staff.receives_new_leads else "Receiving new leads",
                "lead_intake_tone": "warning" if staff.is_active and not staff.receives_new_leads else "success",
                "session_state": session.last_known_state if session else "stopped",
                "active_hours_today": _format_hours(active_totals.get(staff.id, 0)),
                "active_seconds_today": active_totals.get(staff.id, 0),
                "calls_today": calls_today,
                "converted_today": converted_today,
                "assigned_leads": assigned_leads,
                "last_seen": _format_datetime(staff.last_seen_at),
                "is_available": is_available,
                "quality_score": quality.get("score", 0),
                "quality_label": quality.get("label", "No Recent Activity"),
                "quality_tone": quality.get("tone", "muted"),
                "quality_note": quality.get("note", "Build more recent call activity for a fuller review."),
                "invalid_short_count": quality.get("invalid_short_count", 0),
                "outcome_consistency_label": quality.get("outcome_consistency_label", "--"),
                "missed_callbacks": quality.get("missed_callbacks", 0),
                "verified_attempt_count": quality.get("verified_attempt_count", 0),
                "attempt_review_label": quality.get("attempt_review_label", "--"),
                "away_review_label": quality.get("away_review_label", "Within limit"),
                "long_away_count": quality.get("long_away_count", 0),
                "suspicious_block_count": quality.get("suspicious_block_count", 0),
                "zero_only_block_count": quality.get("zero_only_block_count", 0),
                "zero_second_attempt_count": quality.get("zero_second_attempt_count", 0),
                "real_call_count": quality.get("real_call_count", 0),
                "expired_followup_count": quality.get("expired_followup_count", 0),
                "expired_followup_penalty_points": quality.get("expired_followup_penalty_points", 0),
                "total_penalty_points": quality.get("total_penalty_points", 0),
            }
        )

    team_rows.sort(
        key=lambda row: (
            0 if row.get("is_active") else 1,
            0 if row.get("status_filter") == "online" else 1,
            0 if row.get("status_filter") == "attention" else 1,
            row.get("name", "").lower(),
        )
    )

    return {
        "today_label": today.strftime("%A, %d %b %Y"),
        "team_summary": {
            "total_staff": len(team_rows),
            "active_accounts": active_accounts,
            "lead_intake_paused": sum(1 for row in team_rows if row.get("is_active") and not row.get("receives_new_leads")),
            "online_now": online_now,
            "attention_needed": attention_needed,
            "total_assigned": total_assigned,
            "total_calls_today": total_calls_today,
            "total_converted_today": total_converted_today,
        },
        "team_rows": team_rows,
    }


def build_work_review_payload(*, search_query="", review_filter="all", now=None, month_value=""):
    rules = _work_review_rules()
    attempt_threshold = int(rules.get("attempt_threshold", MIN_REAL_CALLS_PER_ATTEMPT_BLOCK) or 1)
    reference_date = _parse_month_value(month_value)
    reference_date = reference_date or timezone.localdate()
    range_start, range_end = _month_range_for_reference(
        reference_date,
        end_at=timezone.localtime(now or timezone.now()),
    )
    team_payload = build_team_management_payload(
        quality_range_start=range_start,
        quality_range_end=range_end,
        rules=rules,
    )
    team_rows = list(team_payload["team_rows"])
    staff_ids = []
    for row in team_rows:
        try:
            staff_ids.append(uuid.UUID(str(row["id"])))
        except (TypeError, ValueError, AttributeError):
            continue

    day_previews = _build_work_review_day_previews(staff_ids, now=now, rules=rules)
    normalized_query = " ".join(str(search_query or "").strip().lower().split())
    review_filter = str(review_filter or "all").strip().lower() or "all"
    if review_filter not in {"all", "review", "attention", "stable", "quiet"}:
        review_filter = "all"

    review_needed_count = 0
    attention_count = 0
    stable_count = 0
    quiet_count = 0
    zero_talk_total = 0
    invalid_short_total = 0
    missed_callbacks_total = 0
    flagged_day_total = 0
    expired_followup_total = 0

    review_rows = []
    zero_talk_blocks = _zero_talk_block_details_by_staff(
        staff_ids,
        range_start=range_start,
        range_end=range_end,
        block_limit=1000,
        include_unverified=True,
        include_invalid_short=True,
        rules=rules,
    )

    for row in team_rows:
        staff_uuid = None
        try:
            staff_uuid = uuid.UUID(str(row["id"]))
        except (TypeError, ValueError, AttributeError):
            staff_uuid = None
        zero_streak_blocks = (zero_talk_blocks.get(staff_uuid) or {}).get("blocks", [])
        quality_label = str(row.get("quality_label") or "")
        verified_attempt_count = int(row.get("verified_attempt_count") or 0)
        real_call_count = int(row.get("real_call_count") or 0)
        zero_second_attempt_count = int(row.get("zero_second_attempt_count") or 0)
        invalid_short_count = int(row.get("invalid_short_count") or 0)
        suspicious_block_count = int(row.get("suspicious_block_count") or 0)
        zero_only_block_count = int(row.get("zero_only_block_count") or 0)
        missed_callbacks = int(row.get("missed_callbacks") or 0)
        expired_followup_count = int(row.get("expired_followup_count") or 0)

        review_state = "stable"
        review_state_label = "Stable"
        review_state_tone = "success"
        review_state_note = "Recent verified activity looks healthy."
        if quality_label == "No Recent Activity" and verified_attempt_count <= 0:
            review_state = "quiet"
            review_state_label = "No Recent Activity"
            review_state_tone = "muted"
            review_state_note = "No recent verified activity is available for review."
        elif (
            quality_label == "Review Needed"
            or suspicious_block_count > 0
            or zero_only_block_count > 0
            or zero_streak_blocks
            or (verified_attempt_count >= attempt_threshold and real_call_count == 0)
        ):
            review_state = "review"
            review_state_label = "Review Needed"
            review_state_tone = "danger"
            review_state_note = "Review calling patterns, empty attempt blocks, and low real-call activity."
        elif (
            quality_label == "Needs Attention"
            or invalid_short_count > 0
            or zero_second_attempt_count > 0
            or missed_callbacks > 0
            or expired_followup_count > 0
        ):
            review_state = "attention"
            review_state_label = "Needs Attention"
            review_state_tone = "warning"
            review_state_note = "Some calling patterns should be reviewed before final decisions are made."

        if review_state == "review":
            review_needed_count += 1
        elif review_state == "attention":
            attention_count += 1
        elif review_state == "quiet":
            quiet_count += 1
        else:
            stable_count += 1

        zero_talk_total += zero_only_block_count
        invalid_short_total += invalid_short_count
        missed_callbacks_total += missed_callbacks
        expired_followup_total += expired_followup_count

        preview_payload = day_previews.get(uuid.UUID(str(row["id"])), {"count": 0, "rows": []})
        flagged_day_total += int(preview_payload.get("count") or 0)
        review_day_rows = preview_payload.get("rows") or []
        extra_review_day_count = max(int(preview_payload.get("count") or 0) - len(review_day_rows), 0)
        if verified_attempt_count > 0:
            ratio_value = (Decimal(real_call_count) / Decimal(verified_attempt_count)) * Decimal("100")
            real_call_ratio_label = f"{int(ratio_value.quantize(Decimal('1')))}%"
        else:
            real_call_ratio_label = "--"

        review_row = {
            **row,
            "review_state": review_state,
            "review_state_label": review_state_label,
            "review_state_tone": review_state_tone,
            "review_state_note": review_state_note,
            "real_call_ratio_label": real_call_ratio_label,
            "review_day_rows": review_day_rows,
            "review_day_count": int(preview_payload.get("count") or 0),
            "extra_review_day_count": extra_review_day_count,
            "zero_only_block_details": zero_streak_blocks,
            "zero_only_block_extra": (zero_talk_blocks.get(staff_uuid) or {}).get("extra_count", 0),
            "search_text": " ".join(
                part.lower()
                for part in (
                    row.get("name", ""),
                    row.get("phone", ""),
                    quality_label,
                    row.get("quality_note", ""),
                    review_state_label,
                    review_state_note,
                    row.get("attempt_review_label", ""),
                    row.get("away_review_label", ""),
                    f"{expired_followup_count} expired follow ups",
                )
                if part
            ),
        }

        if normalized_query and normalized_query not in review_row["search_text"]:
            continue
        if review_filter != "all" and review_state != review_filter:
            continue
        review_rows.append(review_row)

    state_rank = {"review": 0, "attention": 1, "stable": 2, "quiet": 3}
    review_rows.sort(
        key=lambda row: (
            state_rank.get(row["review_state"], 9),
            int(row.get("quality_score") or 0),
            -int(row.get("verified_attempt_count") or 0),
            str(row.get("name") or "").lower(),
        )
    )

    zero_talk_staff_rows = []
    for row in team_rows:
        try:
            staff_uuid = uuid.UUID(str(row["id"]))
        except (TypeError, ValueError, AttributeError):
            continue
        blocks = (zero_talk_blocks.get(staff_uuid) or {}).get("blocks", [])
        if not blocks:
            continue
        zero_talk_staff_rows.append(
            {
                "id": row.get("id"),
                "name": row.get("name"),
                "phone": row.get("phone"),
                "block_count": len(blocks),
                "blocks": blocks,
            }
        )

    return {
        "today_label": team_payload["today_label"],
        "lookback_days": QUALITY_SCORE_LOOKBACK_DAYS,
        "search_query": search_query,
        "review_filter": review_filter,
        "month_value": reference_date.strftime("%Y-%m"),
        "month_options": _month_option_rows(
            reference_date=timezone.localdate(),
            selected_value=reference_date.strftime("%Y-%m"),
            months_back=6,
        ),
        "period_label": reference_date.strftime("%b %Y"),
        "work_review_rules": {
            "attempt_threshold": attempt_threshold,
            "idle_gap_seconds": int(rules.get("idle_gap_seconds", CALL_ACTIVITY_IDLE_BREAK_SECONDS) or 1),
            "connected_cooldown_seconds": int(
                rules.get("connected_cooldown_seconds", CONNECTED_CALL_COOLDOWN_SECONDS) or 0
            ),
            "followup_expired_penalty_points": int(
                rules.get("followup_expired_penalty_points", FOLLOWUP_EXPIRED_SCORE_PENALTY_POINTS) or 0
            ),
            "followup_expired_penalty_cap": int(
                rules.get("followup_expired_penalty_cap", FOLLOWUP_EXPIRED_SCORE_PENALTY_CAP) or 0
            ),
        },
        "review_summary": {
            "total_staff": len(team_rows),
            "filtered_staff_count": len(review_rows),
            "review_needed_count": review_needed_count,
            "attention_count": attention_count,
            "stable_count": stable_count,
            "quiet_count": quiet_count,
            "zero_talk_total": zero_talk_total,
            "invalid_short_total": invalid_short_total,
            "missed_callbacks_total": missed_callbacks_total,
            "flagged_day_total": flagged_day_total,
            "expired_followup_total": expired_followup_total,
        },
        "review_rows": review_rows,
        "zero_talk_staff_rows": zero_talk_staff_rows,
    }


def build_referral_monitoring_payload(*, search_query="", stage_filter="all", reward_filter="all"):
    company_profile = get_company_profile()
    sync_referral_rewards(company_profile)
    required_hours = _quantized_decimal(company_profile.referral_required_hours or 0)
    reward_amount = _money(company_profile.referral_reward_amount or 0)

    submissions = list(
        ReferralSubmission.objects.select_related("referrer", "joined_staff").order_by("-created_at")
    )
    joined_staff_ids = [submission.joined_staff_id for submission in submissions if submission.joined_staff_id]
    active_hours_map = _staff_active_hours_map(joined_staff_ids)
    rewards = {
        reward.referred_staff_id: reward
        for reward in ReferralReward.objects.filter(referred_staff_id__in=joined_staff_ids).select_related(
            "referred_staff",
            "referrer",
        )
    }

    normalized_query = " ".join(str(search_query or "").strip().lower().split())
    stage_filter = str(stage_filter or "all").strip().lower() or "all"
    if stage_filter not in {"all", "not_joined", "joined", "started_working", "completed"}:
        stage_filter = "all"
    reward_filter = str(reward_filter or "all").strip().lower() or "all"
    if reward_filter not in {"all", "pending", "paid", "none"}:
        reward_filter = "all"

    summary = {
        "total": len(submissions),
        "not_joined": 0,
        "joined": 0,
        "started_working": 0,
        "completed": 0,
        "pending_rewards": 0,
        "paid_rewards": 0,
        "pending_total_label": _format_currency(0),
        "program_enabled": company_profile.referral_program_enabled,
        "required_hours_label": f"{float(company_profile.referral_required_hours or 0):,.1f}h",
        "reward_amount_label": _format_currency(company_profile.referral_reward_amount or 0),
    }
    pending_total = Decimal("0.00")
    rows = []

    for submission in submissions:
        required_hours = _submission_required_hours(submission, company_profile=company_profile)
        reward_amount = _submission_reward_amount(submission, company_profile=company_profile)
        referrer = submission.referrer
        joined_staff = submission.joined_staff
        active_hours = active_hours_map.get(joined_staff.id, Decimal("0.00")) if joined_staff else Decimal("0.00")
        reward = rewards.get(joined_staff.id) if joined_staff else None
        is_completed = bool(reward) or (
            joined_staff is not None
            and required_hours <= Decimal("0.00")
        ) or (
            joined_staff is not None
            and required_hours > Decimal("0.00")
            and active_hours >= required_hours
        )

        if joined_staff is None:
            stage = "not_joined"
            stage_label = "Not Joined"
            progress_label = "Waiting to join the team."
            reward_status_label = "Reward unlocks after completion"
            reward_amount_label = "--"
            reward_state = "none"
            summary["not_joined"] += 1
        elif is_completed:
            stage = "completed"
            stage_label = "Completed"
            progress_label = (
                f"{float(active_hours):,.1f}h completed"
                if required_hours <= Decimal("0.00")
                else f"{float(active_hours):,.1f}h of {float(required_hours):,.1f}h completed"
            )
            reward_status_label = "Reward Paid" if reward and reward.is_paid else "Pending Reward"
            reward_amount_label = _format_currency(reward.reward_amount if reward else reward_amount)
            reward_state = "paid" if reward and reward.is_paid else "pending" if reward_amount_label != "--" else "none"
            summary["completed"] += 1
            if reward and reward.is_paid:
                summary["paid_rewards"] += 1
            else:
                summary["pending_rewards"] += 1
                pending_total += reward.reward_amount if reward else reward_amount
        elif active_hours > Decimal("0.00"):
            stage = "started_working"
            stage_label = "Started Working"
            progress_label = (
                f"{float(active_hours):,.1f}h started"
                if required_hours <= Decimal("0.00")
                else f"{float(active_hours):,.1f}h of {float(required_hours):,.1f}h completed"
            )
            reward_status_label = "Reward unlocks after completion"
            reward_amount_label = "--"
            reward_state = "none"
            summary["started_working"] += 1
        else:
            stage = "joined"
            stage_label = "Joined"
            progress_label = "Joined the team and waiting for work hours to begin."
            reward_status_label = "Reward unlocks after completion"
            reward_amount_label = "--"
            reward_state = "none"
            summary["joined"] += 1

        search_text = " ".join(
            part.lower()
            for part in (
                submission.referred_name,
                submission.referred_phone,
                referrer.name if referrer else "",
                referrer.phone if referrer else "",
                joined_staff.name if joined_staff else "",
                joined_staff.phone if joined_staff else "",
                stage_label,
                reward_status_label,
            )
            if part
        )
        if normalized_query and normalized_query not in search_text:
            continue
        if stage_filter != "all" and stage != stage_filter:
            continue
        if reward_filter != "all" and reward_state != reward_filter:
            continue

        rows.append(
            {
                "id": str(submission.id),
                "referrer_name": referrer.name if referrer else "",
                "referrer_phone": referrer.phone if referrer else "",
                "referred_name": submission.referred_name,
                "referred_phone": submission.referred_phone,
                "workflow_stage": stage,
                "workflow_stage_label": stage_label,
                "progress_label": progress_label,
                "active_hours_label": f"{float(active_hours):,.1f}h",
                "required_hours_label": f"{float(required_hours):,.1f}h",
                "reward_amount_label": reward_amount_label,
                "reward_status_label": reward_status_label,
                "reward_state": reward_state,
                "joined_staff_name": joined_staff.name if joined_staff else "",
                "created_at": _format_datetime(submission.created_at),
                "program_label": "Active at submit" if submission.program_enabled_at_submit else "Submitted while paused",
            }
        )

    summary["pending_total_label"] = _format_currency(pending_total)
    return {
        "today_label": timezone.localdate().strftime("%A, %d %b %Y"),
        "search_query": search_query,
        "stage_filter": stage_filter,
        "reward_filter": reward_filter,
        "summary": summary,
        "rows": rows,
    }


def build_staff_profile_payload(request, staff):
    company_profile = get_company_profile()
    sync_referral_rewards(company_profile)
    referral_tracking = _build_referral_tracking_payload(
        staff,
        company_profile=company_profile,
    )
    today, start, end = _today_range()
    previous_month = _shift_month(today.replace(day=1), -1)
    report_month_default = previous_month.strftime("%Y-%m")
    current_month_start, current_month_end = _month_range_for_reference(
        timezone.localdate(),
        end_at=timezone.localtime(timezone.now()),
    )
    active_cutoff = timezone.now() - timedelta(seconds=ONLINE_WINDOW_SECONDS)
    open_session = get_open_session(staff, reconcile=True)
    live_call_staff_ids = _live_call_staff_ids()
    latest_session = Session.objects.filter(staff=staff).order_by("-login_time").first()
    sessions_today = Session.objects.filter(staff=staff, login_time__range=(start, end))
    recent_sessions = Session.objects.filter(staff=staff).order_by("-login_time")[:12]
    calls_today = Call.objects.filter(staff=staff, start_time__range=(start, end))
    qualifying_calls_today = calls_today.filter(is_qualifying=True)
    recent_calls = Call.objects.filter(staff=staff).select_related("lead").order_by("-start_time")
    assigned_leads = _ordered_lead_queryset(
        Lead.objects.filter(assigned_to=staff, status__in=ACTIVE_QUEUE_STATUSES).select_related("assigned_to"),
        now=timezone.now(),
    )
    quality = _build_staff_quality_metrics(
        [staff.id],
        range_start=current_month_start,
        range_end=current_month_end,
    ).get(staff.id, {})
    monthly_salary_history_rows = build_staff_monthly_salary_history_rows(staff, limit=12)
    salary_history_rows = build_staff_salary_history_rows(staff, limit=12)
    review_lead_rows = _staff_review_call_rows(staff)

    active_seconds_today = _effective_active_seconds_for_staff(
        staff=staff,
        start_at=start,
        end_at=end,
    )
    converted_today = qualifying_calls_today.filter(status=Call.Status.CONVERTED).count()
    no_answer_today = qualifying_calls_today.filter(status=Call.Status.NO_ANSWER).count()

    assigned_lead_rows = [
        {
            "id": str(lead.id),
            "name": lead.name,
            "phone": lead.phone,
            "status": lead.status,
            "status_label": lead.get_status_display(),
            "callback_window_label": lead.get_callback_window_display() if lead.callback_window else "",
            "callback_date_label": _format_callback_date_label(lead.callback_date),
            "callback_schedule_label": _format_callback_schedule_label(
                lead.callback_date,
                lead.callback_window,
            ),
            "last_contacted": _format_datetime(lead.last_contacted_at, fallback="Not called yet"),
            "updated_at": _format_datetime(lead.updated_at),
        }
        for lead in assigned_leads
    ]
    recent_call_rows = []
    recent_call_groups = []
    for call in recent_calls:
        local_start = timezone.localtime(call.start_time) if call.start_time else None
        date_key = local_start.date().isoformat() if local_start else "unknown"
        date_label = local_start.strftime("%d %b %Y") if local_start else "Unknown date"
        day_label = local_start.strftime("%A") if local_start else ""
        start_clock_label = local_start.strftime("%I:%M %p").lstrip("0") if local_start else "--"
        row = {
            "id": str(call.id),
            "lead_name": call.lead.name,
            "lead_phone": call.lead.phone,
            "start_time": _format_datetime(call.start_time),
            "start_clock_label": start_clock_label,
            "end_time": _format_datetime(call.end_time),
            "duration_label": _format_duration(call.duration_seconds),
            "status": call.status,
            "status_label": call.get_status_display(),
            "callback_window_label": call.get_callback_window_display() if call.callback_window else "",
            "callback_date_label": _format_callback_date_label(call.callback_date),
            "callback_schedule_label": _format_callback_schedule_label(
                call.callback_date,
                call.callback_window,
            ),
            "activity_date_key": date_key,
            "activity_date_label": date_label,
            "activity_day_label": day_label,
            "is_qualifying": call.is_qualifying,
            "search_text": " ".join(
                part.lower()
                for part in (
                    call.lead.name,
                    call.lead.phone,
                    _format_datetime(call.start_time),
                    start_clock_label,
                    _format_duration(call.duration_seconds),
                    call.get_status_display(),
                    _format_callback_schedule_label(
                        call.callback_date,
                        call.callback_window,
                    ),
                    date_label,
                    day_label,
                )
                if part
            ),
        }
        recent_call_rows.append(row)
        if not recent_call_groups or recent_call_groups[-1]["date_key"] != date_key:
            recent_call_groups.append(
                {
                    "date_key": date_key,
                    "date_label": date_label,
                    "day_label": day_label,
                    "call_count": 0,
                    "rows": [],
                }
            )
        recent_call_groups[-1]["rows"].append(row)
        recent_call_groups[-1]["call_count"] += 1
    recent_session_rows = [
        {
            "id": str(session.id),
            "login_time": _format_datetime(session.login_time),
            "logout_time": _format_datetime(session.logout_time),
            "active_label": _format_duration(session.active_seconds),
            "state_label": _session_status_label(session if session.is_open else None, latest_session=session),
            "close_reason": session.close_reason or "Manual",
        }
        for session in recent_sessions
    ]

    return {
        "today_label": today.strftime("%A, %d %b %Y"),
        "staff_member": staff,
        "summary": {
            "online_label": _staff_online_label(
                open_session,
                active_cutoff,
                is_in_customer_call=staff.id in live_call_staff_ids,
            ),
            "status_label": _session_status_label(open_session, latest_session=latest_session),
            "last_seen": _format_datetime(staff.last_seen_at),
            "assigned_leads": assigned_leads.count(),
            "calls_today": qualifying_calls_today.count(),
            "converted_today": converted_today,
            "no_answer_today": no_answer_today,
            "active_hours_today": _format_hours(active_seconds_today),
            "hourly_rate": _format_currency(staff.hourly_rate),
            "call_rate": _format_currency(staff.call_rate),
            "bonus_per_conversion": _format_currency(staff.bonus_per_conversion),
            "compensation_type_label": staff.get_compensation_type_display(),
            "queue_target": get_lead_queue_limit(),
            "quality_score": quality.get("score", 0),
            "quality_label": quality.get("label", "No Recent Activity"),
            "quality_tone": quality.get("tone", "muted"),
            "quality_note": quality.get("note", "Build more recent call activity for a fuller review."),
            "quality_lookback_days": quality.get("lookback_days", QUALITY_SCORE_LOOKBACK_DAYS),
            "quality_period_label": quality.get("period_label", ""),
            "outcome_consistency_label": quality.get("outcome_consistency_label", "--"),
            "callback_discipline_label": quality.get("callback_discipline_label", "--"),
            "callback_total": quality.get("callback_total", 0),
            "missed_callbacks": quality.get("missed_callbacks", 0),
            "attempt_review_label": quality.get("attempt_review_label", "--"),
            "away_review_label": quality.get("away_review_label", "Within limit"),
            "suspicious_block_count": quality.get("suspicious_block_count", 0),
            "zero_only_block_count": quality.get("zero_only_block_count", 0),
            "zero_second_attempt_count": quality.get("zero_second_attempt_count", 0),
            "real_call_count": quality.get("real_call_count", 0),
            "verified_attempt_count": quality.get("verified_attempt_count", 0),
            "long_away_count": quality.get("long_away_count", 0),
            "expired_followup_count": quality.get("expired_followup_count", 0),
            "expired_followup_penalty_points": quality.get("expired_followup_penalty_points", 0),
            "total_penalty_points": quality.get("total_penalty_points", 0),
        },
        "identity_details": {
            "email": staff.email or "--",
            "bank_account_name": staff.bank_account_name or "--",
            "bank_name": staff.bank_name or "--",
            "bank_account_number": staff.bank_account_number or "--",
            "bank_ifsc_code": staff.bank_ifsc_code or "--",
            "aadhar_number": staff.aadhar_number or "--",
            "aadhar_photo_url": build_staff_document_url(staff, "aadhar"),
            "passbook_photo_url": build_staff_document_url(staff, "passbook"),
        },
        "referral_summary": {
            **referral_tracking["summary"],
        },
        "referral_submission_rows": referral_tracking["rows"][:20],
        "review_lead_summary": {
            "count": len(review_lead_rows),
            "invalid_short_count": sum(1 for row in review_lead_rows if row["is_invalid_short"]),
        },
        "review_lead_rows": review_lead_rows,
        "quality_history_rows": build_staff_quality_history(staff, months=6),
        "monthly_salary_history_rows": monthly_salary_history_rows,
        "salary_history_rows": salary_history_rows,
        "salary_history_summary": {
            "month_count": len(monthly_salary_history_rows),
            "payment_count": len(salary_history_rows),
            "salary_detail_url": reverse("salary-detail-page", args=[staff.id]),
        },
        "report_month_default": report_month_default,
        "report_month_options": _month_option_rows(
            reference_date=today,
            selected_value=report_month_default,
            months_back=6,
        ),
        "assigned_lead_rows": assigned_lead_rows,
        "recent_call_rows": recent_call_rows,
        "recent_call_groups": recent_call_groups,
        "recent_session_rows": recent_session_rows,
    }


def build_salary_control_payload(request):
    company_profile = get_company_profile()
    sync_referral_rewards(company_profile)
    hourly_tracking_count = 0
    weekly_count = 0
    monthly_count = 0
    salary_rows = []
    staff_queryset = list(_staff_queryset())
    conversion_bonus_values = {
        _money(staff.bonus_per_conversion)
        for staff in staff_queryset
    }
    if conversion_bonus_values:
        default_conversion_bonus = min(conversion_bonus_values)
    else:
        default_conversion_bonus = Decimal("10.00")
    referrer_options = [
        {"id": str(staff.id), "name": staff.name}
        for staff in staff_queryset
    ]
    for staff in staff_queryset:
        hourly_tracking_count += 1
        if staff.compensation_type == Staff.CompensationType.WEEKLY:
            weekly_count += 1
        elif staff.compensation_type == Staff.CompensationType.MONTHLY:
            monthly_count += 1

        pending_referral_total = (
            ReferralReward.objects.filter(referrer=staff, is_paid=False)
            .aggregate(total=Sum("reward_amount"))
            .get("total")
            or Decimal("0.00")
        )
        pending_referral_count = ReferralReward.objects.filter(referrer=staff, is_paid=False).count()

        salary_rows.append(
            {
                "id": str(staff.id),
                "name": staff.name,
                "phone": staff.phone,
                "email": staff.email or "--",
                "is_active": staff.is_active,
                "compensation_type": staff.compensation_type,
                "compensation_type_label": _payout_cycle_label(staff),
                "hourly_rate": _format_currency(staff.hourly_rate),
                "hourly_rate_raw": str(staff.hourly_rate),
                "weekly_salary": _format_currency(staff.weekly_salary),
                "weekly_salary_raw": str(staff.weekly_salary),
                "monthly_salary": _format_currency(staff.monthly_salary),
                "monthly_salary_raw": str(staff.monthly_salary),
                "target_hours_per_week": float(staff.target_hours_per_week),
                "target_hours_per_month": float(staff.target_hours_per_month),
                "weekly_payout_day": staff.weekly_payout_day,
                "weekly_payout_day_label": staff.get_weekly_payout_day_display(),
                "monthly_payout_label": "Last day of month",
                "call_rate": _format_currency(staff.call_rate),
                "call_rate_raw": str(staff.call_rate),
                "bonus_per_conversion": _format_currency(staff.bonus_per_conversion),
                "bonus_per_conversion_raw": str(staff.bonus_per_conversion),
                "target_label": _salary_setting_target_label(staff),
                "bank_account_name": staff.bank_account_name or "--",
                "bank_name": staff.bank_name or "--",
                "bank_account_number": staff.bank_account_number or "--",
                "bank_ifsc_code": staff.bank_ifsc_code or "--",
                "has_passbook_photo": bool(staff.passbook_photo),
                "passbook_photo_url": build_staff_document_url(staff, "passbook"),
                "profile_url": reverse("staff-profile-page", args=[staff.id]),
                "referred_by_id": str(staff.referred_by_id) if staff.referred_by_id else "",
                "referred_by_name": staff.referred_by.name if staff.referred_by else "--",
                "pending_referral_reward_count": pending_referral_count,
                "pending_referral_reward_total": _format_currency(pending_referral_total),
            }
        )

    return {
        "summary": {
            "hourly_tracking_count": hourly_tracking_count,
            "weekly_count": weekly_count,
            "monthly_count": monthly_count,
        },
        "referral_settings": {
            "enabled": company_profile.referral_program_enabled,
            "required_hours": f"{Decimal(company_profile.referral_required_hours or 0):.2f}",
            "reward_amount": f"{Decimal(company_profile.referral_reward_amount or 0):.2f}",
            "required_hours_label": f"{float(company_profile.referral_required_hours or 0):,.1f}h",
            "reward_amount_label": _format_currency(company_profile.referral_reward_amount or 0),
        },
        "call_bonus_settings": {
            "enabled": company_profile.hourly_call_bonus_enabled,
            "threshold": int(company_profile.hourly_call_bonus_threshold or 0),
            "rate": f"{Decimal(company_profile.hourly_call_bonus_rate or 0):.2f}",
            "threshold_label": f"{int(company_profile.hourly_call_bonus_threshold or 0)} calls / hour",
            "rate_label": _format_currency(company_profile.hourly_call_bonus_rate or 0),
        },
        "conversion_bonus_settings": {
            "amount": f"{default_conversion_bonus:.2f}",
            "amount_label": _format_currency(default_conversion_bonus),
            "mixed_values": len(conversion_bonus_values) > 1,
            "staff_count": len(staff_queryset),
        },
        "referrer_options": referrer_options,
        "salary_rows": salary_rows,
    }


def build_salary_page_payload():
    today = timezone.localdate()
    company_profile = get_company_profile()
    sync_referral_rewards(company_profile)
    paid_totals_by_staff = {
        row["staff_id"]: row["paid_total"] or Decimal("0.00")
        for row in Salary.objects.filter(paid_amount__gt=Decimal("0.00")).values("staff_id").annotate(paid_total=Sum("paid_amount"))
    }
    paid_meta_by_staff = {
        row["staff_id"]: row["last_paid_at"]
        for row in Salary.objects.filter(paid_amount__gt=Decimal("0.00")).values("staff_id").annotate(last_paid_at=Max("paid_at"))
    }

    pending_total = Decimal("0.00")
    credited_total = Decimal("0.00")
    pending_staff_count = 0
    advance_total = Decimal("0.00")
    advance_staff_count = 0
    pending_referral_total = Decimal("0.00")
    pending_referral_count = 0
    salary_rows = []

    for staff in _staff_queryset():
        (due_start, due_end), due_cycle = _due_period_for_staff(staff, today)
        due_snapshot = _salary_period_snapshot(
            staff,
            period_start=due_start,
            period_end=due_end,
            payout_cycle=due_cycle,
        )
        running_start, running_end = _running_period_for_staff(staff, today)
        running_snapshot = _salary_period_snapshot(
            staff,
            period_start=running_start,
            period_end=running_end,
            payout_cycle=_running_payout_cycle_for_staff(staff),
        )
        running_snapshot = _snapshot_with_paid_total(
            running_snapshot,
            _running_cycle_advance_paid_total(
                staff,
                period_start=running_start,
                period_end=running_end,
                payout_cycle=_running_payout_cycle_for_staff(staff),
            ),
        )
        advance_available = running_snapshot["balance"]
        same_period_as_due = (
            due_snapshot["period_start"] == running_snapshot["period_start"]
            and due_snapshot["period_end"] == running_snapshot["period_end"]
        )
        if same_period_as_due:
            advance_available = Decimal("0.00")

        pending_total += due_snapshot["balance"]
        credited_total += paid_totals_by_staff.get(staff.id, Decimal("0.00"))
        if due_snapshot["balance"] > Decimal("0.00"):
            pending_staff_count += 1
        if advance_available > Decimal("0.00"):
            advance_total += advance_available
            advance_staff_count += 1

        pending_referral_rewards = list(
            ReferralReward.objects.filter(referrer=staff, is_paid=False)
            .select_related("referred_staff")
            .order_by("qualified_at", "created_at")
        )
        pending_referral_amount = sum(
            (reward.reward_amount for reward in pending_referral_rewards),
            Decimal("0.00"),
        )
        pending_referral_total += pending_referral_amount
        pending_referral_count += len(pending_referral_rewards)

        salary_rows.append(
            {
                "id": str(staff.id),
                "name": staff.name,
                "phone": staff.phone,
                "email": staff.email or "",
                "compensation_type": staff.compensation_type,
                "compensation_type_label": _payout_cycle_label(staff),
                "schedule_label": _salary_schedule_label(staff),
                "weekly_payout_day_label": staff.get_weekly_payout_day_display(),
                "hourly_rate": _format_currency(staff.hourly_rate),
                "bank_name": staff.bank_name or "Bank details not added",
                "bank_account_number": staff.bank_account_number or "Account number not added",
                "due_payout_cycle": due_cycle,
                "due_period_start": due_snapshot["period_start"].isoformat(),
                "due_period_end": due_snapshot["period_end"].isoformat(),
                "due_period_label": due_snapshot["period_label"],
                "due_hours_label": due_snapshot["hours_label"],
                "due_earned_total": due_snapshot["earned_total_label"],
                "due_paid_total": due_snapshot["paid_total_label"],
                "due_balance_raw": f"{due_snapshot['balance']:.2f}",
                "due_balance": due_snapshot["balance_label"],
                "due_base_pay": due_snapshot["base_pay_label"],
                "due_call_earnings": due_snapshot["call_earnings_label"],
                "due_bonus_earnings": due_snapshot["bonus_earnings_label"],
                "running_payout_cycle": running_snapshot["payout_cycle"],
                "running_period_start": running_snapshot["period_start"].isoformat(),
                "running_period_end": running_snapshot["period_end"].isoformat(),
                "running_period_label": running_snapshot["period_label"],
                "running_hours_label": running_snapshot["hours_label"],
                "running_earned_total": running_snapshot["earned_total_label"],
                "running_paid_total": running_snapshot["paid_total_label"],
                "advance_available_raw": f"{_money(advance_available):.2f}",
                "advance_available": _format_currency(advance_available),
                "credited_total": _format_currency(paid_totals_by_staff.get(staff.id, Decimal("0.00"))),
                "last_paid_at": _format_datetime(paid_meta_by_staff.get(staff.id)),
                "can_pay_salary": due_snapshot["balance"] > Decimal("0.00"),
                "can_pay_advance": advance_available > Decimal("0.00"),
                "can_pay_referral_rewards": bool(pending_referral_rewards),
                "pending_referral_reward_total": _format_currency(pending_referral_amount),
                "pending_referral_reward_total_raw": f"{pending_referral_amount:.2f}",
                "pending_referral_reward_count": len(pending_referral_rewards),
                "pending_referral_rewards": [
                    {
                        "id": str(reward.id),
                        "referred_staff_name": reward.referred_staff.name,
                        "referred_staff_phone": reward.referred_staff.phone,
                        "reward_amount_label": _format_currency(reward.reward_amount),
                        "qualified_at_label": _format_datetime(reward.qualified_at),
                        "required_hours_label": f"{float(reward.required_hours or 0):,.1f}h",
                    }
                    for reward in pending_referral_rewards
                ],
            }
        )

    salary_rows.sort(
        key=lambda row: (
            -Decimal(row["due_balance_raw"]),
            -Decimal(row["advance_available_raw"]),
            row["name"].lower(),
        )
    )
    pending_salary_rows = [row for row in salary_rows if row["can_pay_salary"] or row["can_pay_advance"]]
    recent_payment_rows = build_recent_salary_payment_rows()
    recent_referral_reward_rows = [
        {
            "id": str(reward.id),
            "referrer_name": reward.referrer.name,
            "referred_staff_name": reward.referred_staff.name,
            "reward_amount_label": _format_currency(reward.reward_amount),
            "qualified_at_label": _format_datetime(reward.qualified_at),
            "paid_at_label": _format_datetime(reward.paid_at),
            "payment_method_label": reward.get_payment_method_display() if reward.payment_method else "Manual",
            "payment_reference": reward.payment_reference or "--",
            "payment_note": reward.payment_note or "--",
        }
        for reward in ReferralReward.objects.filter(is_paid=True)
        .select_related("referrer", "referred_staff")
        .order_by("-paid_at", "-created_at")[:30]
    ]

    return {
        "today_label": today.strftime("%A, %d %b %Y"),
        "summary": {
            "pending_total": _format_currency(pending_total),
            "credited_total": _format_currency(credited_total),
            "pending_staff_count": pending_staff_count,
            "advance_total": _format_currency(advance_total),
            "advance_staff_count": advance_staff_count,
            "pending_referral_total": _format_currency(pending_referral_total),
            "pending_referral_count": pending_referral_count,
            "referral_enabled": company_profile.referral_program_enabled,
            "paid_transaction_count": len(recent_payment_rows),
        },
        "salary_rows": salary_rows,
        "pending_salary_rows": pending_salary_rows,
        "recent_payment_rows": recent_payment_rows,
        "recent_referral_reward_rows": recent_referral_reward_rows,
        "payment_method_options": [
            {"value": value, "label": label}
            for value, label in Salary.PaymentMethod.choices
        ],
    }


def build_salary_detail_payload(staff):
    today = timezone.localdate()
    company_profile = get_company_profile()
    sync_referral_rewards(company_profile)
    (due_start, due_end), due_cycle = _due_period_for_staff(staff, today)
    due_snapshot = _salary_period_snapshot(
        staff,
        period_start=due_start,
        period_end=due_end,
        payout_cycle=due_cycle,
    )
    running_start, running_end = _running_period_for_staff(staff, today)
    running_snapshot = _salary_period_snapshot(
        staff,
        period_start=running_start,
        period_end=running_end,
        payout_cycle=_running_payout_cycle_for_staff(staff),
    )
    running_snapshot = _snapshot_with_paid_total(
        running_snapshot,
        _running_cycle_advance_paid_total(
            staff,
            period_start=running_start,
            period_end=running_end,
            payout_cycle=_running_payout_cycle_for_staff(staff),
        ),
    )
    previous_month_start, previous_month_end = _previous_month_range(today)
    previous_month_snapshot = _salary_period_snapshot(
        staff,
        period_start=previous_month_start,
        period_end=previous_month_end,
        payout_cycle=Salary.PayoutCycle.MONTHLY,
    )
    salary_history = build_staff_salary_history_rows(staff, limit=40)
    total_paid = (
        Salary.objects.filter(staff=staff, paid_amount__gt=Decimal("0.00"))
        .aggregate(total=Sum("paid_amount"))
        .get("total")
        or Decimal("0.00")
    )
    latest_transaction = (
        SalaryPaymentTransaction.objects.filter(salary_record__staff=staff)
        .select_related("salary_record")
        .order_by("-paid_at", "-created_at")
        .first()
    )
    same_period_as_due = (
        due_snapshot["period_start"] == running_snapshot["period_start"]
        and due_snapshot["period_end"] == running_snapshot["period_end"]
    )
    due_period_closed = due_snapshot["period_end"] < today
    show_previous_due_card = due_period_closed and due_snapshot["balance"] > Decimal("0.00")
    advance_available = Decimal("0.00") if same_period_as_due else running_snapshot["balance"]
    pending_referral_rewards = list(
        ReferralReward.objects.filter(referrer=staff, is_paid=False)
        .select_related("referred_staff")
        .order_by("qualified_at", "created_at")
    )
    pending_referral_amount = sum(
        (reward.reward_amount for reward in pending_referral_rewards),
        Decimal("0.00"),
    )

    return {
        "staff_member": staff,
        "summary": {
            "due_salary": due_snapshot["balance_label"],
            "current_earned": running_snapshot["earned_total_label"],
            "advance_available": _format_currency(advance_available),
            "total_paid": _format_currency(total_paid),
            "last_paid_at": _format_datetime(latest_transaction.paid_at) if latest_transaction else "--",
            "last_paid_amount": _format_currency(latest_transaction.amount) if latest_transaction else _format_currency(0),
            "compensation_type_label": _payout_cycle_label(staff),
            "schedule_label": _salary_schedule_label(staff),
            "weekly_payout_day_label": staff.get_weekly_payout_day_display() if staff.weekly_payout_day else "",
        },
        "identity_details": {
            "email": staff.email or "--",
            "bank_account_name": staff.bank_account_name or "--",
            "bank_name": staff.bank_name or "--",
            "bank_account_number": staff.bank_account_number or "--",
            "bank_ifsc_code": staff.bank_ifsc_code or "--",
        },
        "due_snapshot": {
            **due_snapshot,
            "period_start_value": due_snapshot["period_start"].isoformat(),
            "period_end_value": due_snapshot["period_end"].isoformat(),
            "balance_raw": f"{due_snapshot['balance']:.2f}",
            "is_payable": show_previous_due_card,
            "is_closed_period": due_period_closed,
            "cycle_label": Salary.PayoutCycle(due_snapshot["payout_cycle"]).label,
        },
        "running_snapshot": {
            **running_snapshot,
            "period_start_value": running_snapshot["period_start"].isoformat(),
            "period_end_value": running_snapshot["period_end"].isoformat(),
            "advance_available": _format_currency(advance_available),
            "advance_available_raw": f"{_money(advance_available):.2f}",
            "can_pay_advance": advance_available > Decimal("0.00"),
            "cycle_label": Salary.PayoutCycle(running_snapshot["payout_cycle"]).label,
        },
        "previous_month_snapshot": {
            **previous_month_snapshot,
            "period_start_value": previous_month_snapshot["period_start"].isoformat(),
            "period_end_value": previous_month_snapshot["period_end"].isoformat(),
            "cycle_label": Salary.PayoutCycle(previous_month_snapshot["payout_cycle"]).label,
        },
        "company_flags": {
            "referral_enabled": company_profile.referral_program_enabled,
        },
        "page_flags": {
            "show_previous_due_card": show_previous_due_card,
        },
        "pending_referral_rewards": [
            {
                "id": str(reward.id),
                "referred_staff_name": reward.referred_staff.name,
                "referred_staff_phone": reward.referred_staff.phone,
                "reward_amount_label": _format_currency(reward.reward_amount),
                "qualified_at_label": _format_datetime(reward.qualified_at),
                "required_hours_label": f"{float(reward.required_hours or 0):,.1f}h",
            }
            for reward in pending_referral_rewards
        ],
        "pending_referral_summary": {
            "count": len(pending_referral_rewards),
            "amount_label": _format_currency(pending_referral_amount),
        },
        "custom_defaults": {
            "period_start": due_snapshot["period_start"].isoformat(),
            "period_end": due_snapshot["period_end"].isoformat(),
            "paid_amount": f"{due_snapshot['balance']:.2f}",
        },
        "payment_method_options": [
            {"value": value, "label": label}
            for value, label in Salary.PaymentMethod.choices
        ],
        "salary_history_rows": salary_history,
        "recent_salary_history_rows": salary_history[:12],
    }


def get_latest_app_release():
    return (
        AppRelease.objects.filter(is_active=True, published_at__lte=timezone.now())
        .select_related("created_by")
        .order_by("-version_code", "-published_at")
        .first()
    )


def publish_app_release(*, created_by, validated_data):
    is_active = validated_data.get("is_active", True)
    app_release = AppRelease.objects.create(created_by=created_by, **validated_data)
    if is_active:
        AppRelease.objects.exclude(id=app_release.id).update(is_active=False)
    return app_release


def set_active_app_release(app_release):
    AppRelease.objects.exclude(id=app_release.id).update(is_active=False)
    if not app_release.is_active:
        app_release.is_active = True
        app_release.save(update_fields=["is_active", "updated_at"])
    return app_release


def delete_app_release(app_release):
    if app_release.is_active:
        raise ValueError("The active release cannot be deleted.")
    if app_release.apk_file:
        app_release.apk_file.delete(save=False)
    app_release.delete()


def build_developer_release_payload(request):
    releases = AppRelease.objects.select_related("created_by").order_by("-version_code", "-published_at")
    latest_release = get_latest_app_release()
    now = timezone.now()
    _, start, end = _today_range()
    active_cutoff = now - timedelta(seconds=ONLINE_WINDOW_SECONDS)
    open_sessions = _open_sessions_by_staff()
    live_call_staff_ids = _live_call_staff_ids()
    staff_queryset = Staff.objects.filter(role=Staff.Role.STAFF, is_active=True)
    staff_by_id = {staff.id: staff for staff in staff_queryset}
    online_now = 0
    away_now = 0
    warning_now = 0
    offline_now = 0
    for staff_id, staff in staff_by_id.items():
        session = open_sessions.get(staff_id)
        online_label = _staff_online_label(
            session,
            active_cutoff,
            is_in_customer_call=staff_id in live_call_staff_ids,
        )
        if online_label in {"Online", "On Call"}:
            online_now += 1
        elif online_label == "Away":
            away_now += 1
        elif online_label == "Warning":
            warning_now += 1
        else:
            offline_now += 1

    worked_today_count = (
        Session.objects.filter(
            staff__role=Staff.Role.STAFF,
            staff__is_active=True,
            login_time__range=(start, end),
        )
        .values("staff_id")
        .distinct()
        .count()
    )
    latest_heartbeat_session = (
        Session.objects.select_related("staff")
        .filter(
            staff__role=Staff.Role.STAFF,
            staff__is_active=True,
            last_heartbeat_at__isnull=False,
        )
        .order_by("-last_heartbeat_at")
        .first()
    )
    if online_now > 0:
        app_status_label = "App active now"
        app_status_tone = "success"
        app_status_note = f"{online_now} staff device(s) are sending live app activity right now."
    elif warning_now > 0 or away_now > 0:
        app_status_label = "App reachable"
        app_status_tone = "warning"
        app_status_note = "Staff devices were seen recently, but no one is actively working in the foreground right now."
    elif worked_today_count > 0:
        app_status_label = "No one online right now"
        app_status_tone = "muted"
        app_status_note = "The mobile app is reachable, but there is no current live heartbeat from staff at this moment."
    else:
        app_status_label = "Waiting for app activity"
        app_status_tone = "muted"
        app_status_note = "No staff app heartbeat has been recorded today yet."

    total_uploaded_bytes = sum(int(release.file_size_bytes or 0) for release in releases)
    max_release_bytes = max((int(release.file_size_bytes or 0) for release in releases), default=0)
    release_rows = [
        {
            "id": str(release.id),
            "version_name": release.version_name,
            "version_code": release.version_code,
            "minimum_supported_version_code": release.minimum_supported_version_code,
            "release_notes": release.release_notes or "No release notes added.",
            "is_mandatory": release.is_mandatory,
            "is_active": release.is_active,
            "published_at": _format_datetime(release.published_at),
            "created_by": release.created_by.name if release.created_by else "Release Desk",
            "download_url": reverse("app-release-download", args=[release.id]) if release.apk_file else "",
            "download_full_url": request.build_absolute_uri(
                reverse("app-release-download", args=[release.id])
            )
            if release.apk_file
            else "",
            "file_size_label": f"{round((release.file_size_bytes or 0) / (1024 * 1024), 2)} MB",
        }
        for release in releases
    ]
    upload_graph_rows = [
        {
            "label": f"{release.version_name} ({release.version_code})",
            "file_size_label": f"{round((release.file_size_bytes or 0) / (1024 * 1024), 2)} MB",
            "relative_percent": round(
                ((int(release.file_size_bytes or 0) / max_release_bytes) * 100) if max_release_bytes else 0,
                2,
            ),
            "is_active": release.is_active,
        }
        for release in reversed(list(releases[:8]))
    ]
    return {
        "latest_release": latest_release,
        "latest_release_download_url": reverse("app-release-download", args=[latest_release.id])
        if latest_release and latest_release.apk_file
        else "",
        "latest_release_download_full_url": request.build_absolute_uri(
            reverse("app-release-download", args=[latest_release.id])
        )
        if latest_release and latest_release.apk_file
        else "",
        "release_summary": {
            "release_count": len(release_rows),
            "total_uploaded_label": f"{round(total_uploaded_bytes / (1024 * 1024), 2)} MB",
            "largest_release_label": f"{round(max_release_bytes / (1024 * 1024), 2)} MB"
            if max_release_bytes
            else "0 MB",
        },
        "app_health": {
            "checked_at": _format_datetime(now),
            "server_status_label": "Developer center reachable",
            "server_status_tone": "success",
            "server_note": f"The release center is loading correctly from {request.get_host()}.",
            "app_status_label": app_status_label,
            "app_status_tone": app_status_tone,
            "app_status_note": app_status_note,
            "online_now": online_now,
            "on_call_now": len(live_call_staff_ids),
            "away_now": away_now,
            "warning_now": warning_now,
            "offline_now": offline_now,
            "worked_today_count": worked_today_count,
            "open_sessions_now": len(open_sessions),
            "latest_heartbeat_label": _format_datetime(
                latest_heartbeat_session.last_heartbeat_at,
                fallback="No heartbeat yet",
            )
            if latest_heartbeat_session
            else "No heartbeat yet",
            "latest_heartbeat_staff": latest_heartbeat_session.staff.name if latest_heartbeat_session else "",
        },
        "upload_graph_rows": upload_graph_rows,
        "release_rows": release_rows,
    }


def build_app_update_payload(request, *, current_version_code=0):
    latest_release = get_latest_app_release()
    if not latest_release or not latest_release.apk_file:
        return {"update_available": False}

    current_version_code = int(current_version_code or 0)
    download_url = request.build_absolute_uri(reverse("app-release-download", args=[latest_release.id]))
    effective_mandatory = bool(
        latest_release.is_mandatory
        or (
            latest_release.minimum_supported_version_code
            and current_version_code < latest_release.minimum_supported_version_code
        )
    )
    return {
        "update_available": latest_release.version_code > current_version_code,
        "version_name": latest_release.version_name,
        "version_code": latest_release.version_code,
        "minimum_supported_version_code": latest_release.minimum_supported_version_code,
        "release_notes": latest_release.release_notes,
        "is_mandatory": effective_mandatory,
        "download_url": download_url,
        "file_name": latest_release.apk_file.name.rsplit("/", 1)[-1],
        "published_at": latest_release.published_at.isoformat(),
        "file_size_bytes": latest_release.file_size_bytes,
    }

def build_settings_payload(current_user):
    company_profile = get_company_profile()
    company_details = [
        ("Company Email", company_profile.company_email or "Not added yet"),
        ("Company Phone", company_profile.company_phone or "Not added yet"),
        ("Support Phone", company_profile.support_phone or "Not added yet"),
        ("Website", company_profile.website or "Not added yet"),
        ("Tax ID", company_profile.tax_identifier or "Not added yet"),
        ("Country", company_profile.country or "Not added yet"),
        ("Lead Target / Staff", str(get_lead_queue_limit())),
    ]
    address_parts = [
        company_profile.address_line_1,
        company_profile.address_line_2,
        company_profile.city,
        company_profile.state,
        company_profile.postal_code,
        company_profile.country,
    ]
    formatted_address = ", ".join(part for part in address_parts if part) or "Not added yet"

    return {
        "company_profile": company_profile,
        "profile_summary": {
            "name": current_user.name,
            "phone": current_user.phone,
            "role_label": current_user.get_role_display(),
            "last_seen": _format_datetime(current_user.last_seen_at),
            "status_label": "Active" if current_user.is_active else "Inactive",
        },
        "company_summary": {
            "address": formatted_address,
            "description": company_profile.description or "Add company details, address, and support information here.",
            "details": company_details,
        },
    }


def build_lead_management_payload(
    *,
    query="",
    status="all",
    assignment="all",
    callback_window="all",
    contact_state="all",
    notes_state="all",
    date_field="updated_at",
    date_from="",
    date_to="",
    sort_by="updated_at",
    sort_dir="desc",
    readd_only=False,
    page=1,
    page_size=25,
):
    company_profile = get_company_profile()
    active_queue = _lead_queue_queryset()
    queue_limit = get_lead_queue_limit()
    staff_options = [
        {"id": str(staff.id), "name": staff.name}
        for staff in _staff_queryset().filter(is_active=True, receives_new_leads=True)
    ]

    valid_statuses = {choice for choice, _label in Lead.Status.choices}
    valid_callback_windows = {choice for choice, _label in Lead.CallbackWindow.choices}
    valid_date_fields = {"updated_at", "created_at", "last_contacted_at", "callback_date"}
    valid_sort_fields = {
        "updated_at",
        "created_at",
        "last_contacted_at",
        "callback_date",
        "name",
        "phone",
        "status",
        "assigned_to",
        "readd_count",
    }

    trimmed_query = (query or "").strip()
    normalized_status = (status or "all").strip()
    normalized_assignment = (assignment or "all").strip()
    normalized_callback_window = (callback_window or "all").strip()
    normalized_contact_state = (contact_state or "all").strip()
    normalized_notes_state = (notes_state or "all").strip()
    normalized_date_field = (date_field or "updated_at").strip()
    normalized_sort_by = (sort_by or "updated_at").strip()
    normalized_sort_dir = (sort_dir or "desc").strip().lower()
    normalized_sort_dir = "asc" if normalized_sort_dir == "asc" else "desc"
    normalized_readd_only = bool(readd_only)
    try:
        normalized_page_size = int(page_size or 25)
    except (TypeError, ValueError):
        normalized_page_size = 25
    normalized_page_size = max(10, min(normalized_page_size, 100))
    try:
        normalized_page = int(page or 1)
    except (TypeError, ValueError):
        normalized_page = 1
    normalized_page = max(1, normalized_page)

    lead_queryset = _lead_management_queryset().select_related("assigned_to")

    if trimmed_query:
        normalized_phone = re.sub(r"\D+", "", trimmed_query)
        query_filters = (
            Q(name__icontains=trimmed_query)
            | Q(phone__icontains=trimmed_query)
            | Q(notes__icontains=trimmed_query)
            | Q(assigned_to__name__icontains=trimmed_query)
        )
        if normalized_phone and normalized_phone != trimmed_query:
            query_filters |= Q(phone__icontains=normalized_phone)
        lead_queryset = lead_queryset.filter(query_filters)

    if normalized_status in valid_statuses:
        lead_queryset = lead_queryset.filter(status=normalized_status)

    if normalized_assignment == "assigned":
        lead_queryset = lead_queryset.filter(assigned_to__isnull=False)
    elif normalized_assignment == "unassigned":
        lead_queryset = lead_queryset.filter(assigned_to__isnull=True)

    if normalized_callback_window in valid_callback_windows:
        lead_queryset = lead_queryset.filter(callback_window=normalized_callback_window)

    if normalized_contact_state == "contacted":
        lead_queryset = lead_queryset.filter(last_contacted_at__isnull=False)
    elif normalized_contact_state == "not_contacted":
        lead_queryset = lead_queryset.filter(last_contacted_at__isnull=True)

    if normalized_notes_state == "with_notes":
        lead_queryset = lead_queryset.exclude(notes__isnull=True).exclude(notes="")
    elif normalized_notes_state == "without_notes":
        lead_queryset = lead_queryset.filter(Q(notes__isnull=True) | Q(notes=""))

    if normalized_readd_only:
        lead_queryset = lead_queryset.filter(readd_count__gt=0)

    from_date = _parse_date_value(date_from)
    to_date = _parse_date_value(date_to)
    if normalized_date_field in valid_date_fields and (from_date or to_date):
        if normalized_date_field == "callback_date":
            if from_date:
                lead_queryset = lead_queryset.filter(callback_date__gte=from_date)
            if to_date:
                lead_queryset = lead_queryset.filter(callback_date__lte=to_date)
        elif normalized_date_field == "last_contacted_at":
            if from_date:
                lead_queryset = lead_queryset.filter(last_contacted_at__date__gte=from_date)
            if to_date:
                lead_queryset = lead_queryset.filter(last_contacted_at__date__lte=to_date)
        elif normalized_date_field == "created_at":
            if from_date:
                lead_queryset = lead_queryset.filter(created_at__date__gte=from_date)
            if to_date:
                lead_queryset = lead_queryset.filter(created_at__date__lte=to_date)
        else:
            if from_date:
                lead_queryset = lead_queryset.filter(updated_at__date__gte=from_date)
            if to_date:
                lead_queryset = lead_queryset.filter(updated_at__date__lte=to_date)

    if normalized_sort_by not in valid_sort_fields:
        normalized_sort_by = "updated_at"

    sort_prefix = "" if normalized_sort_dir == "asc" else "-"
    if normalized_sort_by == "assigned_to":
        lead_queryset = lead_queryset.order_by(
            f"{sort_prefix}assigned_to__name",
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )
    elif normalized_sort_by in {"name", "phone", "status", "readd_count"}:
        lead_queryset = lead_queryset.order_by(
            f"{sort_prefix}{normalized_sort_by}",
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )
    elif normalized_sort_by == "last_contacted_at":
        lead_queryset = lead_queryset.order_by(
            f"{sort_prefix}last_contacted_at",
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )
    elif normalized_sort_by == "callback_date":
        lead_queryset = lead_queryset.order_by(
            f"{sort_prefix}callback_date",
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )
    elif normalized_sort_by == "created_at":
        lead_queryset = lead_queryset.order_by(
            f"{sort_prefix}created_at",
            f"{sort_prefix}updated_at",
            "id",
        )
    else:
        lead_queryset = lead_queryset.order_by(
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )

    total_lead_count = _lead_management_queryset().count()
    filtered_lead_count = lead_queryset.count()
    status_counts = {
        status_value: lead_queryset.filter(status=status_value).count()
        for status_value, _label in Lead.Status.choices
    }
    assignment_counts = {
        "assigned": lead_queryset.filter(assigned_to__isnull=False).count(),
        "unassigned": lead_queryset.filter(assigned_to__isnull=True).count(),
    }
    contact_counts = {
        "contacted": lead_queryset.filter(last_contacted_at__isnull=False).count(),
        "not_contacted": lead_queryset.filter(last_contacted_at__isnull=True).count(),
    }
    callback_counts = {
        callback_value: lead_queryset.filter(callback_window=callback_value).count()
        for callback_value, _label in Lead.CallbackWindow.choices
    }
    notes_count = lead_queryset.exclude(Q(notes__isnull=True) | Q(notes="")).count()
    readded_lead_count = lead_queryset.filter(readd_count__gt=0).count()

    paginator = Paginator(lead_queryset, normalized_page_size)
    try:
        page_obj = paginator.page(normalized_page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages or 1)

    lead_rows = []
    for lead in page_obj.object_list:
        moved_back_active = bool(lead.followup_moved_back_at and lead.status == Lead.Status.EXPIRED_FOLLOWUP)
        status_label = lead.get_status_display()
        loan_stage_meta = _lead_loan_stage_meta(lead)
        status_tone = {
            Lead.Status.NEW: "new",
            Lead.Status.INTERESTED: "warning",
            Lead.Status.CALL_BACK: "warning",
            Lead.Status.EXPIRED_FOLLOWUP: "danger",
            Lead.Status.NOT_INTERESTED: "muted",
            Lead.Status.NO_ANSWER: "primary",
            Lead.Status.CONVERTED: "primary",
        }.get(lead.status, "muted")
        if moved_back_active:
            status_label = "Follow Up"
            status_tone = "warning"
        lead_rows.append(
            {
                "id": str(lead.id),
                "name": lead.name,
                "phone": lead.phone,
                "status": lead.status,
                "status_label": status_label,
                "status_tone": status_tone,
                "loan_stage": lead.loan_stage or "",
                "loan_stage_label": loan_stage_meta["label"],
                "loan_stage_tone": loan_stage_meta["tone"],
                "callback_window": lead.callback_window,
                "callback_window_label": lead.get_callback_window_display() if lead.callback_window else "",
                "callback_date": lead.callback_date.isoformat() if lead.callback_date else "",
                "callback_date_label": _format_callback_date_label(lead.callback_date),
                "callback_schedule_label": _format_callback_schedule_label(
                    lead.callback_date,
                    lead.callback_window,
                ),
                "assigned_to_id": str(lead.assigned_to_id) if lead.assigned_to_id else "",
                "assigned_to": lead.assigned_to.name if lead.assigned_to else "Unassigned",
                "notes": lead.notes,
                "last_contacted": _format_datetime(lead.last_contacted_at, fallback="Not called yet"),
                "updated_at": _format_datetime(lead.updated_at),
                "created_at": _format_datetime(lead.created_at),
                "followup_moved_back_at": _format_datetime(lead.followup_moved_back_at, fallback="Not moved back yet"),
                "followup_moved_back_present": bool(lead.followup_moved_back_at),
                "readd_count": lead.readd_count,
                "has_notes": bool(lead.notes.strip()),
                "is_unassigned": lead.assigned_to_id is None,
                "is_contacted": lead.last_contacted_at is not None,
            }
        )

    page_window = 2
    start_page = max(1, page_obj.number - page_window)
    end_page = min(paginator.num_pages or 1, page_obj.number + page_window)
    page_numbers = [
        {
            "number": page_number,
            "is_current": page_number == page_obj.number,
        }
        for page_number in range(start_page, end_page + 1)
    ]
    base_query_params = [
        ("q", trimmed_query),
        ("status", normalized_status if normalized_status in valid_statuses else "all"),
        ("assignment", normalized_assignment),
        (
            "callback_window",
            normalized_callback_window if normalized_callback_window in valid_callback_windows else "all",
        ),
        ("contact_state", normalized_contact_state),
        ("notes_state", normalized_notes_state),
        ("date_field", normalized_date_field if normalized_date_field in valid_date_fields else "updated_at"),
        ("date_from", from_date.isoformat() if from_date else ""),
        ("date_to", to_date.isoformat() if to_date else ""),
        ("sort_by", normalized_sort_by),
        ("sort_dir", normalized_sort_dir),
        ("page_size", str(normalized_page_size)),
    ]
    if normalized_readd_only:
        base_query_params.append(("readd_only", "on"))
    base_query_params = [(key, value) for key, value in base_query_params if value not in ("", None)]
    base_querystring = f"?{urlencode(base_query_params)}" if base_query_params else ""

    return {
        "lead_rows": lead_rows,
        "staff_options": staff_options,
        "lead_filters": {
            "query": trimmed_query,
            "status": normalized_status if normalized_status in valid_statuses else "all",
            "assignment": normalized_assignment,
            "callback_window": normalized_callback_window if normalized_callback_window in valid_callback_windows else "all",
            "contact_state": normalized_contact_state,
            "notes_state": normalized_notes_state,
            "date_field": normalized_date_field if normalized_date_field in valid_date_fields else "updated_at",
            "date_from": from_date.isoformat() if from_date else "",
            "date_to": to_date.isoformat() if to_date else "",
            "sort_by": normalized_sort_by,
            "sort_dir": normalized_sort_dir,
            "readd_only": normalized_readd_only,
            "page_size": normalized_page_size,
        },
        "lead_filter_options": {
            "statuses": [{"value": "all", "label": "All statuses"}]
            + [{"value": value, "label": label} for value, label in Lead.Status.choices],
            "assignments": [
                {"value": "all", "label": "All assignments"},
                {"value": "assigned", "label": "Only assigned"},
                {"value": "unassigned", "label": "Only unassigned"},
            ],
            "contact_states": [
                {"value": "all", "label": "All contacts"},
                {"value": "contacted", "label": "Called at least once"},
                {"value": "not_contacted", "label": "Not called yet"},
            ],
            "notes_states": [
                {"value": "all", "label": "Any notes"},
                {"value": "with_notes", "label": "Has notes"},
                {"value": "without_notes", "label": "No notes"},
            ],
            "callback_windows": [{"value": "all", "label": "All callback slots"}]
            + [{"value": value, "label": label} for value, label in Lead.CallbackWindow.choices],
            "date_fields": [
                {"value": "updated_at", "label": "Updated date"},
                {"value": "created_at", "label": "Created date"},
                {"value": "last_contacted_at", "label": "Last contacted date"},
                {"value": "callback_date", "label": "Callback date"},
            ],
            "sort_fields": [
                {"value": "updated_at", "label": "Updated time"},
                {"value": "created_at", "label": "Created time"},
                {"value": "last_contacted_at", "label": "Last contacted time"},
                {"value": "callback_date", "label": "Callback date"},
                {"value": "name", "label": "Lead name"},
                {"value": "phone", "label": "Phone number"},
                {"value": "status", "label": "Status"},
                {"value": "assigned_to", "label": "Assigned staff"},
                {"value": "readd_count", "label": "Re-add count"},
            ],
            "sort_directions": [
                {"value": "desc", "label": "Newest / highest first"},
                {"value": "asc", "label": "Oldest / lowest first"},
            ],
            "page_sizes": [
                {"value": 25, "label": "25 per page"},
                {"value": 50, "label": "50 per page"},
                {"value": 100, "label": "100 per page"},
            ],
        },
        "lead_summary": {
            "total_count": total_lead_count,
            "filtered_count": filtered_lead_count,
            "new_count": int(status_counts.get(Lead.Status.NEW, 0)),
            "follow_up_count": int(status_counts.get(Lead.Status.INTERESTED, 0)),
            "expired_follow_up_count": int(status_counts.get(Lead.Status.EXPIRED_FOLLOWUP, 0)),
            "callback_count": int(status_counts.get(Lead.Status.CALL_BACK, 0)),
            "converted_count": int(status_counts.get(Lead.Status.CONVERTED, 0)),
            "rejected_count": int(status_counts.get(Lead.Status.NOT_INTERESTED, 0)),
            "no_response_count": int(status_counts.get(Lead.Status.NO_ANSWER, 0)),
            "assigned_count": int(assignment_counts.get("assigned", 0)),
            "unassigned_count": int(assignment_counts.get("unassigned", 0)),
            "contacted_count": int(contact_counts.get("contacted", 0)),
            "not_contacted_count": int(contact_counts.get("not_contacted", 0)),
            "notes_count": int(notes_count),
            "readded_count": int(readded_lead_count),
            "callback_window_breakdown": {
                "noon": int(callback_counts.get(Lead.CallbackWindow.NOON, 0)),
                "evening": int(callback_counts.get(Lead.CallbackWindow.EVENING, 0)),
                "night": int(callback_counts.get(Lead.CallbackWindow.NIGHT, 0)),
            },
        },
        "lead_pagination": {
            "page_number": page_obj.number,
            "page_size": normalized_page_size,
            "num_pages": paginator.num_pages or 1,
            "filtered_count": filtered_lead_count,
            "has_previous": page_obj.has_previous(),
            "has_next": page_obj.has_next(),
            "previous_page_number": page_obj.previous_page_number() if page_obj.has_previous() else 1,
            "next_page_number": page_obj.next_page_number() if page_obj.has_next() else (paginator.num_pages or 1),
            "start_index": page_obj.start_index() if filtered_lead_count else 0,
            "end_index": page_obj.end_index() if filtered_lead_count else 0,
            "page_numbers": page_numbers,
            "base_querystring": base_querystring,
        },
        "queue_summary": {
            "active_queue_total": active_queue.count(),
            "unassigned_total": _staff_call_queue_queryset(
                active_queue.filter(assigned_to=None)
            ).count(),
            "staff_active_count": _staff_queryset().filter(is_active=True).count(),
            "queue_limit": queue_limit,
        },
        "readd_summary": {
            "has_readded": readded_lead_count > 0,
            "readded_count": readded_lead_count,
        },
        "lead_cleanup_settings": {
            "auto_enabled": company_profile.lead_auto_delete_enabled,
            "mode": company_profile.lead_auto_delete_mode,
            "days": company_profile.lead_auto_delete_days,
            "count": company_profile.lead_auto_delete_count,
            "last_run_on": company_profile.lead_auto_delete_last_run_on.isoformat()
            if company_profile.lead_auto_delete_last_run_on
            else "",
        },
    }


def build_followup_payload(*, paginate=False, page=1, page_size=25):
    company_profile = get_company_profile()
    expiry_settings = _followup_expiry_settings(company_profile=company_profile)
    work_review_rules = _work_review_rules()
    expire_stale_followups(
        company_profile=company_profile,
        enabled=expiry_settings["enabled"],
        expiry_days=expiry_settings["expiry_days"],
    )
    today = timezone.localdate()
    reference = timezone.now()
    try:
        normalized_page_size = int(page_size or 25)
    except (TypeError, ValueError):
        normalized_page_size = 25
    normalized_page_size = max(10, min(normalized_page_size, 100))
    try:
        normalized_page = int(page or 1)
    except (TypeError, ValueError):
        normalized_page = 1
    normalized_page = max(1, normalized_page)
    followups = list(
        _follow_up_queryset()
        .select_related("assigned_to", "interested_detail__staff")
        .order_by("-updated_at", "-last_contacted_at")
    )
    rejected_candidates = list(
        Lead.objects.select_related("assigned_to", "interested_detail__staff")
        .filter(status=Lead.Status.NOT_INTERESTED)
        .order_by("-updated_at", "-last_contacted_at")[:200]
    )
    expired_candidates = list(
        Lead.objects.select_related("assigned_to", "interested_detail__staff")
        .filter(status=Lead.Status.EXPIRED_FOLLOWUP)
        .order_by("-updated_at", "-last_contacted_at")[:200]
    )

    owner_lookup_ids = [lead.id for lead in followups]
    owner_lookup_ids.extend(lead.id for lead in rejected_candidates)
    owner_lookup_ids.extend(lead.id for lead in expired_candidates)
    owner_lookup_ids = list(dict.fromkeys(owner_lookup_ids))
    call_rows_by_lead = defaultdict(list)
    if owner_lookup_ids:
        progress_call_rows = (
            Call.objects.filter(lead_id__in=owner_lookup_ids)
            .values(
                "id",
                "lead_id",
                "staff_id",
                "status",
                "duration_seconds",
                "start_time",
                "end_time",
                "created_at",
            )
            .order_by("lead_id", "-start_time", "-created_at")
        )
        for row in progress_call_rows:
            call_rows_by_lead[row["lead_id"]].append(row)

    latest_call_owner_map = {}
    if owner_lookup_ids:
        owner_statuses = (
            Call.Status.INTERESTED,
            Call.Status.CALL_BACK,
            Call.Status.NO_ANSWER,
            Call.Status.NOT_INTERESTED,
            Call.Status.CONVERTED,
        )
        latest_owner_calls = (
            Call.objects.filter(lead_id__in=owner_lookup_ids, status__in=owner_statuses)
            .select_related("staff")
            .order_by("lead_id", "-start_time", "-created_at", "-id")
        )
        for call in latest_owner_calls:
            latest_call_owner_map.setdefault(call.lead_id, call.staff)

    def _followup_owner(lead):
        assigned_staff = getattr(lead, "assigned_to", None)
        if assigned_staff:
            return assigned_staff
        interested_detail = getattr(lead, "interested_detail", None)
        if interested_detail and interested_detail.staff_id:
            return interested_detail.staff
        return latest_call_owner_map.get(lead.id)

    mark_statuses = {Call.Status.INTERESTED, Call.Status.CALL_BACK}
    uncalled_alert_enabled = bool(expiry_settings["uncalled_alert_enabled"])
    uncalled_alert_hours = max(1, int(expiry_settings["uncalled_alert_hours"] or 1))
    warning_days = max(1, int(expiry_settings["warning_days"] or FOLLOWUP_STAFF_WARNING_DAYS))
    call_status_labels = dict(Call.Status.choices)

    def _followup_mark_tracking(lead, call_rows):
        latest_mark_index = None
        latest_mark_stamp = None
        for index, row in enumerate(call_rows):
            if row.get("status") not in mark_statuses:
                continue
            stamp = _call_activity_stamp(row)
            if not stamp:
                continue
            latest_mark_index = index
            latest_mark_stamp = timezone.localtime(stamp)
            break

        anchor_at = latest_mark_stamp
        if anchor_at is None:
            fallback_anchor = _followup_activity_anchor(lead)
            anchor_at = timezone.localtime(fallback_anchor) if fallback_anchor else timezone.localtime(reference)

        has_followup_call_after_mark = False
        if latest_mark_index is not None:
            for row in call_rows[:latest_mark_index]:
                if row.get("status") != Call.Status.STARTED:
                    has_followup_call_after_mark = True
                    break

        due_at = anchor_at + timedelta(hours=uncalled_alert_hours)
        is_uncalled_after_mark = not has_followup_call_after_mark
        is_uncalled_overdue = bool(
            uncalled_alert_enabled and is_uncalled_after_mark and timezone.localtime(reference) >= due_at
        )
        if not is_uncalled_after_mark:
            state_label = "Follow-up contacted"
            state_tone = "success"
        elif is_uncalled_overdue:
            state_label = "No call after follow-up mark"
            state_tone = "danger"
        else:
            state_label = "Waiting first follow-up call"
            state_tone = "warning"

        return {
            "latest_mark_at": anchor_at,
            "due_at": due_at,
            "is_uncalled_after_mark": is_uncalled_after_mark,
            "is_uncalled_overdue": is_uncalled_overdue,
            "state_label": state_label,
            "state_tone": state_tone,
        }

    def _build_closed_followup_row(lead, *, row_type):
        owner_staff = _followup_owner(lead)
        activity_anchor = _followup_activity_anchor(lead) or timezone.localtime(reference)
        days_idle = max(
            0,
            (timezone.localdate(reference) - timezone.localdate(activity_anchor)).days,
        )
        row_type_label = "Expired" if row_type == "expired" else "Rejected"
        row_type_tone = "danger" if row_type == "expired" else "secondary"
        return {
            "id": str(lead.id),
            "name": lead.name,
            "phone": lead.phone,
            "status": lead.status,
            "status_label": lead.get_status_display(),
            "status_tone": "danger" if row_type == "expired" else "muted",
            "row_type": row_type,
            "row_type_label": row_type_label,
            "row_type_tone": row_type_tone,
            "is_expired_followup": row_type == "expired",
            "is_rejected_followup": row_type == "rejected",
            "assigned_to_id": str(owner_staff.id) if owner_staff else "",
            "assigned_to": owner_staff.name if owner_staff else "Unassigned",
            "assigned_to_phone": owner_staff.phone if owner_staff else "",
            "notes": lead.notes,
            "last_contacted": _format_datetime(lead.last_contacted_at, fallback="Not called yet"),
            "last_contacted_sort": lead.last_contacted_at.isoformat() if lead.last_contacted_at else "",
            "followup_moved_back_at": _format_datetime(lead.followup_moved_back_at, fallback="Not moved back yet"),
            "followup_moved_back_sort": lead.followup_moved_back_at.isoformat() if lead.followup_moved_back_at else "",
            "updated_at": _format_datetime(lead.updated_at),
            "updated_at_sort": lead.updated_at.isoformat() if lead.updated_at else "",
            "days_idle": days_idle,
        }

    followup_rows = []
    for lead in followups:
        owner_staff = _followup_owner(lead)
        lead_call_rows = call_rows_by_lead.get(lead.id, [])
        is_scheduled = _is_scheduled_followup(lead)
        is_due_today = bool(lead.callback_date and lead.callback_date == today)
        is_due_now = _is_callback_due(lead.callback_date, lead.callback_window, now=reference)
        is_overdue = bool(lead.callback_date and lead.callback_date < today)
        is_unscheduled = not is_scheduled
        followup_progress = _followup_no_response_progress(lead, preloaded_rows=lead_call_rows)
        mark_tracking = _followup_mark_tracking(lead, lead_call_rows)
        activity_anchor = _followup_activity_anchor(lead) or timezone.localtime(reference)
        days_since_update = max(
            0,
            (timezone.localdate(reference) - timezone.localdate(activity_anchor)).days,
        )
        is_warning_due = days_since_update >= warning_days
        days_to_auto_expiry = max(0, int(expiry_settings["expiry_days"]) - days_since_update)
        completed_owner_calls = [
            row
            for row in lead_call_rows
            if row.get("status") != Call.Status.STARTED
            and (
                not owner_staff
                or str(row.get("staff_id") or "") == str(owner_staff.id)
            )
        ]
        followup_call_count = len(completed_owner_calls)
        followup_work_seconds = sum(max(0, int(row.get("duration_seconds") or 0)) for row in completed_owner_calls)
        recent_call_details = []
        for call_row in completed_owner_calls[:3]:
            call_stamp = _call_activity_stamp(call_row)
            recent_call_details.append(
                {
                    "status_label": call_status_labels.get(call_row.get("status"), str(call_row.get("status") or "--")),
                    "time_label": _format_datetime(call_stamp),
                    "duration_label": _format_duration(call_row.get("duration_seconds") or 0),
                }
            )
        followup_attempt_count = followup_progress["attempt_count"]
        attempts_remaining = followup_progress["remaining"]
        can_close_as_no_response = followup_progress["can_close"]

        if is_due_now:
            schedule_state_label = "Due now"
            schedule_state_tone = "warning"
            next_step_label = "Staff should contact now"
        elif is_overdue:
            schedule_state_label = "Overdue"
            schedule_state_tone = "danger"
            next_step_label = "Needs staff follow-up"
        elif is_due_today:
            schedule_state_label = "Due today"
            schedule_state_tone = "primary"
            next_step_label = "Watch for today's slot"
        elif is_unscheduled:
            schedule_state_label = "No time set"
            schedule_state_tone = "muted"
            next_step_label = "Staff can open from Follow Ups menu anytime"
        else:
            schedule_state_label = "Upcoming"
            schedule_state_tone = "success"
            next_step_label = "Wait for scheduled slot"

        followup_rows.append(
            {
                "id": str(lead.id),
                "name": lead.name,
                "phone": lead.phone,
                "status": lead.status,
                "status_label": lead.get_status_display(),
                "handover_status": lead.handover_status,
                "handover_status_label": lead.get_handover_status_display(),
                "callback_window": lead.callback_window,
                "callback_window_label": lead.get_callback_window_display() if lead.callback_window else "",
                "callback_date": lead.callback_date.isoformat() if lead.callback_date else "",
                "callback_date_label": _format_callback_date_label(lead.callback_date),
                "callback_schedule_label": _format_callback_schedule_label(
                    lead.callback_date,
                    lead.callback_window,
                ),
                "assigned_to_id": str(owner_staff.id) if owner_staff else "",
                "assigned_to": owner_staff.name if owner_staff else "Unassigned",
                "assigned_to_phone": owner_staff.phone if owner_staff else "",
                "notes": lead.notes,
                "last_contacted": _format_datetime(lead.last_contacted_at, fallback="Not called yet"),
                "updated_at": _format_datetime(lead.updated_at),
                "updated_at_sort": lead.updated_at.isoformat() if lead.updated_at else "",
                "is_scheduled": is_scheduled,
                "is_due_today": is_due_today,
                "is_due_now": is_due_now,
                "is_overdue": is_overdue,
                "is_unscheduled": is_unscheduled,
                "schedule_state_label": schedule_state_label,
                "schedule_state_tone": schedule_state_tone,
                "next_step_label": next_step_label,
                "followup_attempt_count": followup_attempt_count,
                "attempts_remaining": attempts_remaining,
                "can_close_as_no_response": can_close_as_no_response,
                "followup_attempt_unique_dates": followup_progress["unique_date_count"],
                "followup_attempt_unique_times": followup_progress["unique_time_count"],
                "is_uncalled_after_mark": mark_tracking["is_uncalled_after_mark"],
                "is_uncalled_overdue": mark_tracking["is_uncalled_overdue"],
                "uncalled_state_label": mark_tracking["state_label"],
                "uncalled_state_tone": mark_tracking["state_tone"],
                "followup_marked_at": _format_datetime(mark_tracking["latest_mark_at"]),
                "followup_mark_due_at": _format_datetime(mark_tracking["due_at"]),
                "days_since_followup_update": days_since_update,
                "is_warning_due": is_warning_due,
                "warning_days": warning_days,
                "days_to_auto_expiry": days_to_auto_expiry,
                "followup_call_count": followup_call_count,
                "followup_work_seconds": followup_work_seconds,
                "followup_work_label": _format_duration(followup_work_seconds),
                "followup_moved_back_at": _format_datetime(lead.followup_moved_back_at, fallback="Not moved back"),
                "followup_moved_back_sort": lead.followup_moved_back_at.isoformat() if lead.followup_moved_back_at else "",
                "followup_moved_back_present": bool(lead.followup_moved_back_at),
                "recent_call_details": recent_call_details,
            }
        )

    def _followup_sort_key(row):
        priority = 4
        if row["is_due_now"]:
            priority = 0
        elif row["is_overdue"]:
            priority = 1
        elif row["is_due_today"]:
            priority = 2
        elif row["is_unscheduled"]:
            priority = 3
        return (
            priority,
            row["callback_date"] or "9999-12-31",
            row["callback_window_label"] or "zzz",
            row["updated_at_sort"],
        )

    followup_rows.sort(key=_followup_sort_key)

    rejected_candidate_ids = [lead.id for lead in rejected_candidates]
    rejected_with_followup_history = set(
        Call.objects.filter(
            lead_id__in=rejected_candidate_ids,
            status=Call.Status.INTERESTED,
        ).values_list("lead_id", flat=True)
    )
    rejected_followup_count = sum(1 for lead in rejected_candidates if lead.id in rejected_with_followup_history)
    expired_followup_rows = []
    for lead in expired_candidates:
        expired_followup_rows.append(_build_closed_followup_row(lead, row_type="expired"))

    rejected_followup_rows = []
    for lead in rejected_candidates:
        if lead.id not in rejected_with_followup_history:
            continue
        rejected_followup_rows.append(_build_closed_followup_row(lead, row_type="rejected"))

    closed_followup_rows = expired_followup_rows + rejected_followup_rows
    closed_followup_rows.sort(
        key=lambda row: row.get("updated_at_sort", ""),
        reverse=True,
    )

    expired_followup_rows.sort(
        key=lambda row: row.get("updated_at_sort", ""),
        reverse=True,
    )

    uncalled_after_mark_count = sum(1 for row in followup_rows if row["is_uncalled_after_mark"])
    uncalled_overdue_count = sum(1 for row in followup_rows if row["is_uncalled_overdue"])
    uncalled_waiting_count = max(uncalled_after_mark_count - uncalled_overdue_count, 0)
    warning_due_count = sum(1 for row in followup_rows if row["is_warning_due"])

    performance_by_staff = {}
    for staff in _staff_queryset().filter(is_active=True):
        performance_by_staff[str(staff.id)] = {
            "staff_id": str(staff.id),
            "staff_name": staff.name,
            "active_followups": 0,
            "due_now_count": 0,
            "overdue_count": 0,
            "uncalled_after_mark_count": 0,
            "uncalled_overdue_count": 0,
            "close_ready_count": 0,
            "expired_count": 0,
        }

    def _resolve_staff_bucket(staff_id, staff_name):
        if not staff_id:
            staff_id = "unassigned"
            staff_name = "Unassigned"
        if staff_id and staff_id in performance_by_staff:
            return performance_by_staff[staff_id]
        if staff_id and staff_id not in performance_by_staff:
            performance_by_staff[staff_id] = {
                "staff_id": staff_id,
                "staff_name": staff_name or "Unknown Staff",
                "active_followups": 0,
                "due_now_count": 0,
                "overdue_count": 0,
                "uncalled_after_mark_count": 0,
                "uncalled_overdue_count": 0,
                "close_ready_count": 0,
                "expired_count": 0,
            }
            return performance_by_staff[staff_id]
        return None

    for row in followup_rows:
        bucket = _resolve_staff_bucket(row.get("assigned_to_id"), row.get("assigned_to"))
        if not bucket:
            continue
        bucket["active_followups"] += 1
        if row.get("is_due_now"):
            bucket["due_now_count"] += 1
        if row.get("is_overdue"):
            bucket["overdue_count"] += 1
        if row.get("is_uncalled_after_mark"):
            bucket["uncalled_after_mark_count"] += 1
        if row.get("is_uncalled_overdue"):
            bucket["uncalled_overdue_count"] += 1
        if row.get("can_close_as_no_response"):
            bucket["close_ready_count"] += 1

    for row in expired_followup_rows:
        bucket = _resolve_staff_bucket(row.get("assigned_to_id"), row.get("assigned_to"))
        if not bucket:
            continue
        bucket["expired_count"] += 1

    followup_performance_rows = []
    for bucket in performance_by_staff.values():
        active_followups = int(bucket["active_followups"])
        overdue_penalty = int(bucket["overdue_count"]) + int(bucket["uncalled_overdue_count"])
        expired_penalty = int(bucket["expired_count"]) * 2
        base_score = 100
        health_score = max(base_score - (overdue_penalty * 8) - expired_penalty - int(bucket["close_ready_count"] * 2), 0)
        if active_followups == 0 and bucket["expired_count"] == 0:
            health_label = "No load"
            health_tone = "muted"
        elif health_score >= 80:
            health_label = "Strong"
            health_tone = "success"
        elif health_score >= 60:
            health_label = "Watch"
            health_tone = "warning"
        else:
            health_label = "Critical"
            health_tone = "danger"
        followup_performance_rows.append(
            {
                **bucket,
                "health_score": health_score,
                "health_label": health_label,
                "health_tone": health_tone,
            }
        )

    followup_performance_rows.sort(
        key=lambda row: (
            row["health_score"],
            -row["active_followups"],
            -row["expired_count"],
            row["staff_name"].lower(),
        )
    )

    followup_rows_display = followup_rows
    followup_pagination = {
        "page_number": 1,
        "page_size": normalized_page_size,
        "num_pages": 1,
        "filtered_count": len(followup_rows),
        "has_previous": False,
        "has_next": False,
        "previous_page_number": 1,
        "next_page_number": 1,
        "start_index": 1 if followup_rows else 0,
        "end_index": len(followup_rows),
        "page_numbers": [{"number": 1, "is_current": True}],
        "base_querystring": f"?page_size={normalized_page_size}",
    }
    if paginate:
        paginator = Paginator(followup_rows, normalized_page_size)
        try:
            page_obj = paginator.page(normalized_page)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages or 1)
        followup_rows_display = list(page_obj.object_list)
        page_window = 2
        start_page = max(1, page_obj.number - page_window)
        end_page = min(paginator.num_pages or 1, page_obj.number + page_window)
        followup_pagination = {
            "page_number": page_obj.number,
            "page_size": normalized_page_size,
            "num_pages": paginator.num_pages or 1,
            "filtered_count": len(followup_rows),
            "has_previous": page_obj.has_previous(),
            "has_next": page_obj.has_next(),
            "previous_page_number": page_obj.previous_page_number() if page_obj.has_previous() else 1,
            "next_page_number": page_obj.next_page_number() if page_obj.has_next() else (paginator.num_pages or 1),
            "start_index": page_obj.start_index() if followup_rows else 0,
            "end_index": page_obj.end_index() if followup_rows else 0,
            "page_numbers": [
                {
                    "number": page_number,
                    "is_current": page_number == page_obj.number,
                }
                for page_number in range(start_page, end_page + 1)
            ],
            "base_querystring": f"?page_size={normalized_page_size}",
        }

    return {
        "followup_rows": followup_rows_display,
        "expired_followup_rows": expired_followup_rows,
        "rejected_followup_rows": rejected_followup_rows,
        "closed_followup_rows": closed_followup_rows,
        "staff_options": [
            {"id": str(staff.id), "name": staff.name}
            for staff in _staff_queryset().filter(is_active=True)
        ],
        "followup_summary": {
            "total_followups": len(followup_rows),
            "rejected_followup_count": rejected_followup_count,
            "expired_followup_count": len(expired_followup_rows),
            "closed_followup_count": len(closed_followup_rows),
            "scheduled_count": sum(1 for row in followup_rows if row["is_scheduled"]),
            "unscheduled_count": sum(1 for row in followup_rows if row["is_unscheduled"]),
            "due_now_count": sum(1 for row in followup_rows if row["is_due_now"]),
            "overdue_count": sum(1 for row in followup_rows if row["is_overdue"]),
            "close_ready_count": sum(1 for row in followup_rows if row["can_close_as_no_response"]),
            "assigned_count": sum(1 for row in followup_rows if row["assigned_to_id"]),
            "unassigned_count": sum(1 for row in followup_rows if not row["assigned_to_id"]),
            "with_notes_count": sum(1 for row in followup_rows if row["notes"]),
            "uncalled_after_mark_count": uncalled_after_mark_count,
            "uncalled_waiting_count": uncalled_waiting_count,
            "uncalled_overdue_count": uncalled_overdue_count,
            "warning_due_count": warning_due_count,
        },
        "followup_expiry_settings": {
            "enabled": expiry_settings["enabled"],
            "expiry_days": expiry_settings["expiry_days"],
            "warning_days": expiry_settings["warning_days"],
            "sla_gate_enabled": expiry_settings["sla_gate_enabled"],
            "sla_gate_mode": expiry_settings["sla_gate_mode"],
            "uncalled_alert_enabled": expiry_settings["uncalled_alert_enabled"],
            "uncalled_alert_hours": expiry_settings["uncalled_alert_hours"],
        },
        "work_review_rules": {
            "followup_expired_penalty_points": int(
                work_review_rules.get("followup_expired_penalty_points", FOLLOWUP_EXPIRED_SCORE_PENALTY_POINTS) or 0
            ),
            "followup_expired_penalty_cap": int(
                work_review_rules.get("followup_expired_penalty_cap", FOLLOWUP_EXPIRED_SCORE_PENALTY_CAP) or 0
            ),
        },
        "followup_performance_rows": followup_performance_rows,
        "followup_pagination": followup_pagination,
        "followup_filter_options": {
            "page_sizes": [
                {"value": 25, "label": "25 per page"},
                {"value": 50, "label": "50 per page"},
                {"value": 100, "label": "100 per page"},
            ],
        },
    }

def reassign_unassigned_followup_owners():
    followups = list(
        Lead.objects.select_related("assigned_to", "interested_detail__staff")
        .filter(
            assigned_to__isnull=True,
            status__in=(Lead.Status.INTERESTED, Lead.Status.CALL_BACK),
        )
        .order_by("-updated_at", "-last_contacted_at")
    )
    if not followups:
        return {
            "scanned_count": 0,
            "reassigned_count": 0,
            "still_unassigned_count": 0,
        }

    lead_ids = [lead.id for lead in followups]
    owner_statuses = (
        Call.Status.INTERESTED,
        Call.Status.CALL_BACK,
        Call.Status.NO_ANSWER,
        Call.Status.NOT_INTERESTED,
        Call.Status.CONVERTED,
    )
    latest_call_owner_map = {}
    latest_owner_calls = (
        Call.objects.filter(lead_id__in=lead_ids, status__in=owner_statuses)
        .select_related("staff")
        .order_by("lead_id", "-start_time", "-created_at", "-id")
    )
    for call in latest_owner_calls:
        latest_call_owner_map.setdefault(call.lead_id, call.staff)

    reassigned_leads = []
    for lead in followups:
        interested_detail = getattr(lead, "interested_detail", None)
        resolved_owner = None
        if (
            interested_detail
            and interested_detail.staff_id
            and interested_detail.staff
            and interested_detail.staff.role == Staff.Role.STAFF
            and interested_detail.staff.is_active
        ):
            resolved_owner = interested_detail.staff
        else:
            latest_owner = latest_call_owner_map.get(lead.id)
            if latest_owner and latest_owner.role == Staff.Role.STAFF and latest_owner.is_active:
                resolved_owner = latest_owner
        if not resolved_owner:
            continue
        lead.assigned_to = resolved_owner
        reassigned_leads.append(lead)

    if reassigned_leads:
        with transaction.atomic():
            Lead.objects.bulk_update(reassigned_leads, ["assigned_to", "updated_at"])

    return {
        "scanned_count": len(followups),
        "reassigned_count": len(reassigned_leads),
        "still_unassigned_count": len(followups) - len(reassigned_leads),
    }


def build_followup_csv_response():
    response = io.StringIO()
    writer = csv.writer(response)
    writer.writerow(
        [
            "Lead ID",
            "Name",
            "Phone",
            "Status",
            "Handover Status",
            "Follow-Up Date",
            "Follow-Up Slot",
            "Assigned Staff",
            "Assigned Staff Phone",
            "Notes",
            "Last Contacted",
            "Moved Back At",
            "Updated At",
        ]
    )

    for row in build_followup_payload(paginate=False)["followup_rows"]:
        writer.writerow(
            [
                row["id"],
                row["name"],
                row["phone"],
                row["status_label"],
                row["handover_status_label"],
                row["callback_date_label"],
                row["callback_window_label"],
                row["assigned_to"],
                row["assigned_to_phone"],
                row["notes"],
                row["last_contacted"],
                row["followup_moved_back_at"],
                row["updated_at"],
            ]
        )

    return response.getvalue()


def build_followup_excel_response():
    if Workbook is None:
        raise ValueError("Excel export is not available right now.")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Follow Ups"
    worksheet.append(
        [
            "Lead ID",
            "Name",
            "Phone",
            "Status",
            "Handover Status",
            "Follow-Up Date",
            "Follow-Up Slot",
            "Assigned Staff",
            "Assigned Staff Phone",
            "Notes",
            "Last Contacted",
            "Moved Back At",
            "Updated At",
        ]
    )

    for row in build_followup_payload(paginate=False)["followup_rows"]:
        worksheet.append(
            [
                row["id"],
                row["name"],
                row["phone"],
                row["status_label"],
                row["handover_status_label"],
                row["callback_date_label"],
                row["callback_window_label"],
                row["assigned_to"],
                row["assigned_to_phone"],
                row["notes"],
                row["last_contacted"],
                row["followup_moved_back_at"],
                row["updated_at"],
            ]
        )

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 36)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_callback_payload():
    today = timezone.localdate()
    reference = timezone.now()
    callbacks = (
        _callback_tracking_queryset()
        .select_related("assigned_to")
        .order_by("callback_date", "callback_window", "-updated_at", "-last_contacted_at")
    )

    callback_rows = []
    for lead in callbacks:
        callback_date_label = _format_callback_date_label(lead.callback_date)
        callback_window_label = lead.get_callback_window_display() if lead.callback_window else ""
        is_due_today = bool(lead.callback_date and lead.callback_date == today)
        is_due_now = _is_callback_due(lead.callback_date, lead.callback_window, now=reference)
        is_overdue = bool(lead.callback_date and lead.callback_date < today)
        is_unscheduled = not lead.callback_date or not lead.callback_window
        if is_due_now:
            schedule_state_label = "Due now"
            schedule_state_tone = "warning"
        elif is_overdue:
            schedule_state_label = "Overdue"
            schedule_state_tone = "danger"
        elif is_due_today:
            schedule_state_label = "Due today"
            schedule_state_tone = "primary"
        elif is_unscheduled:
            schedule_state_label = "Schedule needed"
            schedule_state_tone = "muted"
        else:
            schedule_state_label = "Upcoming"
            schedule_state_tone = "success"

        callback_rows.append(
            {
                "id": str(lead.id),
                "name": lead.name,
                "phone": lead.phone,
                "status": lead.status,
                "status_label": lead.get_status_display(),
                "callback_window": lead.callback_window,
                "callback_window_label": callback_window_label,
                "callback_date": lead.callback_date.isoformat() if lead.callback_date else "",
                "callback_date_label": callback_date_label,
                "callback_schedule_label": _format_callback_schedule_label(
                    lead.callback_date,
                    lead.callback_window,
                ),
                "assigned_to_id": str(lead.assigned_to_id) if lead.assigned_to_id else "",
                "assigned_to": lead.assigned_to.name if lead.assigned_to else "Unassigned",
                "assigned_to_phone": lead.assigned_to.phone if lead.assigned_to else "",
                "notes": lead.notes,
                "last_contacted": _format_datetime(lead.last_contacted_at, fallback="Not called yet"),
                "updated_at": _format_datetime(lead.updated_at),
                "is_due_now": is_due_now,
                "is_due_today": is_due_today,
                "is_overdue": is_overdue,
                "is_unscheduled": is_unscheduled,
                "schedule_state_label": schedule_state_label,
                "schedule_state_tone": schedule_state_tone,
            }
        )

    return {
        "callback_rows": callback_rows,
        "staff_options": [
            {"id": str(staff.id), "name": staff.name}
            for staff in _staff_queryset().filter(is_active=True)
        ],
        "callback_summary": {
            "total_callbacks": len(callback_rows),
            "due_now_count": sum(1 for row in callback_rows if row["is_due_now"]),
            "due_today_count": sum(1 for row in callback_rows if row["is_due_today"]),
            "overdue_count": sum(1 for row in callback_rows if row["is_overdue"]),
            "assigned_count": sum(1 for row in callback_rows if row["assigned_to_id"]),
            "unscheduled_count": sum(1 for row in callback_rows if row["is_unscheduled"]),
        },
    }


def build_callback_csv_response():
    response = io.StringIO()
    writer = csv.writer(response)
    writer.writerow(
        [
            "Lead ID",
            "Name",
            "Phone",
            "Callback Date",
            "Callback Window",
            "Schedule State",
            "Assigned Staff",
            "Assigned Staff Phone",
            "Notes",
            "Last Contacted",
            "Updated At",
        ]
    )

    for row in build_callback_payload()["callback_rows"]:
        writer.writerow(
            [
                row["id"],
                row["name"],
                row["phone"],
                row["callback_date_label"],
                row["callback_window_label"],
                row["schedule_state_label"],
                row["assigned_to"],
                row["assigned_to_phone"],
                row["notes"],
                row["last_contacted"],
                row["updated_at"],
            ]
        )

    return response.getvalue()


def build_callback_excel_response():
    if Workbook is None:
        raise ValueError("Excel export is not available right now.")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Callbacks"
    worksheet.append(
        [
            "Lead ID",
            "Name",
            "Phone",
            "Callback Date",
            "Callback Window",
            "Schedule State",
            "Assigned Staff",
            "Assigned Staff Phone",
            "Notes",
            "Last Contacted",
            "Updated At",
        ]
    )

    for row in build_callback_payload()["callback_rows"]:
        worksheet.append(
            [
                row["id"],
                row["name"],
                row["phone"],
                row["callback_date_label"],
                row["callback_window_label"],
                row["schedule_state_label"],
                row["assigned_to"],
                row["assigned_to_phone"],
                row["notes"],
                row["last_contacted"],
                row["updated_at"],
            ]
        )

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 36)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _recovery_status_scope(scope):
    if scope == "rejected":
        return (Lead.Status.NOT_INTERESTED,)
    if scope == "no_response":
        return (Lead.Status.NO_ANSWER,)
    return RECOVERY_LEAD_STATUSES


def _readd_recovery_status_scope(scope):
    if scope == "rejected":
        return ()
    return (Lead.Status.NO_ANSWER,)


def _recovery_owner_staff_map(recovery_leads):
    owner_map = {}
    if not recovery_leads:
        return owner_map

    lead_ids = [lead.id for lead in recovery_leads]
    for lead in recovery_leads:
        assigned = getattr(lead, "assigned_to", None)
        if assigned and assigned.role == Staff.Role.STAFF and assigned.is_active:
            owner_map[lead.id] = assigned

    owner_statuses = (
        Call.Status.INTERESTED,
        Call.Status.CALL_BACK,
        Call.Status.NO_ANSWER,
        Call.Status.NOT_INTERESTED,
        Call.Status.CONVERTED,
    )
    latest_owner_calls = (
        Call.objects.filter(lead_id__in=lead_ids, status__in=owner_statuses)
        .select_related("staff")
        .order_by("lead_id", "-start_time", "-created_at", "-id")
    )
    for call in latest_owner_calls:
        if call.lead_id in owner_map:
            continue
        staff = call.staff
        if staff and staff.role == Staff.Role.STAFF and staff.is_active:
            owner_map[call.lead_id] = staff

    missing_ids = [lead_id for lead_id in lead_ids if lead_id not in owner_map]
    if missing_ids:
        detail_rows = (
            InterestedLeadDetail.objects.filter(lead_id__in=missing_ids)
            .select_related("staff")
            .order_by("lead_id", "-updated_at", "-created_at")
        )
        for detail in detail_rows:
            if detail.lead_id in owner_map:
                continue
            staff = detail.staff
            if staff and staff.role == Staff.Role.STAFF and staff.is_active:
                owner_map[detail.lead_id] = staff

    return owner_map


def build_recovery_lead_payload(
    *,
    query="",
    status="all",
    assignment="all",
    sort_by="updated_at",
    sort_dir="desc",
    page=1,
    page_size=25,
    readd_min="",
    readd_max="",
):
    trimmed_query = " ".join(str(query or "").strip().lower().split())
    normalized_status = str(status or "all").strip()
    normalized_assignment = str(assignment or "all").strip()
    normalized_sort_by = str(sort_by or "updated_at").strip()
    normalized_sort_dir = "asc" if str(sort_dir or "desc").strip().lower() == "asc" else "desc"
    try:
        normalized_page = max(1, int(page or 1))
    except (TypeError, ValueError):
        normalized_page = 1
    try:
        normalized_page_size = int(page_size or 25)
    except (TypeError, ValueError):
        normalized_page_size = 25
    normalized_page_size = max(10, min(normalized_page_size, 100))
    try:
        normalized_readd_min = int(readd_min) if str(readd_min).strip() != "" else None
    except (TypeError, ValueError):
        normalized_readd_min = None
    try:
        normalized_readd_max = int(readd_max) if str(readd_max).strip() != "" else None
    except (TypeError, ValueError):
        normalized_readd_max = None

    recovery_queryset = _recovery_lead_queryset().select_related("assigned_to")
    if trimmed_query:
        normalized_phone = re.sub(r"\D+", "", trimmed_query)
        query_filters = (
            Q(name__icontains=trimmed_query)
            | Q(phone__icontains=trimmed_query)
            | Q(notes__icontains=trimmed_query)
            | Q(assigned_to__name__icontains=trimmed_query)
        )
        if normalized_phone and normalized_phone != trimmed_query:
            query_filters |= Q(phone__icontains=normalized_phone)
        recovery_queryset = recovery_queryset.filter(query_filters)

    if normalized_status in {Lead.Status.NOT_INTERESTED, Lead.Status.NO_ANSWER}:
        recovery_queryset = recovery_queryset.filter(status=normalized_status)

    if normalized_assignment == "assigned":
        recovery_queryset = recovery_queryset.filter(assigned_to__isnull=False)
    elif normalized_assignment == "unassigned":
        recovery_queryset = recovery_queryset.filter(assigned_to__isnull=True)

    if normalized_readd_min is not None:
        recovery_queryset = recovery_queryset.filter(readd_count__gte=normalized_readd_min)
    if normalized_readd_max is not None:
        recovery_queryset = recovery_queryset.filter(readd_count__lte=normalized_readd_max)

    valid_sort_fields = {"updated_at", "created_at", "last_contacted_at", "readd_count", "name", "phone", "status", "assigned_to"}
    if normalized_sort_by not in valid_sort_fields:
        normalized_sort_by = "updated_at"
    sort_prefix = "" if normalized_sort_dir == "asc" else "-"
    if normalized_sort_by == "assigned_to":
        recovery_queryset = recovery_queryset.order_by(
            f"{sort_prefix}assigned_to__name",
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )
    elif normalized_sort_by in {"name", "phone", "status", "readd_count"}:
        recovery_queryset = recovery_queryset.order_by(
            f"{sort_prefix}{normalized_sort_by}",
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )
    elif normalized_sort_by == "last_contacted_at":
        recovery_queryset = recovery_queryset.order_by(
            f"{sort_prefix}last_contacted_at",
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )
    elif normalized_sort_by == "created_at":
        recovery_queryset = recovery_queryset.order_by(
            f"{sort_prefix}created_at",
            f"{sort_prefix}updated_at",
            "id",
        )
    else:
        recovery_queryset = recovery_queryset.order_by(
            f"{sort_prefix}updated_at",
            f"{sort_prefix}created_at",
            "id",
        )

    paginator = Paginator(recovery_queryset, normalized_page_size)
    try:
        page_obj = paginator.page(normalized_page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages or 1)

    recovery_leads = list(page_obj.object_list)
    owner_map = _recovery_owner_staff_map(recovery_leads)
    staff_options = [
        {"id": str(staff.id), "name": staff.name}
        for staff in _staff_queryset().filter(is_active=True, receives_new_leads=True)
    ]

    recovery_rows = []
    for lead in recovery_leads:
        recovery_owner = owner_map.get(lead.id)
        recovery_rows.append(
            {
                "id": str(lead.id),
                "name": lead.name,
                "phone": lead.phone,
                "status": lead.status,
                "status_label": lead.get_status_display(),
                "callback_window": lead.callback_window,
                "callback_date": lead.callback_date.isoformat() if lead.callback_date else "",
                "assigned_to_id": str(lead.assigned_to_id) if lead.assigned_to_id else "",
                "assigned_to": lead.assigned_to.name if lead.assigned_to else "Unassigned",
                "notes": lead.notes,
                "last_contacted": _format_datetime(lead.last_contacted_at, fallback="Not called yet"),
                "updated_at": _format_datetime(lead.updated_at),
                "created_at": _format_datetime(lead.created_at),
                "readd_count": int(lead.readd_count or 0),
                "recovery_owner_id": str(recovery_owner.id) if recovery_owner else "",
                "recovery_owner_name": recovery_owner.name if recovery_owner else "",
            }
        )

    total_recovery_count = _recovery_lead_queryset().count()
    filtered_count = recovery_queryset.count()
    rejected_count = recovery_queryset.filter(status=Lead.Status.NOT_INTERESTED).count()
    no_response_count = recovery_queryset.filter(status=Lead.Status.NO_ANSWER).count()
    oldest_row = recovery_rows[0] if recovery_rows else None
    total_readds = sum(row.get("readd_count", 0) for row in recovery_rows)
    max_readd = max((row.get("readd_count", 0) for row in recovery_rows), default=0)
    page_window = 2
    start_page = max(1, page_obj.number - page_window)
    end_page = min(paginator.num_pages or 1, page_obj.number + page_window)
    page_numbers = [
        {
            "number": page_number,
            "is_current": page_number == page_obj.number,
        }
        for page_number in range(start_page, end_page + 1)
    ]
    base_query_params = [
        ("q", trimmed_query),
        ("status", normalized_status if normalized_status in {Lead.Status.NOT_INTERESTED, Lead.Status.NO_ANSWER} else "all"),
        ("assignment", normalized_assignment),
        ("sort_by", normalized_sort_by),
        ("sort_dir", normalized_sort_dir),
        ("page_size", str(normalized_page_size)),
        ("readd_min", str(normalized_readd_min) if normalized_readd_min is not None else ""),
        ("readd_max", str(normalized_readd_max) if normalized_readd_max is not None else ""),
    ]
    base_query_params = [(key, value) for key, value in base_query_params if value not in ("", None)]
    base_querystring = f"?{urlencode(base_query_params)}" if base_query_params else ""
    return {
        "recovery_rows": recovery_rows,
        "staff_options": staff_options,
        "recovery_filters": {
            "query": trimmed_query,
            "status": normalized_status,
            "assignment": normalized_assignment,
            "sort_by": normalized_sort_by,
            "sort_dir": normalized_sort_dir,
            "page_size": normalized_page_size,
            "readd_min": readd_min,
            "readd_max": readd_max,
        },
        "recovery_filter_options": {
            "statuses": [
                {"value": "all", "label": "All statuses"},
                {"value": Lead.Status.NOT_INTERESTED, "label": "Rejected only"},
                {"value": Lead.Status.NO_ANSWER, "label": "No response only"},
            ],
            "assignments": [
                {"value": "all", "label": "All ownership"},
                {"value": "assigned", "label": "Only assigned"},
                {"value": "unassigned", "label": "Only unassigned"},
            ],
            "sort_fields": [
                {"value": "updated_at", "label": "Marked time"},
                {"value": "created_at", "label": "Created time"},
                {"value": "last_contacted_at", "label": "Last contact time"},
                {"value": "readd_count", "label": "Re-add count"},
                {"value": "name", "label": "Lead name"},
                {"value": "phone", "label": "Phone number"},
                {"value": "status", "label": "Status"},
                {"value": "assigned_to", "label": "Assigned staff"},
            ],
            "sort_directions": [
                {"value": "desc", "label": "Newest / highest first"},
                {"value": "asc", "label": "Oldest / lowest first"},
            ],
            "page_sizes": [
                {"value": 25, "label": "25 per page"},
                {"value": 50, "label": "50 per page"},
                {"value": 100, "label": "100 per page"},
            ],
        },
        "recovery_pagination": {
            "current_page": page_obj.number,
            "total_pages": paginator.num_pages,
            "has_previous": page_obj.has_previous(),
            "has_next": page_obj.has_next(),
            "previous_page_number": page_obj.previous_page_number() if page_obj.has_previous() else None,
            "next_page_number": page_obj.next_page_number() if page_obj.has_next() else None,
            "page_numbers": [
                {"number": item["number"], "is_current": item["is_current"]}
                for item in page_numbers
            ],
            "base_querystring": base_querystring,
            "start_index": page_obj.start_index(),
            "end_index": page_obj.end_index(),
        },
        "recovery_summary": {
            "total_count": total_recovery_count,
            "filtered_count": filtered_count,
            "rejected_count": rejected_count,
            "no_response_count": no_response_count,
            "total_readd_count": total_readds,
            "max_readd_count": max_readd,
            "queue_limit": get_lead_queue_limit(),
            "oldest_lead_name": oldest_row["name"] if oldest_row else "No leads in this list",
            "oldest_lead_updated_at": oldest_row["updated_at"] if oldest_row else "No waiting recovery leads",
        },
    }


def recover_recovery_lead_to_owner(lead_id, *, target_status=Lead.Status.NEW):
    normalized_status = str(target_status or "").strip()
    if normalized_status not in {Lead.Status.NEW, Lead.Status.INTERESTED}:
        raise ValueError("Recovery target must be New or Interested.")

    try:
        parsed_id = uuid.UUID(str(lead_id))
    except (TypeError, ValueError, AttributeError):
        raise ValueError("Select a valid recovery lead.") from None

    lead = (
        Lead.objects.select_related("assigned_to")
        .filter(id=parsed_id, status__in=RECOVERY_LEAD_STATUSES)
        .first()
    )
    if lead is None:
        raise ValueError("Recovery lead not found or already moved from recovery queue.")

    owner_map = _recovery_owner_staff_map([lead])
    owner = owner_map.get(lead.id)
    if owner is None:
        raise ValueError("No active previous staff owner found for this customer.")

    previous_status = lead.status
    update_fields = ["assigned_to", "status", "loan_stage", "callback_window", "callback_date", "updated_at"]
    lead.assigned_to = owner
    lead.status = normalized_status
    if normalized_status == Lead.Status.INTERESTED and not lead.loan_stage:
        lead.loan_stage = Lead.LoanStage.OFFICE_REVIEW
    elif normalized_status == Lead.Status.NEW:
        lead.loan_stage = ""
    lead.callback_window = ""
    lead.callback_date = None
    if normalized_status == Lead.Status.NEW:
        lead.readd_count = int(lead.readd_count or 0) + 1
        update_fields.append("readd_count")
    lead.save(update_fields=update_fields)

    if normalized_status == Lead.Status.NEW:
        auto_allocate_leads(target_staff=owner, prioritized_lead_ids=[lead.id])

    _log_staff_action(
        owner,
        StaffAction.ActionType.CALL_STATUS_UPDATED,
        lead=lead,
        metadata={
            "source": "recovery_restore",
            "target_status": normalized_status,
            "owner_name": owner.name,
            "previous_status": previous_status,
        },
    )

    return {
        "lead_id": str(lead.id),
        "lead_name": lead.name,
        "owner_name": owner.name,
        "target_status": normalized_status,
        "target_status_label": "Interested" if normalized_status == Lead.Status.INTERESTED else "New",
    }


def recover_interested_lead_to_owner(lead_id):
    try:
        parsed_id = uuid.UUID(str(lead_id))
    except (TypeError, ValueError, AttributeError):
        raise ValueError("Select a valid interested lead.") from None

    lead = (
        Lead.objects.select_related("assigned_to", "interested_detail__staff")
        .filter(id=parsed_id)
        .first()
    )
    if lead is None:
        raise ValueError("Interested lead not found.")

    if lead.status == Lead.Status.EXPIRED_FOLLOWUP:
        return reallocate_expired_followup_to_owner(lead_id)
    return recover_recovery_lead_to_owner(lead_id, target_status=Lead.Status.INTERESTED)


def reallocate_expired_followup_to_owner(lead_id):
    try:
        parsed_id = uuid.UUID(str(lead_id))
    except (TypeError, ValueError, AttributeError):
        raise ValueError("Select a valid expired follow-up lead.") from None

    lead = (
        Lead.objects.select_related("assigned_to", "interested_detail__staff")
        .filter(id=parsed_id, status=Lead.Status.EXPIRED_FOLLOWUP)
        .first()
    )
    if lead is None:
        raise ValueError("Expired follow-up lead not found or already reallocated.")

    owner = lead.assigned_to
    if not (
        owner
        and owner.role == Staff.Role.STAFF
        and owner.is_active
    ):
        owner = _recovery_owner_staff_map([lead]).get(lead.id)

    update_fields = [
        "assigned_to",
        "status",
        "loan_stage",
        "callback_window",
        "callback_date",
        "followup_moved_back_at",
        "updated_at",
    ]
    moved_back_at = timezone.now()
    lead.status = Lead.Status.INTERESTED
    if not lead.loan_stage:
        lead.loan_stage = Lead.LoanStage.OFFICE_REVIEW
    lead.callback_window = ""
    lead.callback_date = None
    lead.followup_moved_back_at = moved_back_at
    if owner is not None:
        lead.assigned_to = owner
    else:
        lead.assigned_to = None
    lead.save(update_fields=update_fields)
    actor_staff = owner or (getattr(getattr(lead, "interested_detail", None), "staff", None))
    if actor_staff:
        _log_staff_action(
            actor_staff,
            StaffAction.ActionType.CALL_STATUS_UPDATED,
            lead=lead,
            metadata={
                "source": "followup_move_back",
                "owner_name": owner.name if owner else "Auto allocated",
                "moved_back_at": moved_back_at.isoformat(),
            },
        )
    if owner is not None:
        auto_allocate_leads(target_staff=owner, prioritized_lead_ids=[lead.id])
    else:
        auto_allocate_leads(prioritized_lead_ids=[lead.id])

    return {
        "lead_id": str(lead.id),
        "lead_name": lead.name,
        "owner_name": owner.name if owner else "Auto allocated",
    }


def reactivate_oldest_recovery_leads(count, *, scope="all", max_readd_count=None):
    count = max(1, int(count))
    recovery_statuses = _readd_recovery_status_scope(scope)
    if not recovery_statuses:
        return {
            "reactivated_count": 0,
            "assigned_count": 0,
            "remaining_unassigned_count": _lead_queue_queryset().filter(assigned_to=None).count(),
            "scope_label": scope,
        }
    recovery_queryset = Lead.objects.filter(status__in=recovery_statuses)
    if max_readd_count is not None:
        recovery_queryset = recovery_queryset.filter(readd_count__lte=max_readd_count)
    selected_leads = list(
        recovery_queryset.order_by("readd_count", "updated_at", "last_contacted_at", "created_at", "id")[
            :count
        ]
    )

    if not selected_leads:
        return {
            "reactivated_count": 0,
            "assigned_count": 0,
            "remaining_unassigned_count": _lead_queue_queryset().filter(assigned_to=None).count(),
            "scope_label": scope,
        }

    now = timezone.now()
    selected_ids = [lead.id for lead in selected_leads]
    Lead.objects.filter(id__in=selected_ids).update(
        status=Lead.Status.NEW,
        callback_window="",
        assigned_to=None,
        readd_count=F("readd_count") + 1,
        updated_at=now,
    )
    allocation = auto_allocate_leads(prioritized_lead_ids=selected_ids)
    return {
        "reactivated_count": len(selected_ids),
        "assigned_count": allocation["assigned_count"],
        "remaining_unassigned_count": allocation["remaining_unassigned_count"],
        "scope_label": scope,
    }


def delete_recovery_leads_by_ids(selected_ids):
    valid_ids = []
    for lead_id in (selected_ids or []):
        if not lead_id:
            continue
        try:
            valid_ids.append(str(uuid.UUID(str(lead_id))))
        except (TypeError, ValueError, AttributeError):
            continue
    if not valid_ids:
        return {"deleted_count": 0}
    queryset = Lead.objects.filter(id__in=valid_ids, status__in=RECOVERY_LEAD_STATUSES)
    lead_count = queryset.count()
    queryset.delete()
    return {"deleted_count": lead_count}


def delete_recovery_leads_filtered(count, *, scope="all", max_readd_count=None):
    count = max(1, int(count))
    recovery_statuses = _recovery_status_scope(scope)
    recovery_queryset = Lead.objects.filter(status__in=recovery_statuses)
    if max_readd_count is not None:
        recovery_queryset = recovery_queryset.filter(readd_count__lte=max_readd_count)
    selected_ids = list(
        recovery_queryset.order_by("readd_count", "updated_at", "last_contacted_at", "created_at", "id").values_list(
            "id", flat=True
        )[:count]
    )
    if not selected_ids:
        return {"deleted_count": 0}
    queryset = Lead.objects.filter(id__in=selected_ids, status__in=RECOVERY_LEAD_STATUSES)
    lead_count = queryset.count()
    queryset.delete()
    return {"deleted_count": lead_count}


def build_call_detail_payload(*, limit=200, date_value=""):
    selected_date = _parse_date_value(date_value)
    if selected_date:
        start, end = _day_range_for_date(selected_date)
        calls = (
            Call.objects.select_related("staff", "lead")
            .filter(start_time__range=(start, end))
            .order_by("-start_time")
        )
    else:
        calls = Call.objects.select_related("staff", "lead").order_by("-start_time")
    calls = calls[:limit]
    call_rows = []
    for call in calls:
        call_rows.append(
            {
                "id": str(call.id),
                "staff_name": call.staff.name,
                "lead_name": call.lead.name,
                "lead_phone": call.lead.phone,
                "start_time": _format_datetime(call.start_time),
                "end_time": _format_datetime(call.end_time),
                "duration_seconds": call.duration_seconds,
                "duration_label": _format_duration(call.duration_seconds),
                "status": call.status,
                "status_label": call.get_status_display(),
                "callback_window": call.callback_window,
                "callback_window_label": call.get_callback_window_display() if call.callback_window else "",
                "callback_date": call.callback_date.isoformat() if call.callback_date else "",
                "callback_date_label": _format_callback_date_label(call.callback_date),
                "callback_schedule_label": _format_callback_schedule_label(
                    call.callback_date,
                    call.callback_window,
                ),
                "is_qualifying": call.is_qualifying,
                "auto_skipped_sync_issue": call.auto_skipped_sync_issue,
                "sync_skip_reason": call.sync_skip_reason,
                "sync_skip_reason_label": call.get_sync_skip_reason_display()
                if call.sync_skip_reason
                else "",
            }
        )

    return {
        "call_rows": call_rows,
        "selected_date": selected_date.isoformat() if selected_date else "",
    }


def build_lead_route_map_payload(lead):
    lead = (
        Lead.objects.select_related("assigned_to", "interested_detail__staff")
        .filter(id=getattr(lead, "id", lead))
        .first()
    )
    if lead is None:
        raise ValueError("Lead not found.")

    calls = list(
        Call.objects.filter(lead=lead).select_related("staff").order_by("start_time", "created_at", "id")
    )
    lead_actions = list(
        StaffAction.objects.filter(lead=lead).select_related("staff", "call").order_by("created_at", "id")
    )
    events = []
    seen_event_ids = set()

    def add_event(event_id, *, stamp, title, description="", tone="muted", icon="circle", meta_lines=None):
        if not stamp or event_id in seen_event_ids:
            return
        seen_event_ids.add(event_id)
        local_stamp = timezone.localtime(stamp)
        events.append(
            {
                "id": event_id,
                "title": title,
                "description": description,
                "tone": tone,
                "icon": icon,
                "time_label": _format_datetime(local_stamp),
                "time_sort": local_stamp.isoformat(),
                "meta_lines": [line for line in (meta_lines or []) if line],
            }
        )

    current_owner = lead.assigned_to
    current_owner_name = current_owner.name if current_owner else "Unassigned"
    current_owner_phone = current_owner.phone if current_owner else ""
    status_meta = _lead_route_status_meta(lead.status)
    loan_stage_meta = _lead_loan_stage_meta(lead)
    lead_created_stamp = lead.created_at or lead.updated_at or timezone.now()
    add_event(
        f"lead-created-{lead.id}",
        stamp=lead_created_stamp,
        title="Lead created",
        description="The lead entered the admin system.",
        tone="primary",
        icon="plus-circle-fill",
        meta_lines=[
            f"Current owner: {current_owner_name}",
            f"Current status: {lead.get_status_display()}",
        ],
    )

    for call in calls:
        call_stamp = call.end_time or call.start_time or call.created_at
        if not call_stamp:
            continue
        call_tone = "muted"
        call_icon = "telephone"
        call_title = "Call logged"
        if call.status == Call.Status.STARTED:
            call_title = "Call started"
            call_tone = "primary"
            call_icon = "telephone-forward-fill"
        elif call.status == Call.Status.INTERESTED:
            call_title = "Marked Follow Up"
            call_tone = "warning"
            call_icon = "arrow-repeat"
        elif call.status == Call.Status.CALL_BACK:
            call_title = "Marked Call Back"
            call_tone = "warning"
            call_icon = "calendar2-check"
        elif call.status == Call.Status.NO_ANSWER:
            call_title = "Marked No Response"
            call_tone = "primary"
            call_icon = "skip-forward-circle-fill"
        elif call.status == Call.Status.NOT_INTERESTED:
            call_title = "Marked Rejected"
            call_tone = "danger"
            call_icon = "x-circle-fill"
        elif call.status == Call.Status.CONVERTED:
            call_title = "Marked Successful"
            call_tone = "success"
            call_icon = "check2-circle-fill"
        elif call.status == Call.Status.INVALID_SHORT:
            call_title = "Invalid short call"
            call_tone = "muted"
            call_icon = "slash-circle"

        callback_parts = []
        callback_schedule_label = _format_callback_schedule_label(call.callback_date, call.callback_window)
        if callback_schedule_label:
            callback_parts.append(f"Callback: {callback_schedule_label}")
        elif call.callback_window:
            callback_parts.append(f"Callback slot: {call.get_callback_window_display()}")
        if call.is_verified:
            callback_parts.append("Verified call")
        add_event(
            f"call-{call.id}",
            stamp=call_stamp,
            title=call_title,
            description=f"{call.staff.name} recorded the call outcome.",
            tone=call_tone,
            icon=call_icon,
            meta_lines=[
                f"Staff: {call.staff.name}",
                f"Duration: {_format_duration(call.duration_seconds)}",
                *callback_parts,
            ],
        )

    for action in lead_actions:
        metadata = action.metadata or {}
        source = str(metadata.get("source") or "").strip()
        if source not in {"admin_lead_edit", "followup_expiry", "followup_move_back", "recovery_restore", "customer_recovery"}:
            continue
        stamp = action.created_at
        if not stamp:
            continue
        if source == "admin_lead_edit":
            change_labels = _lead_route_change_labels(metadata)
            if not change_labels:
                continue
            add_event(
                f"action-{action.id}",
                stamp=stamp,
                title="Lead details updated",
                description="Admin changed the lead details.",
                tone="info",
                icon="pencil-square",
                meta_lines=change_labels,
            )
            continue

        if source == "followup_expiry":
            expiry_days = metadata.get("expiry_days")
            idle_days = metadata.get("idle_days")
            meta_lines = []
            if expiry_days:
                meta_lines.append(f"Expiry rule: {expiry_days} day(s)")
            if idle_days:
                meta_lines.append(f"Idle for: {idle_days} day(s)")
            add_event(
                f"action-{action.id}",
                stamp=stamp,
                title="Auto-expired from follow-up",
                description="The lead crossed the follow-up inactivity rule.",
                tone="danger",
                icon="hourglass-split",
                meta_lines=meta_lines,
            )
            continue

        if source == "followup_move_back":
            owner_name = metadata.get("owner_name") or current_owner_name
            add_event(
                f"action-{action.id}",
                stamp=stamp,
                title="Moved back to active follow-up",
                description="The lead returned to the live follow-up queue.",
                tone="warning",
                icon="arrow-counterclockwise",
                meta_lines=[f"Owner: {owner_name}" if owner_name else "Owner: Unassigned"],
            )
            continue

        if source == "recovery_restore":
            target_status = metadata.get("target_status") or Lead.Status.NEW
            target_label = "New lead" if target_status == Lead.Status.NEW else "Interested"
            add_event(
                f"action-{action.id}",
                stamp=stamp,
                title=f"Recovered to {target_label}",
                description="The lead returned from the recovery list.",
                tone="primary",
                icon="arrow-repeat",
                meta_lines=[f"Owner: {metadata.get('owner_name') or current_owner_name}"],
            )
            continue

        if source == "customer_recovery":
            recovery_status = metadata.get("status") or Lead.Status.INTERESTED
            recovery_status_label = dict(Lead.Status.choices).get(recovery_status, recovery_status)
            recovery_callback_window = metadata.get("callback_window") or ""
            recovery_callback_window_label = dict(Lead.CallbackWindow.choices).get(
                recovery_callback_window,
                recovery_callback_window or "Not set",
            )
            add_event(
                f"action-{action.id}",
                stamp=stamp,
                title="Lead recovered by staff",
                description="The lead was restored from staff history.",
                tone="primary",
                icon="arrow-repeat",
                meta_lines=[
                    f"Status: {recovery_status_label}",
                    f"Callback: {recovery_callback_window_label}",
                ],
            )

    interested_detail = getattr(lead, "interested_detail", None)
    if interested_detail and interested_detail.created_at:
        add_event(
            f"detail-{interested_detail.id}",
            stamp=interested_detail.created_at,
            title="Interested details captured",
            description="The lead was opened up with follow-up notes and preferred timing.",
            tone="warning",
            icon="clipboard2-check",
            meta_lines=[
                f"Product: {interested_detail.product_enquired}",
                f"Preferred time: {interested_detail.preferred_call_time or 'Not set'}",
            ],
        )

    if lead.handover_updated_at:
        add_event(
            f"handover-{lead.id}",
            stamp=lead.handover_updated_at,
            title=f"Handover {lead.get_handover_status_display()}",
            description="The handover state changed for this lead.",
            tone="info",
            icon="box-seam",
            meta_lines=[f"Current handover: {lead.get_handover_status_display()}"],
        )

    events.sort(key=lambda row: row["time_sort"])
    last_touch_stamp = lead.last_contacted_at or (calls[-1].end_time if calls else None) or lead.updated_at or lead.created_at
    first_call_stamp = calls[0].start_time if calls else None
    last_call_stamp = None
    for call in reversed(calls):
        last_call_stamp = call.end_time or call.start_time or call.created_at
        if last_call_stamp:
            break

    return {
        "lead": {
            "id": str(lead.id),
            "name": lead.name,
            "phone": lead.phone,
            "status": lead.status,
            "status_label": lead.get_status_display(),
            "status_tone": status_meta["tone"],
            "loan_stage": lead.loan_stage or "",
            "loan_stage_label": loan_stage_meta["label"],
            "loan_stage_tone": loan_stage_meta["tone"],
            "owner_name": current_owner_name,
            "owner_phone": current_owner_phone,
            "created_at": _format_datetime(lead.created_at),
            "updated_at": _format_datetime(lead.updated_at),
            "last_contacted_at": _format_datetime(lead.last_contacted_at, fallback="Not contacted yet"),
            "followup_moved_back_at": _format_datetime(lead.followup_moved_back_at, fallback="Not moved back yet"),
            "callback_schedule_label": _format_callback_schedule_label(
                lead.callback_date,
                lead.callback_window,
            ),
            "handover_status_label": lead.get_handover_status_display(),
            "notes": lead.notes or "",
        },
        "summary": {
            "event_count": len(events),
            "call_count": len(calls),
            "qualifying_call_count": sum(1 for call in calls if call.is_qualifying),
            "first_call_at": _format_datetime(first_call_stamp, fallback="No call yet"),
            "last_call_at": _format_datetime(last_call_stamp, fallback="No call yet"),
            "last_touch_at": _format_datetime(last_touch_stamp),
            "current_status_label": lead.get_status_display(),
            "current_status_tone": status_meta["tone"],
            "current_status_icon": status_meta["icon"],
            "current_owner_name": current_owner_name,
            "current_owner_phone": current_owner_phone,
            "route_state_label": "Success" if lead.status == Lead.Status.CONVERTED else (
                "Closed" if lead.status in {Lead.Status.NOT_INTERESTED, Lead.Status.NO_ANSWER} else (
                    "Expired" if lead.status == Lead.Status.EXPIRED_FOLLOWUP else (
                        "Office pipeline" if lead.status == Lead.Status.INTERESTED else "Open"
                    )
                )
            ),
            "loan_stage_label": loan_stage_meta["label"],
        },
        "timeline_rows": events,
    }


def build_work_hours_payload(*, date_value=""):
    rules = _work_review_rules()
    selected_date = _parse_date_value(date_value) or timezone.localdate()
    start, end = _day_range_for_date(selected_date)
    today = selected_date
    active_cutoff = timezone.now() - timedelta(seconds=ONLINE_WINDOW_SECONDS)
    open_sessions = _open_sessions_by_staff()
    live_call_staff_ids = _live_call_staff_ids()

    sessions_today = Session.objects.filter(login_time__range=(start, end)).select_related("staff").order_by("-login_time")
    staff_list = list(_staff_queryset())
    totals = _effective_active_seconds_map(
        start_at=start,
        end_at=end,
        staff_ids=[staff.id for staff in staff_list],
    )
    gap_summary_map = _work_gap_summary_map(
        start_at=start,
        end_at=end,
        staff_ids=[staff.id for staff in staff_list],
    )
    session_counts = {
        row["staff_id"]: row["count"]
        for row in sessions_today.values("staff_id").annotate(count=Count("id"))
    }
    first_login_map = {}
    last_logout_map = {}
    for session in sessions_today:
        first_login_map[session.staff_id] = min(
            session.login_time,
            first_login_map.get(session.staff_id, session.login_time),
        )
        if session.logout_time:
            last_logout_map[session.staff_id] = max(
                session.logout_time,
                last_logout_map.get(session.staff_id, session.logout_time),
            )

    summary_rows = []
    for staff in staff_list:
        session = open_sessions.get(staff.id)
        gap_summary = gap_summary_map.get(staff.id, {})
        summary_rows.append(
            {
                "id": str(staff.id),
                "name": staff.name,
                "phone": staff.phone,
                "active_hours_today": _format_hours(totals.get(staff.id, 0)),
                "active_seconds_today": totals.get(staff.id, 0),
                "session_count_today": session_counts.get(staff.id, 0),
                "first_login": _format_datetime(first_login_map.get(staff.id)),
                "last_logout": _format_datetime(last_logout_map.get(staff.id)),
                "state_label": _session_status_label(session),
                "online_label": _staff_online_label(
                    session,
                    active_cutoff,
                    is_in_customer_call=staff.id in live_call_staff_ids,
                ),
                "gap_count": gap_summary.get("gap_count", 0),
                "gap_total_label": gap_summary.get("gap_total_label", "0s"),
                "gap_uncounted_label": gap_summary.get("gap_uncounted_label", "0s"),
                "gap_buffer_label": gap_summary.get("gap_buffer_label", "0s"),
                "gap_rows": gap_summary.get("gap_rows", []),
                "gap_extra_count": gap_summary.get("gap_extra_count", 0),
                "call_time_label": gap_summary.get("call_time_label", "0s"),
                "first_call": gap_summary.get("first_call", "--"),
                "last_call": gap_summary.get("last_call", "--"),
            }
        )

    session_rows = []
    for session in sessions_today[:200]:
        session_rows.append(
            {
                "id": str(session.id),
                "staff_name": session.staff.name,
                "login_time": _format_datetime(session.login_time),
                "logout_time": _format_datetime(session.logout_time),
                "active_seconds": session.active_seconds,
                "active_label": _format_duration(session.active_seconds),
                "last_known_state": session.last_known_state,
                "close_reason": session.close_reason or "Manual",
                "is_open": session.is_open,
            }
        )

    return {
        "today_label": today.strftime("%A, %d %b %Y"),
        "selected_date": selected_date.isoformat(),
        "work_gap_rules": {
            "idle_gap_seconds": int(rules.get("idle_gap_seconds", CALL_ACTIVITY_IDLE_BREAK_SECONDS) or 1),
            "connected_cooldown_seconds": int(
                rules.get("connected_cooldown_seconds", CONNECTED_CALL_COOLDOWN_SECONDS) or 0
            ),
            "attempt_threshold": int(rules.get("attempt_threshold", MIN_REAL_CALLS_PER_ATTEMPT_BLOCK) or 1),
        },
        "summary_rows": summary_rows,
        "session_rows": session_rows,
    }


def _work_gap_summary_map(*, start_at=None, end_at=None, staff_ids=None):
    call_queryset = _payable_work_hour_call_queryset()
    rules = _work_review_rules()
    idle_gap_seconds = int(rules.get("idle_gap_seconds", CALL_ACTIVITY_IDLE_BREAK_SECONDS) or 1)
    connected_cooldown_seconds = int(
        rules.get("connected_cooldown_seconds", CONNECTED_CALL_COOLDOWN_SECONDS) or 0
    )

    if start_at is not None and end_at is not None:
        call_queryset = call_queryset.filter(start_time__range=(start_at, end_at))

    if staff_ids is not None:
        staff_ids = list(staff_ids)
        call_queryset = call_queryset.filter(staff_id__in=staff_ids)

    calls_by_staff = defaultdict(list)
    for call in call_queryset.only(
        "staff_id",
        "start_time",
        "end_time",
        "duration_seconds",
    ).order_by("staff_id", "start_time", "end_time", "id"):
        calls_by_staff[call.staff_id].append(call)

    summary_map = {}
    for staff_id in staff_ids or []:
        summary_map[staff_id] = {
            "gap_count": 0,
            "gap_total_label": "0s",
            "gap_uncounted_label": "0s",
            "gap_buffer_label": "0s",
            "gap_rows": [],
            "gap_extra_count": 0,
            "call_time_label": "0s",
            "first_call": "--",
            "last_call": "--",
        }

    for staff_id, calls in calls_by_staff.items():
        gap_rows = []
        gap_total_seconds = 0
        gap_uncounted_seconds = 0
        gap_buffer_seconds = 0
        total_call_seconds = 0
        first_call = None
        last_call = None
        previous_end = None
        previous_call_info = None

        def _gap_call_label(call_record, duration_value):
            if duration_value and duration_value > 0:
                return "Connected call"
            return "Zero-second attempt"

        for call in calls:
            start_time, end_time, duration_seconds = _call_activity_window(call)
            if not start_time:
                continue
            if duration_seconds is None:
                duration_seconds = 0
            total_call_seconds += max(0, int(duration_seconds))
            first_call = min(first_call, start_time) if first_call else start_time
            last_call = max(last_call, end_time) if last_call else end_time

            if previous_end:
                gap_seconds = max(0, int((start_time - previous_end).total_seconds()))
                if gap_seconds > idle_gap_seconds:
                    buffer_seconds = min(gap_seconds, connected_cooldown_seconds)
                    uncounted_seconds = max(0, gap_seconds - connected_cooldown_seconds)
                    if uncounted_seconds <= 0:
                        continue
                    gap_total_seconds += gap_seconds
                    gap_buffer_seconds += buffer_seconds
                    gap_uncounted_seconds += uncounted_seconds
                    previous_label = previous_call_info.get("label") if previous_call_info else "Previous call"
                    current_label = _gap_call_label(call, duration_seconds)
                    gap_rows.append(
                        {
                            "start_time": _format_datetime(previous_end),
                            "end_time": _format_datetime(start_time),
                            "gap_label": _format_duration(gap_seconds),
                            "uncounted_label": _format_duration(uncounted_seconds),
                            "buffer_label": _format_duration(buffer_seconds),
                            "previous_call_label": previous_label,
                            "previous_call_time": _format_datetime(previous_call_info.get("start_time")) if previous_call_info else "--",
                            "previous_call_duration": _format_duration(previous_call_info.get("duration_seconds")) if previous_call_info else "--",
                            "previous_call_status": previous_call_info.get("status_label") if previous_call_info else "--",
                            "current_call_label": current_label,
                            "current_call_time": _format_datetime(call.start_time),
                            "current_call_duration": _format_duration(call.duration_seconds),
                            "current_call_status": call.get_status_display(),
                        }
                    )

            previous_end = end_time
            previous_call_info = {
                "start_time": call.start_time,
                "duration_seconds": call.duration_seconds,
                "status_label": call.get_status_display(),
                "label": _gap_call_label(call, duration_seconds),
            }

        visible_rows = gap_rows
        summary_map[staff_id] = {
            "gap_count": len(gap_rows),
            "gap_total_label": _format_duration(gap_total_seconds),
            "gap_uncounted_label": _format_duration(gap_uncounted_seconds),
            "gap_buffer_label": _format_duration(gap_buffer_seconds),
            "gap_rows": visible_rows,
            "gap_extra_count": 0,
            "call_time_label": _format_duration(total_call_seconds),
            "first_call": _format_datetime(first_call),
            "last_call": _format_datetime(last_call),
        }

    return summary_map


def start_staff_session(staff, *, source="manual_start"):
    now = timezone.now()
    session = get_open_session(staff, reconcile=True)
    if session:
        _touch_session_interaction(session, now, metadata={"source": source, "reused": True})
        return session, False

    pending_lessons = list(get_pending_mandatory_lessons(staff)[:20])
    if pending_lessons:
        pending_payload = build_staff_learning_payload(staff)
        _log_staff_action(
            staff,
            StaffAction.ActionType.TRAINING_REQUIRED_BLOCKED,
            metadata={
                "pending_mandatory_count": pending_payload["summary"]["pending_mandatory_count"],
                "next_required_title": pending_payload["summary"]["next_required_title"],
            },
        )
        raise TrainingRequiredError(pending_payload)

    auto_allocate_leads(target_staff=staff)
    session = Session.objects.create(
        staff=staff,
        login_time=now,
        last_heartbeat_at=now,
        last_interaction_at=now,
        state_changed_at=now,
        last_known_state=Session.AppState.OFFLINE,
        is_open=True,
    )
    mark_staff_seen(staff, now)
    _log_staff_action(
        staff,
        StaffAction.ActionType.SESSION_STARTED,
        session=session,
        app_state=Session.AppState.OFFLINE,
        metadata={"source": source},
    )
    return session, True


def record_session_heartbeat(staff, state, *, interaction=False, source="timer"):
    session = get_open_session(staff)
    if not session:
        return None

    now = timezone.now()
    session = reconcile_session(session, now=now)
    if not session or not session.is_open:
        return None

    has_live_customer_call = staff.id in _reconcile_open_calls(staff=staff, now=now)
    previous_state = session.last_known_state
    session.active_seconds += _active_elapsed_until(
        session,
        now,
        has_live_customer_call=has_live_customer_call,
    )
    session.last_heartbeat_at = now
    session.heartbeat_count += 1

    if interaction:
        session.last_interaction_at = now

    requested_state = _resolve_requested_state(
        session,
        state,
        now,
        interaction,
        has_live_customer_call=has_live_customer_call,
    )
    update_fields = ["active_seconds", "last_heartbeat_at", "heartbeat_count"]

    if interaction:
        update_fields.append("last_interaction_at")

    if requested_state != previous_state:
        session.last_known_state = requested_state
        session.state_changed_at = now
        update_fields.extend(["last_known_state", "state_changed_at"])

    if requested_state == Session.AppState.WARNING:
        if previous_state != Session.AppState.WARNING or session.warning_started_at is None:
            session.warning_started_at = now
            update_fields.append("warning_started_at")
    elif session.warning_started_at is not None:
        session.warning_started_at = None
        update_fields.append("warning_started_at")

    session.save(update_fields=_dedupe_fields(update_fields))
    mark_staff_seen(staff, now)

    action_type = None
    if requested_state == Session.AppState.BACKGROUND and previous_state != Session.AppState.BACKGROUND:
        action_type = StaffAction.ActionType.APP_BACKGROUNDED
    elif requested_state == Session.AppState.WARNING and previous_state != Session.AppState.WARNING:
        action_type = StaffAction.ActionType.IDLE_WARNING
    elif requested_state == Session.AppState.OFFLINE and previous_state != Session.AppState.OFFLINE:
        action_type = StaffAction.ActionType.MARKED_OFFLINE
    elif requested_state == Session.AppState.FOREGROUND and previous_state == Session.AppState.WARNING:
        action_type = StaffAction.ActionType.IDLE_WARNING_ACKNOWLEDGED if interaction else None
    elif requested_state == Session.AppState.FOREGROUND and previous_state == Session.AppState.OFFLINE and interaction:
        action_type = StaffAction.ActionType.RETURNED_ONLINE
    elif requested_state == Session.AppState.FOREGROUND and previous_state == Session.AppState.BACKGROUND:
        action_type = StaffAction.ActionType.APP_FOREGROUNDED

    if action_type:
        _log_staff_action(
            staff,
            action_type,
            session=session,
            app_state=requested_state,
            metadata={"source": source, "interaction": interaction},
        )

    return session


def end_staff_session(staff, *, close_reason="manual"):
    session = get_open_session(staff)
    if not session:
        return None

    now = timezone.now()
    session = reconcile_session(session, now=now)
    if not session or not session.is_open:
        return None

    return _close_session(
        session,
        now,
        close_reason=close_reason,
        auto_generated=close_reason != "manual",
    )


def get_pending_status_call(staff):
    now = timezone.now()
    stale_cutoff = now - timedelta(hours=PENDING_STATUS_BLOCK_HOURS)
    pending_calls = list(
        Call.objects.filter(
            staff=staff,
            status=Call.Status.STARTED,
            end_time__isnull=False,
        )
        .select_related("lead")
        .order_by("-end_time", "-start_time")
    )

    for call in pending_calls:
        if call.end_time and call.end_time < stale_cutoff:
            continue

        superseded_by_newer_call = Call.objects.filter(
            staff=staff,
            start_time__gt=call.end_time or call.start_time,
        ).exclude(id=call.id).exists()
        if superseded_by_newer_call:
            continue

        return call

    return None


def build_staff_today_payload(staff):
    today, start, end = _today_range()
    now = timezone.now()
    sessions_today = Session.objects.filter(staff=staff, login_time__range=(start, end))
    calls_today = Call.objects.filter(staff=staff, start_time__range=(start, end))
    qualifying_calls = calls_today.filter(is_qualifying=True)
    assigned_leads = _visible_staff_lead_queryset(
        Lead.objects.filter(assigned_to=staff, status__in=ACTIVE_QUEUE_STATUSES),
        now=now,
    )
    positive_leads = Lead.objects.filter(
        assigned_to=staff,
        status=Lead.Status.INTERESTED,
    )
    scheduled_followup_leads = positive_leads.filter(
        callback_date__isnull=False,
    ).exclude(callback_window="")
    converted_leads = Lead.objects.filter(
        assigned_to=staff,
        status=Lead.Status.CONVERTED,
    )
    open_session = get_open_session(staff, reconcile=True)
    latest_session = sessions_today.order_by("-login_time").first()
    learning_payload = build_staff_learning_payload(staff)
    pending_status_call = get_pending_status_call(staff)
    recoverable_open_call = get_recoverable_open_call(staff, now=now)
    followup_sla_gate_status = build_staff_followup_sla_gate_status(staff, now=now)
    notifications = build_staff_active_notifications_payload(staff, now=now)

    active_seconds = _effective_active_seconds_for_staff(
        staff=staff,
        start_at=start,
        end_at=end,
    )
    calls_count = qualifying_calls.count()
    follow_up_count = positive_leads.count()
    scheduled_followup_count = scheduled_followup_leads.count()
    converted_count = converted_leads.count()

    return {
        "today": today.isoformat(),
        "notifications": notifications,
        "summary": {
            "active_seconds": active_seconds,
            "active_label": _format_hours(active_seconds),
            "calls_count": calls_count,
            "interested_count": follow_up_count,
            "converted_count": converted_count,
            "result_label": f"{follow_up_count} follow up / {scheduled_followup_count} scheduled",
            "working_now": bool(open_session),
            "current_state": open_session.last_known_state if open_session else "stopped",
            "status_label": _session_status_label(open_session, latest_session=latest_session),
            "close_reason": latest_session.close_reason if latest_session else "",
            "pending_training_count": learning_payload["summary"]["pending_mandatory_count"],
            "training_required": learning_payload["summary"]["has_pending_mandatory"],
            "next_training_title": learning_payload["summary"]["next_required_title"],
            "pending_call_status_required": bool(pending_status_call),
            "pending_call_id": str(pending_status_call.id) if pending_status_call else "",
            "pending_call_lead_id": str(pending_status_call.lead_id) if pending_status_call else "",
            "pending_call_lead_name": pending_status_call.lead.name if pending_status_call else "",
            "pending_call_lead_phone": pending_status_call.lead.phone if pending_status_call else "",
            "recoverable_call_required": bool(recoverable_open_call),
            "recoverable_call_id": str(recoverable_open_call.id) if recoverable_open_call else "",
            "recoverable_call_lead_id": str(recoverable_open_call.lead_id) if recoverable_open_call else "",
            "recoverable_call_lead_name": recoverable_open_call.lead.name if recoverable_open_call else "",
            "recoverable_call_lead_phone": recoverable_open_call.lead.phone if recoverable_open_call else "",
            "recoverable_call_started_at": recoverable_open_call.start_time.isoformat()
            if recoverable_open_call
            else None,
            "followup_sla_crossed_count": followup_sla_gate_status["crossed_sla_count"],
            "followup_sla_warning_days": followup_sla_gate_status["warning_days"],
            "followup_sla_gate_enabled": followup_sla_gate_status["gate_enabled"],
            "followup_sla_gate_mode": followup_sla_gate_status["gate_mode"],
            "followup_sla_followup_calls_today": followup_sla_gate_status["followup_calls_today"],
            "normal_lead_calls_blocked_by_sla": not followup_sla_gate_status["normal_lead_calls_allowed"],
            "followup_sla_block_message": followup_sla_gate_status["block_message"],
        },
        "session": {
            "id": str(open_session.id),
            "is_open": open_session.is_open,
            "last_known_state": open_session.last_known_state,
            "active_seconds": open_session.active_seconds,
            "last_interaction_at": open_session.last_interaction_at.isoformat() if open_session.last_interaction_at else None,
            "warning_started_at": open_session.warning_started_at.isoformat() if open_session.warning_started_at else None,
            "close_reason": open_session.close_reason,
        }
        if open_session
        else None,
    }


def build_staff_followups_payload(staff):
    company_profile = get_company_profile()
    expiry_settings = _followup_expiry_settings(company_profile=company_profile)
    expire_stale_followups(
        company_profile=company_profile,
        enabled=expiry_settings["enabled"],
        expiry_days=expiry_settings["expiry_days"],
    )
    now = timezone.now()
    warning_days = max(1, int(expiry_settings["warning_days"] or FOLLOWUP_STAFF_WARNING_DAYS))
    followups = list(
        Lead.objects.select_related("assigned_to")
        .filter(
            assigned_to=staff,
            status=Lead.Status.CALL_BACK,
        )
        .order_by("callback_date", "callback_window", "-updated_at", "-last_contacted_at")
    )
    lead_ids = [lead.id for lead in followups]
    call_rows_by_lead = defaultdict(list)
    if lead_ids:
        call_rows = (
            Call.objects.filter(lead_id__in=lead_ids, staff=staff)
            .values(
                "id",
                "lead_id",
                "staff_id",
                "status",
                "duration_seconds",
                "start_time",
                "end_time",
                "created_at",
            )
            .order_by("lead_id", "-start_time", "-created_at")
        )
        for row in call_rows:
            call_rows_by_lead[row["lead_id"]].append(row)

    rows = []
    warning_count = 0
    oldest_warning_days = 0
    for lead in followups:
        lead_call_rows = call_rows_by_lead.get(lead.id, [])
        followup_progress = _followup_no_response_progress(
            lead,
            staff=staff,
            preloaded_rows=lead_call_rows,
        )
        no_answer_attempt_count = followup_progress["attempt_count"]
        is_scheduled_followup = _is_scheduled_followup(lead)
        is_due_now = _is_followup_highlighted(lead, now=now) if is_scheduled_followup else False
        activity_anchor = _followup_activity_anchor(lead) or timezone.localtime(now)
        days_since_update = max(
            0,
            (timezone.localdate(now) - timezone.localdate(activity_anchor)).days,
        )
        is_warning_due = days_since_update >= warning_days
        if is_warning_due:
            warning_count += 1
            oldest_warning_days = max(oldest_warning_days, days_since_update)
        days_to_auto_expiry = max(0, int(expiry_settings["expiry_days"]) - days_since_update)
        completed_staff_calls = [
            row for row in lead_call_rows if row.get("status") != Call.Status.STARTED
        ]
        followup_work_seconds = sum(
            max(0, int(row.get("duration_seconds") or 0)) for row in completed_staff_calls
        )

        rows.append(
            {
                "id": str(lead.id),
                "name": lead.name,
                "phone": lead.phone,
                "status": Lead.Status.INTERESTED,
                "status_label": "Follow Up",
                "callback_window": lead.callback_window,
                "callback_window_label": lead.get_callback_window_display() if lead.callback_window else "",
                "callback_date": lead.callback_date.isoformat() if lead.callback_date else "",
                "callback_date_label": _format_callback_date_label(lead.callback_date),
                "callback_schedule_label": _format_callback_schedule_label(
                    lead.callback_date,
                    lead.callback_window,
                ),
                "notes": lead.notes,
                "last_contacted_at": lead.last_contacted_at.isoformat() if lead.last_contacted_at else "",
                "updated_at": lead.updated_at.isoformat(),
                "is_due_now": is_due_now,
                "followup_attempt_count": no_answer_attempt_count,
                "followup_attempts_remaining": followup_progress["remaining"],
                "can_mark_followup_no_response": followup_progress["can_close"],
                "followup_attempt_unique_dates": followup_progress["unique_date_count"],
                "followup_attempt_unique_times": followup_progress["unique_time_count"],
                "is_scheduled_followup": is_scheduled_followup,
                "followup_warning_due": is_warning_due,
                "followup_warning_days": warning_days,
                "followup_stale_days": days_since_update,
                "followup_warning_label": (
                    f"Pending for {days_since_update} day(s). Call now."
                    if is_warning_due
                    else ""
                ),
                "days_to_auto_expiry": (
                    days_to_auto_expiry if expiry_settings["enabled"] else None
                ),
                "followup_work_seconds": followup_work_seconds,
                "followup_work_label": _format_duration(followup_work_seconds),
                "followup_call_count": len(completed_staff_calls),
            }
        )
    rows.sort(
        key=lambda row: (
            0 if row["followup_warning_due"] else 1,
            0 if row["is_due_now"] else 1,
            0 if row["is_scheduled_followup"] else 1,
            row["callback_date"] or "9999-12-31",
            row["callback_window"] or "zzzz",
            row["updated_at"],
        )
    )
    return {
        "followups": rows,
        "warning_summary": {
            "warning_days": warning_days,
            "warning_count": warning_count,
            "oldest_warning_days": oldest_warning_days,
            "popup_required": warning_count > 0,
            "title": "Follow-up calls pending",
            "message": (
                f"{warning_count} follow-up lead(s) have been pending for {warning_days} day(s) or more. "
                "Open Follow Ups now and complete those calls."
                if warning_count > 0
                else "No pending follow-up warning items right now."
            ),
        },
        "followup_settings": {
            "auto_expire_enabled": bool(expiry_settings["enabled"]),
            "auto_expire_days": int(expiry_settings["expiry_days"]),
            "warning_days": warning_days,
            "sla_gate_enabled": bool(expiry_settings["sla_gate_enabled"]),
            "sla_gate_mode": str(expiry_settings["sla_gate_mode"]),
        },
        "sla_gate_status": build_staff_followup_sla_gate_status(
            staff,
            now=now,
            company_profile=company_profile,
        ),
    }


def save_interested_lead_detail(
    call,
    *,
    customer_name,
    customer_phone,
    product_enquired,
    enquiry_notes="",
    preferred_call_time="",
):
    if call.status != Call.Status.INTERESTED:
        raise ValueError("Interested enquiry details can only be saved after marking the call as Interested.")

    detail, _ = InterestedLeadDetail.objects.update_or_create(
        lead=call.lead,
        defaults={
            "staff": call.staff,
            "call": call,
            "customer_name": customer_name.strip(),
            "customer_phone": customer_phone.strip(),
            "product_enquired": product_enquired.strip(),
            "enquiry_notes": (enquiry_notes or "").strip(),
            "preferred_call_time": preferred_call_time.strip(),
        },
    )
    return detail


def _normalize_page_number(page_value, default=1):
    try:
        page_number = int(page_value or default)
    except (TypeError, ValueError):
        page_number = default
    return max(1, page_number)


def _build_interested_page_url(base_params, page_key, page_number):
    query_params = dict(base_params)
    query_params[page_key] = str(max(1, int(page_number or 1)))
    return f"?{urlencode(query_params)}" if query_params else ""


def _paginate_interested_rows(rows, *, page_value, page_size, base_params, page_key):
    normalized_page_size = max(1, int(page_size or 25))
    normalized_page = _normalize_page_number(page_value)
    if not rows:
        return [], {
            "page_number": 1,
            "page_size": normalized_page_size,
            "num_pages": 0,
            "total_count": 0,
            "has_previous": False,
            "has_next": False,
            "previous_page_number": 1,
            "next_page_number": 1,
            "start_index": 0,
            "end_index": 0,
            "page_links": [],
            "previous_url": "",
            "next_url": "",
        }

    paginator = Paginator(rows, normalized_page_size)
    try:
        page_obj = paginator.page(normalized_page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages or 1)

    page_window = 2
    start_page = max(1, page_obj.number - page_window)
    end_page = min(paginator.num_pages or 1, page_obj.number + page_window)
    page_links = [
        {
            "number": page_number,
            "is_current": page_number == page_obj.number,
            "url": _build_interested_page_url(base_params, page_key, page_number),
        }
        for page_number in range(start_page, end_page + 1)
    ]
    return list(page_obj.object_list), {
        "page_number": page_obj.number,
        "page_size": normalized_page_size,
        "num_pages": paginator.num_pages,
        "total_count": len(rows),
        "has_previous": page_obj.has_previous(),
        "has_next": page_obj.has_next(),
        "previous_page_number": page_obj.previous_page_number() if page_obj.has_previous() else 1,
        "next_page_number": page_obj.next_page_number() if page_obj.has_next() else (paginator.num_pages or 1),
        "start_index": page_obj.start_index() if rows else 0,
        "end_index": page_obj.end_index() if rows else 0,
        "page_links": page_links,
        "previous_url": _build_interested_page_url(base_params, page_key, page_obj.previous_page_number()) if page_obj.has_previous() else "",
        "next_url": _build_interested_page_url(base_params, page_key, page_obj.next_page_number()) if page_obj.has_next() else "",
    }


def build_interested_lead_payload(
    *,
    query="",
    date_from="",
    date_to="",
    staff_id="",
    outcome="all",
    page_size=25,
    pending_page=1,
    success_page=1,
    unsuccessful_page=1,
):
    queryset = InterestedLeadDetail.objects.select_related("lead", "staff", "call")
    trimmed_query = (query or "").strip()
    normalized_staff_id = (staff_id or "").strip()
    normalized_outcome = (outcome or "all").strip().lower() or "all"
    normalized_page_size = max(10, min(_normalize_page_number(page_size, default=25), 100))
    if normalized_outcome in {"converted"}:
        normalized_outcome = "successful"
    elif normalized_outcome in {"rejected", "no_response", "expired_followup"}:
        normalized_outcome = "unsuccessful"
    if normalized_outcome not in {"all", "successful", "unsuccessful", "open"}:
        normalized_outcome = "all"

    from_value = (date_from or "").strip()
    to_value = (date_to or "").strip()
    parsed_from = None
    parsed_to = None
    try:
        parsed_from = date.fromisoformat(from_value) if from_value else None
    except ValueError:
        parsed_from = None
        from_value = ""
    try:
        parsed_to = date.fromisoformat(to_value) if to_value else None
    except ValueError:
        parsed_to = None
        to_value = ""

    if trimmed_query:
        queryset = queryset.filter(
            Q(customer_name__icontains=trimmed_query)
            | Q(customer_phone__icontains=trimmed_query)
            | Q(product_enquired__icontains=trimmed_query)
            | Q(enquiry_notes__icontains=trimmed_query)
            | Q(staff__name__icontains=trimmed_query)
            | Q(lead__name__icontains=trimmed_query)
            | Q(lead__phone__icontains=trimmed_query)
        )
    if normalized_staff_id:
        queryset = queryset.filter(staff_id=normalized_staff_id)
    if parsed_from:
        queryset = queryset.filter(created_at__date__gte=parsed_from)
    if parsed_to:
        queryset = queryset.filter(created_at__date__lte=parsed_to)
    if normalized_outcome == "successful":
        queryset = queryset.filter(lead__status=Lead.Status.CONVERTED)
    elif normalized_outcome == "unsuccessful":
        queryset = queryset.filter(lead__status__in=[Lead.Status.NOT_INTERESTED, Lead.Status.NO_ANSWER, Lead.Status.EXPIRED_FOLLOWUP])
    elif normalized_outcome == "open":
        queryset = queryset.filter(lead__status=Lead.Status.INTERESTED)

    rows = []
    staff_breakdown = defaultdict(
        lambda: {
            "staff_name": "",
            "staff_phone": "",
            "total_count": 0,
            "converted_count": 0,
            "rejected_count": 0,
            "no_response_count": 0,
            "open_count": 0,
        }
    )
    for detail in queryset.order_by("-updated_at", "-created_at"):
        lead_status = detail.lead.status
        loan_stage_meta = _lead_loan_stage_meta(detail.lead)
        outcome_group = {
            Lead.Status.CONVERTED: "successful",
            Lead.Status.INTERESTED: "open",
            Lead.Status.NOT_INTERESTED: "unsuccessful",
            Lead.Status.NO_ANSWER: "unsuccessful",
            Lead.Status.EXPIRED_FOLLOWUP: "unsuccessful",
        }.get(lead_status, "other")
        status_tone = {
            Lead.Status.CONVERTED: "success",
            Lead.Status.INTERESTED: "warning",
            Lead.Status.NOT_INTERESTED: "danger",
            Lead.Status.NO_ANSWER: "danger",
            Lead.Status.EXPIRED_FOLLOWUP: "warning",
            Lead.Status.NEW: "muted",
        }.get(lead_status, "muted")
        outcome_key = {
            Lead.Status.CONVERTED: "converted",
            Lead.Status.NOT_INTERESTED: "rejected",
            Lead.Status.NO_ANSWER: "no_response",
            Lead.Status.INTERESTED: "open",
            Lead.Status.EXPIRED_FOLLOWUP: "unsuccessful",
        }.get(lead_status, "other")
        if lead_status == Lead.Status.CONVERTED:
            lead_status_label = "Successful"
        elif lead_status == Lead.Status.INTERESTED:
            lead_status_label = detail.lead.get_status_display()
        elif lead_status == Lead.Status.EXPIRED_FOLLOWUP:
            lead_status_label = "Expired"
        elif lead_status in {Lead.Status.NOT_INTERESTED, Lead.Status.NO_ANSWER}:
            lead_status_label = "Unsuccessful"
        else:
            lead_status_label = detail.lead.get_status_display()
        if detail.call:
            result_at = detail.call.end_time or detail.call.created_at or detail.updated_at or detail.created_at
        else:
            result_at = detail.updated_at or detail.created_at
        rows.append(
            {
                "id": str(detail.id),
                "lead_id": str(detail.lead_id),
                "customer_name": detail.customer_name,
                "customer_phone": detail.customer_phone,
                "product_enquired": detail.product_enquired,
                "enquiry_notes": detail.enquiry_notes,
                "preferred_call_time": detail.preferred_call_time,
                "staff_name": detail.staff.name,
                "staff_phone": detail.staff.phone,
                "updated_at": _format_datetime(detail.updated_at),
                "created_at": _format_datetime(detail.created_at),
                "created_at_date": detail.created_at.date().isoformat(),
                "result_at": _format_datetime(result_at),
                "lead_status": lead_status,
                "lead_status_label": lead_status_label,
                "lead_status_tone": status_tone,
                "loan_stage": detail.lead.loan_stage or "",
                "loan_stage_label": loan_stage_meta["label"],
                "loan_stage_tone": loan_stage_meta["tone"],
                "outcome_group": outcome_group,
                "outcome_group_label": {
                    "open": "Interested",
                    "successful": "Successful",
                    "unsuccessful": "Unsuccessful",
                }.get(outcome_group, "Other"),
                "outcome_group_tone": {
                    "open": "warning",
                    "successful": "success",
                    "unsuccessful": "danger",
                }.get(outcome_group, "muted"),
                "is_converted": lead_status == Lead.Status.CONVERTED,
                "reward_amount_label": _format_currency(detail.staff.bonus_per_conversion) if lead_status == Lead.Status.CONVERTED else "",
                "outcome_key": outcome_key,
            }
        )
        breakdown = staff_breakdown[detail.staff_id]
        breakdown["staff_name"] = detail.staff.name
        breakdown["staff_phone"] = detail.staff.phone
        breakdown["total_count"] += 1
        if lead_status == Lead.Status.CONVERTED:
            breakdown["converted_count"] += 1
        elif lead_status == Lead.Status.NOT_INTERESTED:
            breakdown["rejected_count"] += 1
        elif lead_status == Lead.Status.NO_ANSWER:
            breakdown["no_response_count"] += 1
        elif lead_status == Lead.Status.INTERESTED:
            breakdown["open_count"] += 1

    interested_open_rows = [row for row in rows if row["outcome_group"] == "open"]
    interested_success_rows = [row for row in rows if row["outcome_group"] == "successful"]
    interested_unsuccessful_rows = [row for row in rows if row["outcome_group"] == "unsuccessful"]

    base_page_params = {
        "q": trimmed_query,
        "date_from": from_value,
        "date_to": to_value,
        "staff_id": normalized_staff_id,
        "outcome": normalized_outcome,
        "page_size": str(normalized_page_size),
        "pending_page": str(_normalize_page_number(pending_page)),
        "success_page": str(_normalize_page_number(success_page)),
        "unsuccessful_page": str(_normalize_page_number(unsuccessful_page)),
    }
    pending_rows_page, pending_pagination = _paginate_interested_rows(
        interested_open_rows,
        page_value=pending_page,
        page_size=normalized_page_size,
        base_params=base_page_params,
        page_key="pending_page",
    )
    success_rows_page, success_pagination = _paginate_interested_rows(
        interested_success_rows,
        page_value=success_page,
        page_size=normalized_page_size,
        base_params=base_page_params,
        page_key="success_page",
    )
    unsuccessful_rows_page, unsuccessful_pagination = _paginate_interested_rows(
        interested_unsuccessful_rows,
        page_value=unsuccessful_page,
        page_size=normalized_page_size,
        base_params=base_page_params,
        page_key="unsuccessful_page",
    )

    staff_breakdown_rows = sorted(
        staff_breakdown.values(),
        key=lambda row: (
            -row["converted_count"],
            -row["open_count"],
            -row["total_count"],
            row["staff_name"].lower(),
        ),
    )
    staff_options = [
        {"id": str(staff.id), "name": staff.name}
        for staff in Staff.objects.filter(role=Staff.Role.STAFF, is_active=True, receives_new_leads=True).order_by("name")
    ]

    return {
        "interested_all_rows": rows,
        "interested_lead_rows": interested_open_rows,
        "interested_success_rows": interested_success_rows,
        "interested_unsuccessful_rows": interested_unsuccessful_rows,
        "interested_pending_rows": pending_rows_page,
        "interested_success_rows_page": success_rows_page,
        "interested_unsuccessful_rows_page": unsuccessful_rows_page,
        "interested_pending_pagination": pending_pagination,
        "interested_success_pagination": success_pagination,
        "interested_unsuccessful_pagination": unsuccessful_pagination,
        "interested_page_size": normalized_page_size,
        "interested_lead_summary": {
            "total_count": len(rows),
            "today_count": sum(
                1
                for row in rows
                if row["created_at"].startswith(timezone.localdate().strftime("%d %b %Y"))
            ),
            "with_notes_count": sum(1 for row in rows if row["enquiry_notes"]),
            "scheduled_count": sum(1 for row in rows if row["preferred_call_time"]),
            "converted_count": len(interested_success_rows),
            "rejected_count": sum(1 for row in rows if row["lead_status"] == Lead.Status.NOT_INTERESTED),
            "no_response_count": sum(1 for row in rows if row["lead_status"] == Lead.Status.NO_ANSWER),
            "open_count": len(interested_open_rows),
            "unsuccessful_count": len(interested_unsuccessful_rows),
            "staff_count": len(staff_breakdown_rows),
        },
        "interested_staff_breakdown_rows": staff_breakdown_rows,
        "interested_staff_options": staff_options,
        "interested_search_query": trimmed_query,
        "interested_filter_date_from": from_value,
        "interested_filter_date_to": to_value,
        "interested_filter_staff_id": normalized_staff_id,
        "interested_filter_outcome": normalized_outcome,
    }


def build_interested_lead_csv_response(*, query="", date_from="", date_to="", staff_id="", outcome="all"):
    payload = build_interested_lead_payload(
        query=query,
        date_from=date_from,
        date_to=date_to,
        staff_id=staff_id,
        outcome=outcome,
    )
    export_rows = list(payload.get("interested_all_rows", []))
    response = io.StringIO()
    writer = csv.writer(response)
    writer.writerow(
        [
            "Lead ID",
            "Customer Name",
            "Customer Phone",
            "Product",
            "Notes",
            "Preferred Call Time",
            "Staff",
            "Staff Phone",
            "Outcome",
            "Loan Stage",
            "Result At",
            "Reward Amount",
            "Created At",
            "Updated At",
        ]
    )
    for row in export_rows:
        writer.writerow(
            [
                row["lead_id"],
                row["customer_name"],
                row["customer_phone"],
                row["product_enquired"],
                row["enquiry_notes"],
                row["preferred_call_time"],
                row["staff_name"],
                row["staff_phone"],
                row["lead_status_label"],
                row["loan_stage_label"],
                row["result_at"],
                row["reward_amount_label"] or "--",
                row["created_at"],
                row["updated_at"],
            ]
        )
    return response.getvalue()


def build_interested_lead_excel_response(*, query="", date_from="", date_to="", staff_id="", outcome="all"):
    if Workbook is None:
        raise ValueError("Excel export is not available right now.")

    payload = build_interested_lead_payload(
        query=query,
        date_from=date_from,
        date_to=date_to,
        staff_id=staff_id,
        outcome=outcome,
    )
    export_rows = list(payload.get("interested_all_rows", []))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Interested Outcomes"
    worksheet.append(
        [
            "Lead ID",
            "Customer Name",
            "Customer Phone",
            "Product",
            "Notes",
            "Preferred Call Time",
            "Staff",
            "Staff Phone",
            "Outcome",
            "Loan Stage",
            "Result At",
            "Reward Amount",
            "Created At",
            "Updated At",
        ]
    )
    for row in export_rows:
        worksheet.append(
            [
                row["lead_id"],
                row["customer_name"],
                row["customer_phone"],
                row["product_enquired"],
                row["enquiry_notes"],
                row["preferred_call_time"],
                row["staff_name"],
                row["staff_phone"],
                row["lead_status_label"],
                row["loan_stage_label"],
                row["result_at"],
                row["reward_amount_label"] or "--",
                row["created_at"],
                row["updated_at"],
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def get_assigned_leads(staff):
    now = timezone.now()
    auto_allocate_leads(target_staff=staff)
    return _ordered_lead_queryset(
        _visible_staff_lead_queryset(
            Lead.objects.filter(
                assigned_to=staff,
                status__in=STAFF_CALL_QUEUE_STATUSES,
            ).select_related("assigned_to"),
            now=now,
        ),
        now=now,
    )


def search_staff_customer_history(staff, *, query="", limit=25):
    called_lead_ids = Call.objects.filter(staff=staff).values_list("lead_id", flat=True)
    queryset = Lead.objects.filter(Q(assigned_to=staff) | Q(id__in=called_lead_ids)).select_related("assigned_to")

    trimmed_query = (query or "").strip()
    if trimmed_query:
        normalized_phone = re.sub(r"\D+", "", trimmed_query)
        filters = Q(name__icontains=trimmed_query) | Q(phone__icontains=trimmed_query)
        if normalized_phone and normalized_phone != trimmed_query:
            filters |= Q(phone__icontains=normalized_phone)
        queryset = queryset.filter(filters).annotate(
            exact_match_priority=Case(
                When(phone__iexact=trimmed_query, then=Value(0)),
                When(name__iexact=trimmed_query, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            )
        ).order_by("exact_match_priority", "-last_contacted_at", "-updated_at", "-created_at")
    else:
        queryset = queryset.order_by("-last_contacted_at", "-updated_at", "-created_at")

    return queryset.distinct()[:limit]


def _save_recovered_interested_detail(staff, lead, interested_detail):
    if not interested_detail:
        return
    customer_name = (interested_detail.get("customer_name") or "").strip()
    customer_phone = (interested_detail.get("customer_phone") or "").strip()
    product_enquired = (interested_detail.get("product_enquired") or "").strip()
    preferred_call_time = (interested_detail.get("preferred_call_time") or "").strip()
    if not all([customer_name, customer_phone, product_enquired, preferred_call_time]):
        return

    latest_call = (
        Call.objects.filter(staff=staff, lead=lead).order_by("-start_time", "-created_at").first()
    )
    InterestedLeadDetail.objects.update_or_create(
        lead=lead,
        defaults={
            "staff": staff,
            "call": latest_call,
            "customer_name": customer_name,
            "customer_phone": customer_phone,
            "product_enquired": product_enquired,
            "enquiry_notes": (interested_detail.get("enquiry_notes") or "").strip(),
            "preferred_call_time": preferred_call_time,
        },
    )


def recover_staff_customer_lead(
    staff,
    lead,
    *,
    status,
    callback_window="",
    callback_date=None,
    interested_detail=None,
):
    has_staff_history = lead.assigned_to_id == staff.id or Call.objects.filter(staff=staff, lead=lead).exists()
    if not has_staff_history:
        raise PermissionError("This customer is not in your calling history.")

    if lead.status == Lead.Status.CONVERTED:
        raise ValueError("This customer is already marked as converted.")

    if lead.assigned_to_id and lead.assigned_to_id != staff.id and _is_active_queue_status(lead.status):
        raise ValueError("This customer is already active in another staff queue.")

    status = _normalize_followup_status(status)
    now = timezone.now()
    session = get_open_session(staff, reconcile=True)
    if session:
        _touch_session_interaction(session, now, metadata={"source": "customer_recovery"})

    resolved_callback_window = callback_window if status == Lead.Status.INTERESTED else ""
    resolved_callback_date = callback_date if status == Lead.Status.INTERESTED else None
    previous_owner = lead.assigned_to.name if lead.assigned_to and lead.assigned_to_id != staff.id else ""

    lead.assigned_to = staff
    lead.status = status
    lead.callback_window = resolved_callback_window
    lead.callback_date = resolved_callback_date
    lead.last_contacted_at = now
    lead.save(
        update_fields=[
            "assigned_to",
            "status",
            "callback_window",
            "callback_date",
            "last_contacted_at",
            "updated_at",
        ]
    )

    mark_staff_seen(staff, now)
    _log_staff_action(
        staff,
        StaffAction.ActionType.CALL_STATUS_UPDATED,
        session=session,
        lead=lead,
        app_state=session.last_known_state if session else None,
        metadata={
            "source": "customer_recovery",
            "status": status,
            "callback_window": resolved_callback_window,
            "callback_date": resolved_callback_date.isoformat() if resolved_callback_date else "",
            "previous_owner": previous_owner,
        },
    )

    if status == Lead.Status.INTERESTED:
        _save_recovered_interested_detail(staff, lead, interested_detail)

    auto_allocate_leads(target_staff=staff, prioritized_lead_ids=[lead.id])
    return lead


def start_staff_call(staff, lead):
    now = timezone.now()
    _reconcile_open_calls(staff=staff, now=now)
    for lingering_call in (
        Call.objects.filter(staff=staff, end_time__isnull=True)
        .select_related("staff", "lead")
        .order_by("-start_time", "-created_at")
    ):
        _close_unresolved_call(lingering_call, reason="replaced_by_new_call")

    session = get_open_session(staff, reconcile=True)
    if session:
        _touch_session_interaction(session, now, metadata={"source": "call_start"})
    else:
        session, _ = start_staff_session(staff, source="auto_call_start")

    call = Call.objects.create(
        staff=staff,
        lead=lead,
        start_time=now,
        status=Call.Status.STARTED,
    )
    lead.last_contacted_at = now
    lead.save(update_fields=["last_contacted_at", "updated_at"])
    mark_staff_seen(staff, now)
    _log_staff_action(
        staff,
        StaffAction.ActionType.CALL_STARTED,
        session=session,
        call=call,
        lead=lead,
        app_state=session.last_known_state if session else None,
        metadata={"lead_name": lead.name},
    )
    return call


def retry_pending_staff_call(call):
    if call.end_time is None or call.status != Call.Status.STARTED:
        return call

    now = timezone.now()
    session = get_open_session(call.staff, reconcile=True)
    if session:
        _touch_session_interaction(session, now, metadata={"source": "call_retry"})
    else:
        session, _ = start_staff_session(call.staff, source="auto_call_retry")

    call.start_time = now
    call.end_time = None
    call.duration_seconds = 0
    call.is_qualifying = False
    call.callback_window = ""
    call.callback_date = None
    call.is_verified = False
    call.verification_source = ""
    call.auto_skipped_sync_issue = False
    call.sync_skip_reason = ""
    call.save(
        update_fields=[
            "start_time",
            "end_time",
            "duration_seconds",
            "is_qualifying",
            "callback_window",
            "callback_date",
            "is_verified",
            "verification_source",
            "auto_skipped_sync_issue",
            "sync_skip_reason",
            "updated_at",
        ]
    )

    call.lead.last_contacted_at = now
    call.lead.save(update_fields=["last_contacted_at", "updated_at"])
    mark_staff_seen(call.staff, now)
    _log_staff_action(
        call.staff,
        StaffAction.ActionType.CALL_STARTED,
        session=session,
        call=call,
        lead=call.lead,
        app_state=session.last_known_state if session else None,
        metadata={"lead_name": call.lead.name, "source": "retry_pending"},
    )
    return call


def end_staff_call(
    call,
    status=None,
    *,
    duration_seconds=None,
    ended_at=None,
    source="app",
    callback_window="",
    callback_date=None,
):
    sync_skip_reason = ""
    auto_skipped_sync_issue = False
    if source == "sync_issue_no_log_access_skip":
        sync_skip_reason = Call.SyncSkipReason.NO_LOG_ACCESS
        auto_skipped_sync_issue = True
    elif source == "sync_issue_no_log_match_skip":
        sync_skip_reason = Call.SyncSkipReason.NO_MATCHING_LOG
        auto_skipped_sync_issue = True
    elif source == "sync_issue_read_error_skip":
        sync_skip_reason = Call.SyncSkipReason.LOG_READ_FAILED
        auto_skipped_sync_issue = True

    force_sync_issue_resolution = (
        bool(call.end_time)
        and call.status == Call.Status.STARTED
        and source
        in {
            "sync_issue_no_log_access_skip",
            "sync_issue_no_log_match_skip",
            "sync_issue_read_error_skip",
            "manual_sync_escape",
        }
    )
    if call.end_time and not force_sync_issue_resolution:
        return call

    now = timezone.now()
    session = get_open_session(call.staff, reconcile=True)
    already_ended = bool(call.end_time)
    is_verified = (
        duration_seconds is not None
        and ended_at is not None
        and source in VERIFIED_CALL_SOURCES
    )

    if force_sync_issue_resolution:
        resolved_end_time = call.end_time or ended_at or now
        if resolved_end_time < call.start_time:
            resolved_end_time = call.start_time
        resolved_duration = max(0, int(call.duration_seconds or 0))
    elif is_verified:
        resolved_end_time = min(ended_at, now + timedelta(seconds=VERIFIED_CALL_TIME_SKEW_SECONDS))
        if resolved_end_time < call.start_time:
            resolved_end_time = call.start_time
        max_verified_duration = max(
            0,
            int((resolved_end_time - call.start_time).total_seconds()) + VERIFIED_CALL_TIME_SKEW_SECONDS,
        )
        resolved_duration = min(max(0, int(duration_seconds)), max_verified_duration)
        call.start_time = resolved_end_time - timedelta(seconds=resolved_duration)
    else:
        resolved_end_time = ended_at or now
        if resolved_end_time < call.start_time:
            resolved_end_time = call.start_time
        resolved_duration = 0

    requested_status = status
    normalized_requested_status = requested_status
    if force_sync_issue_resolution and not normalized_requested_status:
        normalized_requested_status = Call.Status.NO_ANSWER
    if requested_status == Call.Status.NOT_INTERESTED and (
        not is_verified or resolved_duration <= 0
    ):
        normalized_requested_status = Call.Status.NO_ANSWER
    call.end_time = resolved_end_time
    call.duration_seconds = resolved_duration
    call.is_qualifying = call.duration_seconds >= SHORT_CALL_SECONDS
    call.is_verified = is_verified
    call.verification_source = source if is_verified else ""
    call.auto_skipped_sync_issue = auto_skipped_sync_issue
    call.sync_skip_reason = sync_skip_reason
    if session and not already_ended:
        _credit_call_duration_to_session(
            session,
            resolved_end_time,
            resolved_duration,
            metadata={
                "source": "call_end",
                "duration_seconds": resolved_duration,
            },
            mark_verified=is_verified,
        )
    if source == "call_log_short_recall":
        call.is_qualifying = False
        call.status = Call.Status.INVALID_SHORT
    elif not call.is_qualifying and normalized_requested_status in {
        Call.Status.NO_ANSWER,
        Call.Status.NOT_INTERESTED,
    }:
        call.status = Call.Status.NO_ANSWER
        if normalized_requested_status == Call.Status.NOT_INTERESTED:
            call.status = Call.Status.NOT_INTERESTED
    else:
        call.status = Call.Status.INVALID_SHORT if not call.is_qualifying else call.status
    call.save(
        update_fields=[
            "start_time",
            "end_time",
            "duration_seconds",
            "is_qualifying",
            "is_verified",
            "verification_source",
            "auto_skipped_sync_issue",
            "sync_skip_reason",
            "status",
            "updated_at",
        ]
    )

    call.lead.last_contacted_at = call.end_time
    call.lead.save(update_fields=["last_contacted_at", "updated_at"])
    if call.status == Call.Status.INVALID_SHORT:
        if source == "call_log_short_recall":
            _return_lead_to_same_staff_after_invalid_short(call.lead, call.staff)
            auto_allocate_leads(target_staff=call.staff, prioritized_lead_ids=[call.lead_id])
        else:
            _return_lead_to_queue_after_invalid_short(call.lead)
            auto_allocate_leads()

    _log_staff_action(
        call.staff,
        StaffAction.ActionType.CALL_ENDED,
        session=session,
        call=call,
        lead=call.lead,
        app_state=session.last_known_state if session else None,
        metadata={
            "duration_seconds": call.duration_seconds,
            "is_qualifying": call.is_qualifying,
            "is_verified": call.is_verified,
            "status": call.status,
            "source": source,
            "auto_skipped_sync_issue": call.auto_skipped_sync_issue,
            "sync_skip_reason": call.sync_skip_reason,
            "ended_at": call.end_time.isoformat(),
        },
    )

    if normalized_requested_status and (
        call.is_qualifying
        or normalized_requested_status in {Call.Status.NO_ANSWER, Call.Status.NOT_INTERESTED}
    ):
        update_staff_call_status(call, normalized_requested_status, callback_window, callback_date)
    return call


def update_staff_call_status(call, status, callback_window="", callback_date=None):
    if call.status == Call.Status.INVALID_SHORT:
        return call

    status = _normalize_followup_status(status)
    now = timezone.now()
    is_followup_lead = _is_followup_status(call.lead.status)
    followup_progress = None
    if is_followup_lead and status in {Call.Status.NO_ANSWER, Call.Status.NOT_INTERESTED}:
        followup_progress = _followup_no_response_progress(
            call.lead,
            staff=call.staff,
            exclude_call_id=call.id,
            include_attempt_at=call.end_time or now,
        )

    if not call.is_verified or int(call.duration_seconds or 0) <= 0:
        if status == Call.Status.NOT_INTERESTED:
            # Allow direct rejection for follow-up only after 3 unanswered tries.
            if not (is_followup_lead and followup_progress and followup_progress["can_close"]):
                status = Call.Status.NO_ANSWER
        elif status != Call.Status.NO_ANSWER:
            raise ValueError(
                "This call could not be verified from the phone log. Sync it again or mark it as No Response."
            )

    session = get_open_session(call.staff, reconcile=True)
    if session:
        _touch_session_interaction(session, now, metadata={"source": "call_status"})

    resolved_callback_window = callback_window if status == Call.Status.INTERESTED else ""
    resolved_callback_date = callback_date if status == Call.Status.INTERESTED else None
    lead_status = status
    lead_callback_window = resolved_callback_window
    lead_callback_date = resolved_callback_date
    if is_followup_lead and status == Call.Status.NO_ANSWER:
        followup_progress = followup_progress or _followup_no_response_progress(
            call.lead,
            staff=call.staff,
            exclude_call_id=call.id,
            include_attempt_at=call.end_time or now,
        )
        if followup_progress["can_close"]:
            lead_status = Lead.Status.NO_ANSWER
        else:
            lead_status = Lead.Status.INTERESTED
            lead_callback_window = call.lead.callback_window
            lead_callback_date = call.lead.callback_date
    update_fields = ["status", "callback_window", "callback_date", "updated_at"]
    if call.end_time is None:
        call.end_time = now
        call.duration_seconds = 0
        call.is_qualifying = False
        call.is_verified = False
        call.verification_source = ""
        update_fields = [
            "end_time",
            "duration_seconds",
            "is_qualifying",
            "is_verified",
            "verification_source",
            *update_fields,
        ]
    call.status = status
    call.callback_window = resolved_callback_window
    call.callback_date = resolved_callback_date
    call.save(update_fields=update_fields)
    call.lead.status = lead_status
    if lead_status == Lead.Status.INTERESTED and not call.lead.loan_stage:
        call.lead.loan_stage = Lead.LoanStage.OFFICE_REVIEW
    elif lead_status == Lead.Status.CONVERTED:
        call.lead.loan_stage = Lead.LoanStage.SUCCESSFUL
    elif lead_status in {Lead.Status.NOT_INTERESTED, Lead.Status.NO_ANSWER}:
        call.lead.loan_stage = Lead.LoanStage.UNSUCCESSFUL
    call.lead.callback_window = lead_callback_window
    call.lead.callback_date = lead_callback_date
    call.lead.last_contacted_at = call.end_time or now
    call.lead.save(
        update_fields=[
            "status",
            "loan_stage",
            "callback_window",
            "callback_date",
            "last_contacted_at",
            "updated_at",
        ]
    )

    _log_staff_action(
        call.staff,
        StaffAction.ActionType.CALL_STATUS_UPDATED,
        session=session,
        call=call,
        lead=call.lead,
        app_state=session.last_known_state if session else None,
        metadata={
            "status": status,
            "callback_window": resolved_callback_window,
            "callback_date": resolved_callback_date.isoformat() if resolved_callback_date else "",
            "lead_status": lead_status,
        },
    )
    if lead_status == Lead.Status.INTERESTED:
        auto_allocate_leads(target_staff=call.staff)
    elif lead_status in TERMINAL_QUEUE_STATUSES:
        auto_allocate_leads(target_staff=call.staff)
    return call


def ensure_converted_call_credit(lead):
    if lead.status != Lead.Status.CONVERTED:
        return None

    existing_converted_call = (
        Call.objects.filter(
            lead=lead,
            status=Call.Status.CONVERTED,
            is_qualifying=True,
        )
        .select_related("staff", "lead")
        .order_by("-end_time", "-start_time", "-created_at")
        .first()
    )
    if existing_converted_call:
        return existing_converted_call

    detail = getattr(lead, "interested_detail", None)
    preferred_call = detail.call if detail and detail.call_id else None
    if preferred_call and preferred_call.lead_id != lead.id:
        preferred_call = None

    candidate_call = preferred_call or (
        Call.objects.filter(lead=lead)
        .select_related("staff", "lead")
        .order_by("-end_time", "-start_time", "-created_at")
        .first()
    )
    if candidate_call is None:
        return None

    update_fields = []
    if candidate_call.status != Call.Status.CONVERTED:
        candidate_call.status = Call.Status.CONVERTED
        update_fields.append("status")
    if not candidate_call.is_qualifying:
        candidate_call.is_qualifying = True
        update_fields.append("is_qualifying")
    if int(candidate_call.duration_seconds or 0) < SHORT_CALL_SECONDS:
        candidate_call.duration_seconds = SHORT_CALL_SECONDS
        update_fields.append("duration_seconds")
    if candidate_call.end_time is None:
        candidate_call.end_time = timezone.now()
        update_fields.append("end_time")
    if not candidate_call.verification_source:
        candidate_call.verification_source = "admin_conversion"
        update_fields.append("verification_source")

    if update_fields:
        candidate_call.save(update_fields=[*update_fields, "updated_at"])
    return candidate_call


def read_root_file(filename):
    return (settings.BASE_DIR / filename).read_text(encoding="utf-8")

